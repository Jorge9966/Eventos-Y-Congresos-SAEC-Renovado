from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User, Group
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.urls import reverse
from django.http import HttpResponseRedirect
from django.core.mail import send_mail
from django.conf import settings
from django.core.files.uploadedfile import UploadedFile
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
import random
import re
from datetime import timedelta
from django.utils import timezone
from django.db import transaction
from functools import wraps
from django.utils.text import slugify
from django.db.models import Q, Count
from .models import (
    PasswordResetCode,
    Congreso,
    AllowedEmailDomain,
    PerformanceLevel,
    UserCongresoMembership,
    Taller,
    Conferencia,
    Concurso,
    TallerInscripcion,
    ConcursoInscripcion,
    ConferenciaInscripcion,
    ExtraField,
    UserExtraFieldValue,
    CongresoAdminScope,
    ConcursoEquipo,
    ConcursoEquipoMiembro,
)

# -------------------
# Helpers de acceso por congreso
# -------------------
def _user_is_congreso_admin(user, congreso: Congreso) -> bool:
    """Determina si el usuario tiene privilegios administrativos en este congreso.

    Reglas:
    - Superuser: siempre sí.
    - admin_user del congreso: sí.
    - Grupo "Administrador": global sí.
    - Grupo "Administradores_Congresos": solo sí si:
        * El congreso tiene restrict_group_admin_scope=False (acceso a todos) OR
        * Existe un CongresoAdminScope que vincula al usuario con este congreso.
    """
    if not user.is_authenticated:
        return False
    if user.is_superuser or congreso.admin_user_id == user.id or user.groups.filter(name="Administrador").exists():
        return True
    if user.groups.filter(name="Administradores_Congresos").exists():
        # Si el congreso no restringe el scope, permitir.
        if not congreso.restrict_group_admin_scope:
            return True
        scope = getattr(user, "admin_congreso_scope", None)
        if scope and scope.congreso_id == congreso.id:
            return True
    return False


def _ensure_congreso_access_or_redirect(request, congreso: Congreso, allow_instructors: bool = False):
    """Devuelve None si el usuario tiene acceso. En caso contrario, agrega un mensaje y devuelve un redirect.

    Reglas:
    - Admin del congreso, superuser y grupo Administradores_Congresos: acceso total.
    - Resto: requiere una membresía APPROVED en UserCongresoMembership para ese congreso.
    """
    user = request.user
    # Bloqueo global: Instructores y Participantes no pueden acceder a vistas protegidas
    # excepto cuando explícitamente se permita a Instructores (p.ej., impresión de listas).
    if user.is_authenticated and not user.is_superuser and not user.groups.filter(name__in=["Administrador", "Administradores_Congresos"]).exists():
        if user.groups.filter(name__in=["Instructores", "Participantes"]).exists():
            if user.groups.filter(name="Instructores").exists() and allow_instructors:
                pass
            else:
                messages.error(request, "No tienes acceso a esta sección.")
                return redirect("login")
    if _user_is_congreso_admin(user, congreso):
        return None
    # Si el usuario está en Administradores_Congresos pero con scope diferente, bloquear
    if user.is_authenticated and user.groups.filter(name="Administradores_Congresos").exists():
        scope = getattr(user, "admin_congreso_scope", None)
        if scope and scope.congreso_id != congreso.id:
            messages.error(request, "No tienes permiso para acceder a este congreso.")
            return redirect("inicio")
    membership = (
        UserCongresoMembership.objects.filter(user=user, congreso=congreso)
        .order_by("-created_at")
        .first()
    )
    if not membership:
        messages.warning(request, "Necesitas solicitar acceso a este congreso para continuar.")
        return redirect("congresos_eventos")
    if membership.status == "pending":
        messages.info(request, "Tu solicitud de acceso a este congreso está pendiente de aprobación.")
        return redirect("congresos_eventos")
    if membership.status == "rejected":
        messages.error(request, "Tu solicitud de acceso a este congreso fue rechazada.")
        return redirect("congresos_eventos")
    return None
def _get_scoped_congreso_for_group_admin(user):
    """Obtiene el congreso asignado a un usuario del grupo
    "Administradores_Congresos". Si no existe el registro de scope pero el
    usuario figura como admin_user de algún congreso, se intenta crearlo.
    """
    if not getattr(user, "is_authenticated", False):
        return None
    if not user.groups.filter(name="Administradores_Congresos").exists():
        return None
    scope = getattr(user, "admin_congreso_scope", None)
    if scope:
        return scope.congreso
    congreso = Congreso.objects.filter(admin_user=user).first()
    if congreso:
        try:
            CongresoAdminScope.objects.get_or_create(user=user, defaults={"congreso": congreso})
        except Exception:
            pass
        return congreso
    return None

# -------------------
# Decoradores de acceso por rol
# -------------------
def instructors_area_only(view_func):
    """Restringe el acceso a vistas del área de instructores.

    Permite: superuser o usuarios en el grupo "Instructores".
    Bloquea: grupo "Participantes" y cualquier otro usuario no instructor.
    """
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        user = request.user
        if not getattr(user, "is_authenticated", False):
            # Conservar comportamiento estándar de login_required
            return redirect("login")
        if user.is_superuser or user.groups.filter(name__in=["Instructores"]).exists():
            return view_func(request, *args, **kwargs)
        # Mensaje claro para participantes y otros roles
        messages.error(request, "Esta sección es solo para instructores.")
        return redirect("inicio")
    return _wrapped

# -------------------
# GUARD: BLOQUEAR ROLES NO ADMIN EN VISTAS PROTEGIDAS
# -------------------
def _deny_non_admin_roles(request):
    """Retorna un redirect si el usuario pertenece a Instructores/Participantes y no es admin/superuser."""
    user = request.user
    if getattr(user, "is_authenticated", False):
        if user.is_superuser:
            return None
        if user.groups.filter(name__in=["Administrador", "Administradores_Congresos"]).exists():
            return None
        if user.groups.filter(name__in=["Instructores", "Participantes"]).exists():
            # Permitir sus páginas de inicio específicas
            allowed_paths = {
                reverse("inicio_instructor"),
                reverse("inicio_participante"),
            }
            # También permitir home público, login, registro y recuperación
            allowed_names = {"home", "login", "register", "password_reset", "code_password"}
            current_path = request.path
            current_name = None
            try:
                # Resolver nombre de la ruta actual para comparar
                current_name = request.resolver_match.view_name if request.resolver_match else None
            except Exception:
                current_name = None
            if current_path in allowed_paths or current_name in allowed_names:
                return None
            messages.error(request, "No tienes acceso a esta sección.")
            return redirect("login")
    return None
def home_congresos(request):
    """Página pública de inicio: lista todos los congresos/eventos y redirige a login al hacer clic."""
    congresos = list(Congreso.objects.all().order_by("name"))
    return render(request, "home.html", {"congresos": congresos})


# -------------------
# LOGIN
# -------------------
def login_view(request):
    # Obtener congreso seleccionado por query (?c=ID) o por campo oculto del formulario
    congreso_id = request.GET.get("c") or request.POST.get("c")
    selected_congreso = None
    if congreso_id:
        try:
            selected_congreso = Congreso.objects.get(pk=congreso_id)
        except Congreso.DoesNotExist:
            selected_congreso = None

    if request.method == "POST":
        # Permitir iniciar sesión con usuario o correo
        identifier = (request.POST.get("username") or "").strip()
        password = request.POST.get("password")

        username_for_auth = identifier
        if identifier and "@" in identifier:
            # 1) Si existe un usuario con username==email, úsalo (es el caso común en este proyecto)
            match_username = User.objects.filter(username__iexact=identifier).values_list("username", flat=True).first()
            if match_username:
                username_for_auth = match_username
            else:
                # 2) Si hay múltiples cuentas con el mismo email, elegir una estable (último login más reciente, si no, id mayor)
                user_by_email = (
                    User.objects.filter(email__iexact=identifier)
                    .order_by("-last_login", "-id")
                    .first()
                )
                if user_by_email:
                    username_for_auth = user_by_email.username

        user = authenticate(request, username=(username_for_auth or identifier), password=password)

        if user is not None:
            # Política: si es Administradores_Congresos con scope y el usuario eligió un congreso distinto, DENEGAR login
            if user.groups.filter(name="Administradores_Congresos").exists():
                scoped = _get_scoped_congreso_for_group_admin(user)
                if scoped:
                    if selected_congreso and scoped.id != selected_congreso.id:
                        messages.error(
                            request,
                            f"Debes iniciar sesión en tu congreso asignado: '{scoped.name}'."
                        )
                        # No iniciar sesión; permanecer en login
                        return render(request, "login.html", {"selected_congreso": selected_congreso})
                    # Si no eligió congreso o coincide, iniciar sesión y enviar a su congreso
                    login(request, user)
                    return redirect("inicio_congreso", scoped.id)
            # Lógica específica para Instructores: sólo permiten login si el congreso seleccionado está aprobado
            if user.groups.filter(name="Instructores").exists():
                approved_qs = UserCongresoMembership.objects.filter(user=user, role="instructor", status="approved")
                # Si no tiene ningún congreso aprobado, negar el login (similar a scope estricto)
                if not approved_qs.exists():
                    messages.error(request, "Tu cuenta de instructor no tiene ningún congreso aprobado todavía.")
                    return render(request, "login.html", {"selected_congreso": selected_congreso})
                # Si viene congreso seleccionado, debe estar aprobado para ese congreso
                chosen_id = None
                if selected_congreso:
                    if not approved_qs.filter(congreso=selected_congreso).exists():
                        messages.error(request, f"Debes iniciar sesión en un congreso donde hayas sido aceptado.")
                        return render(request, "login.html", {"selected_congreso": selected_congreso})
                    chosen_id = selected_congreso.id
                else:
                    # Si no eligió, usar el más reciente aprobado
                    chosen = approved_qs.order_by("-decided_at", "-id").first()
                    chosen_id = chosen.congreso_id if chosen else None

                # Ahora sí iniciar sesión y fijar congreso activo
                login(request, user)
                if chosen_id:
                    request.session["instructor_active_congreso_id"] = chosen_id
                return redirect("inicio_instructor")

            # Participantes: permitir login aunque aún no tengan un congreso aprobado.
            # Si hay aprobado, fijar congreso activo; si no, ir al inicio de participante con botones deshabilitados.
            if user.groups.filter(name="Participantes").exists():
                approved_qs = UserCongresoMembership.objects.filter(user=user, role="participante", status="approved")
                chosen_id = None
                if approved_qs.exists():
                    # Determinar congreso activo (seleccionado aprobado o más reciente)
                    if selected_congreso:
                        if not approved_qs.filter(congreso=selected_congreso).exists():
                            messages.error(request, "Debes iniciar sesión en un congreso donde hayas sido aceptado como participante.")
                            return render(request, "login.html", {"selected_congreso": selected_congreso})
                        chosen_id = selected_congreso.id
                    else:
                        chosen = approved_qs.order_by("-decided_at", "-id").first()
                        chosen_id = chosen.congreso_id if chosen else None
                # Login válido y (opcionalmente) fijar sesión
                login(request, user)
                if chosen_id:
                    request.session["participant_active_congreso_id"] = chosen_id
                return redirect("inicio_participante")

            # Resto de roles: login normal
            login(request, user)
            return redirect("inicio")  # por defecto administradores
            return redirect("inicio")  # por defecto administradores
        else:
            messages.error(request, "Usuario o contraseña incorrectos.")
        return render(request, "login.html", {"selected_congreso": selected_congreso})

    return render(request, "login.html", {"selected_congreso": selected_congreso})


# -------------------
# REGISTRO
# -------------------
def registro_view(request):
    # Congreso seleccionado para mostrar logo en registro
    congreso_id = request.GET.get("c") or request.POST.get("c")
    selected_congreso = None
    if congreso_id:
        try:
            selected_congreso = Congreso.objects.get(pk=congreso_id)
        except Congreso.DoesNotExist:
            selected_congreso = None

    levels = []
    if selected_congreso:
        levels = list(PerformanceLevel.objects.filter(congreso=selected_congreso).order_by("name"))

    if request.method == "POST":
        first_name = request.POST.get("first_name")
        apellido_paterno = request.POST.get("apellido_paterno")
        apellido_materno = request.POST.get("apellido_materno")
        email = request.POST.get("email")
        password = request.POST.get("password")
        tipo_usuario = request.POST.get("tipo_usuario")
        # Nueva contraseña de acceso de congreso (sustituye clave de instructor)
        congreso_password = request.POST.get("congreso_password")
        performance_level_id = request.POST.get("performance_level_id")
        # Campos extra dinámicos activos para el congreso
        extra_fields = []
        if selected_congreso:
            extra_fields = list(
                ExtraField.objects.filter(
                    (Q(congreso=selected_congreso) | Q(congreso__isnull=True)),
                    active=True,
                    section="registro",
                ).order_by("order", "name")
            )

        # Validar formato de correo (username y email deben ser correos válidos)
        try:
            validate_email(email or "")
        except ValidationError:
            messages.error(request, "Ingresa un correo electrónico válido.")
            return render(request, "register.html", {"selected_congreso": selected_congreso, "levels": levels})



        # Unicidad por congreso: si ya existe un usuario con este correo vinculado a este congreso, bloquear
        if selected_congreso:
            existing_user_for_email = (
                User.objects.filter(Q(username__iexact=email) | Q(email__iexact=email))
                .order_by("-last_login", "-id")
                .first()
            )
            if existing_user_for_email:
                if UserCongresoMembership.objects.filter(user=existing_user_for_email, congreso=selected_congreso).exists():
                    messages.error(request, "Este correo ya está registrado en este congreso. Inicia sesión para continuar.")
                    return render(request, "register.html", {"selected_congreso": selected_congreso, "levels": levels})

        # Validar contraseña de acceso del congreso (si aplica)
        if selected_congreso and selected_congreso.has_access_password():
            if not congreso_password or not selected_congreso.check_access_password(congreso_password):
                messages.error(request, "La contraseña de acceso del congreso es incorrecta.")
                return render(request, "register.html", {"selected_congreso": selected_congreso, "levels": levels})

        # Validar dominio permitido por congreso (si hay dominios configurados)
        if selected_congreso:
            dominios = list(AllowedEmailDomain.objects.filter(congreso=selected_congreso).values_list("domain", flat=True))
            if dominios:
                email_lc = (email or "").lower().strip()
                is_allowed = any(email_lc.endswith(d) for d in dominios)
                if not is_allowed:
                    pretty = ", ".join(sorted(dominios))
                    messages.error(request, f"Este correo no pertenece a un dominio permitido para este congreso. Dominios permitidos: {pretty}")
                    return render(request, "register.html", {"selected_congreso": selected_congreso, "levels": levels})

        # Validación de política de contraseña: 8-16 caracteres, al menos 1 mayúscula, 1 minúscula y 1 símbolo especial (.!%&$})
        specials = set(".!%&$}")
        if not password or len(password) < 8 or len(password) > 16:
            messages.error(request, "La contraseña debe tener entre 8 y 16 caracteres.")
            return render(request, "register.html", {"selected_congreso": selected_congreso, "levels": levels})
        if not any(c.isupper() for c in password) or not any(c.islower() for c in password) or not any(c in specials for c in password):
            messages.error(request, "La contraseña debe contener al menos una letra mayúscula, una letra minúscula y un símbolo especial (.!%&$}).")
            return render(request, "register.html", {"selected_congreso": selected_congreso, "levels": levels})

        # Validar campos extra antes de crear usuario (solo para el rol elegido)
        extra_values = {}
        for f in extra_fields:
            if f.role_scope == "participante" and tipo_usuario != "participante":
                continue
            if f.role_scope == "instructor" and tipo_usuario != "instructor":
                continue
            key = f"extra_{f.id}"
            raw_val = (request.POST.get(key) or "").strip()
            if f.required and not raw_val:
                messages.error(request, f"El campo '{f.name}' es obligatorio.")
                return render(request, "register.html", {"selected_congreso": selected_congreso, "levels": levels})
            if raw_val:
                if f.field_type == "email":
                    try:
                        validate_email(raw_val)
                    except ValidationError:
                        messages.error(request, f"Ingresa un correo válido en el campo '{f.name}'.")
                        return render(request, "register.html", {"selected_congreso": selected_congreso, "levels": levels})
                elif f.field_type == "number":
                    try:
                        float(raw_val)
                    except ValueError:
                        messages.error(request, f"Ingresa un número válido en el campo '{f.name}'.")
                        return render(request, "register.html", {"selected_congreso": selected_congreso, "levels": levels})
                elif f.field_type == "select":
                    choices = f.get_choices()
                    if choices and raw_val not in choices:
                        messages.error(request, f"Selecciona una opción válida para '{f.name}'.")
                        return render(request, "register.html", {"selected_congreso": selected_congreso, "levels": levels})
                # boolean y date no requieren validación extra aquí (navegador ayuda)
            extra_values[f.id] = raw_val

        # Encontrar o crear usuario (permitir "usuarios únicos por congreso")
        user = None
        existing = (
            User.objects.filter(Q(username__iexact=email) | Q(email__iexact=email))
            .order_by("-last_login", "-id")
            .first()
        )
        if existing:
            # Si ya existe, exigir que la contraseña ingresada coincida para asociar nueva membresía
            user = authenticate(request, username=existing.username, password=password)
            if not user:
                messages.error(request, "Ya existe una cuenta con ese correo. Inicia sesión o verifica tu contraseña.")
                return render(request, "register.html", {"selected_congreso": selected_congreso, "levels": levels})
            # Actualizar nombre si viene vacío en la cuenta
            if first_name and not user.first_name:
                user.first_name = first_name
            full_last = f"{apellido_paterno} {apellido_materno}".strip()
            if full_last and not user.last_name:
                user.last_name = full_last
            # Sincronizar username/email con el correo ingresado si es diferente y ya pasó las validaciones
            if email:
                # Evitar colisión con otros usuarios para username
                if user.username.lower() != email.lower():
                    if User.objects.exclude(pk=user.pk).filter(username__iexact=email).exists():
                        messages.error(request, "Ese correo ya está en uso por otra cuenta. Inicia sesión con ese correo o usa uno distinto.")
                        return render(request, "register.html", {"selected_congreso": selected_congreso, "levels": levels})
                    user.username = email
                if user.email.lower() != email.lower():
                    user.email = email
            user.save()
        else:
            # Crear usuario nuevo: forzar que username sea el correo (formato obligatorio)
            # Ya pasó validaciones de dominio si aplica.
            user = User.objects.create_user(
                username=email,
                email=email,
                password=password,
                first_name=first_name,
                last_name=f"{apellido_paterno} {apellido_materno}"
            )

        # Asignar grupo según tipo
        if tipo_usuario == "instructor":
            grupo = Group.objects.get(name="Instructores")
            user.groups.add(grupo)
        else:
            grupo = Group.objects.get(name="Participantes")
            user.groups.add(grupo)

        # Asociar solicitud de membresía al congreso (pendiente de aprobación)
        if selected_congreso:
            role_value = "instructor" if tipo_usuario == "instructor" else "participante"
            # Validar nivel de desempeño si es participante
            selected_level = None
            if role_value == "participante" and performance_level_id:
                try:
                    selected_level = PerformanceLevel.objects.get(id=performance_level_id, congreso=selected_congreso)
                except PerformanceLevel.DoesNotExist:
                    selected_level = None
            if role_value == "participante" and PerformanceLevel.objects.filter(congreso=selected_congreso).exists() and not selected_level:
                messages.error(request, "Selecciona un nivel de desempeño válido para este congreso.")
                return render(request, "register.html", {"selected_congreso": selected_congreso, "levels": levels})
            try:
                UserCongresoMembership.objects.create(
                    user=user,
                    congreso=selected_congreso,
                    role=role_value,
                    status="pending",
                    performance_level=selected_level,
                )
            except Exception:
                # Evitar que una excepción silencie el registro; el admin podrá corregir desde el panel
                pass

        # Tras el registro, redirigir al login con aviso
        # Guardar valores extra si aplica
        if selected_congreso and user and extra_values:
            for fid, val in extra_values.items():
                try:
                    fobj = next((x for x in extra_fields if x.id == fid), None)
                    if not fobj:
                        continue
                    UserExtraFieldValue.objects.update_or_create(
                        user=user,
                        congreso=selected_congreso if fobj.congreso_id else None,
                        field=fobj,
                        defaults={"value": val},
                    )
                except Exception:
                    # No bloquear registro por errores en extras
                    pass

        messages.success(request, "Cuenta creada. Inicia sesión para continuar.")
        login_url = reverse("login")
        if selected_congreso:
            return redirect(f"{login_url}?c={selected_congreso.id}")
        return redirect(login_url)

    # GET: cargar campos extra activos (congreso + globales)
    extra_fields = []
    if selected_congreso:
        extra_fields = list(
            ExtraField.objects.filter((Q(congreso=selected_congreso) | Q(congreso__isnull=True)), active=True, section="registro")
            .order_by("order", "name")
        )
    return render(request, "register.html", {"selected_congreso": selected_congreso, "levels": levels, "extra_fields": extra_fields})


@login_required
def campos_view(request):
    # Redirigir a admins de congresos con scope al panel del congreso
    scoped = _get_scoped_congreso_for_group_admin(request.user)
    if scoped:
        return redirect("inicio_congreso", scoped.id)
    # Rol en sidebar
    grupos = request.user.groups.values_list("name", flat=True)
    if request.user.is_superuser or "Administrador" in grupos:
        rol = "Administrador"
    elif "Instructores" in grupos:
        rol = "Instructor"
    elif "Participantes" in grupos:
        rol = "Participante"
    else:
        rol = "Administrador"

    # Asegurar un congreso base
    if not Congreso.objects.exists():
        Congreso.objects.create(name="Congreso General")

    # Selección de congreso
    selected_id = request.POST.get("congreso_id") or request.GET.get("c")
    try:
        congreso = Congreso.objects.get(pk=selected_id) if selected_id else Congreso.objects.order_by("name").first()
    except Congreso.DoesNotExist:
        congreso = Congreso.objects.order_by("name").first()

    # Acciones
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "add_field":
            name = (request.POST.get("name") or "").strip()
            role_scope = (request.POST.get("role_scope") or "both").strip()
            required = request.POST.get("required") == "on"
            active = request.POST.get("active") == "on"
            unique_value = request.POST.get("unique_value") == "on"

            if not name:
                messages.error(request, "El nombre del campo no puede estar vacío.")
                return HttpResponseRedirect(reverse("campos") + f"?c={congreso.id}")
            if role_scope not in {"both", "participante", "instructor"}:
                role_scope = "both"

            # Generar code único por congreso
            base = slugify(name)[:200]
            code = base or f"campo-{ExtraField.objects.filter(congreso=congreso).count()+1}"
            suffix = 1
            while ExtraField.objects.filter(congreso=congreso, code=code).exists():
                suffix += 1
                code = f"{base}-{suffix}"[:220]

            # Calcular siguiente orden automáticamente
            last_order = (
                ExtraField.objects.filter(congreso=congreso)
                .order_by("-order")
                .values_list("order", flat=True)
                .first()
            ) or 0
            next_order = int(last_order) + 1

            ExtraField.objects.create(
                congreso=congreso,
                name=name,
                code=code,
                field_type="text",      # por defecto, no editable aquí
                role_scope=role_scope,
                section="registro",     # por defecto, no editable aquí
                required=required,
                active=active,
                unique_value=unique_value,
                order=next_order,
                choices_text="",
            )
            messages.success(request, "Campo agregado.")
            return HttpResponseRedirect(reverse("campos") + f"?c={congreso.id}")

        elif action == "delete_field":
            fid = request.POST.get("field_id")
            ExtraField.objects.filter(id=fid, congreso=congreso).delete()
            messages.success(request, "Campo eliminado.")
            return HttpResponseRedirect(reverse("campos") + f"?c={congreso.id}")

        elif action == "update_field":
            fid = request.POST.get("field_id")
            name = (request.POST.get("name") or "").strip()
            role_scope = (request.POST.get("role_scope") or "both").strip()
            required = request.POST.get("required") == "on"
            active = request.POST.get("active") == "on"
            unique_value = request.POST.get("unique_value") == "on"

            try:
                ef = ExtraField.objects.get(id=fid, congreso=congreso)
            except ExtraField.DoesNotExist:
                messages.error(request, "El campo a actualizar no existe.")
                return HttpResponseRedirect(reverse("campos") + f"?c={congreso.id}")

            if not name:
                messages.error(request, "El nombre del campo no puede estar vacío.")
                return HttpResponseRedirect(reverse("campos") + f"?c={congreso.id}&edit={ef.id}")
            if role_scope not in {"both", "participante", "instructor"}:
                role_scope = ef.role_scope

            # Solo actualizar atributos visibles en esta página
            ef.name = name
            ef.role_scope = role_scope
            ef.required = required
            ef.active = active
            ef.unique_value = unique_value
            ef.save(update_fields=["name", "role_scope", "required", "active", "unique_value"])
            messages.success(request, "Campo actualizado.")
            return HttpResponseRedirect(reverse("campos") + f"?c={congreso.id}")

    congresos = list(Congreso.objects.all().order_by("name"))
    fields = list(ExtraField.objects.filter(congreso=congreso).order_by("order", "name"))
    edit_id = request.GET.get("edit")
    edit_field = None
    if edit_id:
        try:
            edit_field = ExtraField.objects.get(id=edit_id, congreso=congreso)
        except ExtraField.DoesNotExist:
            edit_field = None

    return render(request, "campos.html", {"rol": rol, "congresos": congresos, "congreso": congreso, "fields": fields, "edit_field": edit_field})


# -------------------
# LOGOUT
# -------------------
def logout_view(request):
    logout(request)
    # Redirigir siempre a la página pública de inicio (lista de congresos)
    return redirect("home")


# -------------------
# VISTAS PROTEGIDAS
# -------------------
@login_required
def inicio_view(request):
    deny = _deny_non_admin_roles(request)
    if deny: return deny
    # Redirigir a admins de congresos acotados
    scoped = _get_scoped_congreso_for_group_admin(request.user)
    if scoped:
        return redirect("inicio_congreso", scoped.id)
    grupos = request.user.groups.values_list("name", flat=True)
    if "Instructores" in grupos:
        rol = "Instructor"
    elif "Participantes" in grupos:
        rol = "Participante"
    else:
        rol = "Administrador"

    return render(request, "inicio.html", {"rol": rol, "is_scoped_admin": bool(_get_scoped_congreso_for_group_admin(request.user))})

# -------------------
# LANDINGS POR ROL (INSTRUCTOR / PARTICIPANTE)
# -------------------
@login_required
@instructors_area_only
def inicio_instructor_view(request):
    # Solo instructores pueden acceder a esta vista
    if not (request.user.is_superuser or request.user.groups.filter(name__in=["Instructores"]).exists()):
        messages.error(request, "Esta página es solo para instructores.")
        return redirect("inicio")
    # Determinar el congreso activo desde sesión (si se fijó al iniciar sesión)
    active_id = request.session.get("instructor_active_congreso_id")
    membership = None
    if active_id:
        membership = UserCongresoMembership.objects.filter(
            user=request.user, role="instructor", status="approved", congreso_id=active_id
        ).first()
    # Si no se encontró por sesión, tomar el más reciente aprobado
    if not membership:
        membership = (
            UserCongresoMembership.objects
            .filter(user=request.user, role="instructor", status="approved")
            .order_by("-decided_at", "-id")
            .first()
        )
        # Ajustar sesión si encontramos uno
        if membership:
            request.session["instructor_active_congreso_id"] = membership.congreso_id
    # Si aún no hay membresía aprobada, redirigir a página de pendiente
    if not membership:
        return redirect("instructor_pendiente")
    congreso = membership.congreso if membership else None
    ctx = {"congreso": congreso}
    return render(request, "instructores_vista/inicio_instructor.html", ctx)


@login_required
@instructors_area_only
def instructor_pendiente_view(request):
    """Página mostrada a instructores que aún no tienen una membresía aprobada.

    Casos:
    - Sin solicitudes: invitar a solicitar acceso a un congreso.
    - Solicitudes pendientes: mostrar estado.
    - Solicitudes rechazadas (sin aprobadas): informar.
    Si eventualmente se aprueba una membresía, se redirige a la landing normal.
    """
    if not (request.user.is_superuser or request.user.groups.filter(name__in=["Instructores"]).exists()):
        messages.error(request, "Esta página es solo para instructores.")
        return redirect("inicio")
    approved_exists = UserCongresoMembership.objects.filter(user=request.user, role="instructor", status="approved").exists()
    if approved_exists:
        # Si se aprobó durante la sesión, ir directo a inicio
        return redirect("inicio_instructor")
    pending_exists = UserCongresoMembership.objects.filter(user=request.user, role="instructor", status="pending").exists()
    rejected_exists = UserCongresoMembership.objects.filter(user=request.user, role="instructor", status="rejected").exists()
    memberships = list(UserCongresoMembership.objects.filter(user=request.user, role="instructor").order_by("-created_at"))
    ctx = {
        "pending_exists": pending_exists,
        "rejected_exists": rejected_exists and not pending_exists,
        "memberships": memberships,
    }
    return render(request, "instructores_vista/instructor_pendiente.html", ctx)


@login_required
def inicio_participante_view(request):
    # Solo participantes (o superuser) pueden acceder a esta vista, incluso si su membresía está pendiente/rechazada.
    if not (request.user.is_superuser or request.user.groups.filter(name__in=["Participantes"]).exists()):
        messages.error(request, "Esta página es solo para participantes.")
        return redirect("inicio")

    # Tomar la membresía más reciente del usuario como participante para determinar estado.
    membership = (
        UserCongresoMembership.objects
        .filter(user=request.user, role="participante")
        .order_by("-decided_at", "-id")
        .first()
    )
    congreso = membership.congreso if membership else None
    status = membership.status if membership else "none"
    approved = status == "approved"
    # Etiqueta en español para el estado
    status_label_map = {
        "pending": "Pendiente de Activación",
        "approved": "Aprobado",
        "rejected": "Rechazado",
        "none": "",
    }
    status_label = status_label_map.get(status, "")

    # Inscripciones (mostrar solo la primera si hay varias, similar al mockup que lista un recurso)
    taller_insc = None
    concurso_insc = None
    conferencia_insc = None
    if approved and congreso:
        taller_insc = (
            TallerInscripcion.objects
            .filter(user=request.user, congreso=congreso)
            .select_related("taller")
            .order_by("-created_at")
            .first()
        )
        concurso_insc = (
            ConcursoInscripcion.objects
            .filter(user=request.user, congreso=congreso)
            .select_related("concurso")
            .order_by("-created_at")
            .first()
        )
        conferencia_insc = (
            ConferenciaInscripcion.objects
            .filter(user=request.user, congreso=congreso)
            .select_related("conferencia")
            .order_by("-created_at")
            .first()
        )

    # URLs para explorar (se deshabilitan en la plantilla si no aprobado)
    # Redirigen a las nuevas páginas de participante
    talleres_url = reverse("part_talleres")
    concursos_url = reverse("part_concursos")
    conferencias_url = reverse("part_conferencias")

    ctx = {
        "congreso": congreso,
        "membership": membership,
        "membership_status": status,
        "membership_status_label": status_label,
        "approved": approved,
        "taller_insc": taller_insc,
        "concurso_insc": concurso_insc,
        "conferencia_insc": conferencia_insc,
        "talleres_url": talleres_url,
        "concursos_url": concursos_url,
        "conferencias_url": conferencias_url,
    }
    return render(request, "participantes_vista/inicio_participante.html", ctx)


@login_required
def participante_base_context(request):
    """Devuelve contexto base (congreso, approved, labels) para vistas de participantes.

    Nota: Permite acceso aunque la membresía esté pendiente o rechazada; 
    las acciones se restringen desde la UI.
    """
    if not (request.user.is_superuser or request.user.groups.filter(name__in=["Participantes"]).exists()):
        messages.error(request, "Esta página es solo para participantes.")
        return None

    membership = (
        UserCongresoMembership.objects
        .filter(user=request.user, role="participante")
        .order_by("-decided_at", "-id")
        .first()
    )
    congreso = membership.congreso if membership else None
    status = membership.status if membership else "none"
    approved = status == "approved"
    status_label_map = {"pending": "Pendiente de Activación", "approved": "Aprobado", "rejected": "Rechazado", "none": ""}
    status_label = status_label_map.get(status, "")

    return {
        "congreso": congreso,
        "membership": membership,
        "membership_status": status,
        "membership_status_label": status_label,
        "approved": approved,
        # Mantener variables usadas por el topbar/landing
        "talleres_url": reverse("part_talleres"),
        "concursos_url": reverse("part_concursos"),
        "conferencias_url": reverse("part_conferencias"),
    }


@login_required
def part_talleres_view(request):
    ctx = participante_base_context(request)
    if ctx is None:
        return redirect("inicio")
    congreso = ctx.get("congreso")
    talleres = []
    if congreso:
        talleres = list(
            Taller.objects
            .filter(congreso=congreso)
            .annotate(inscritos_count=Count("inscripciones"))
            .order_by("created_at")
        )
    ctx.update({"talleres": talleres})
    return render(request, "participantes_vista/part_talleres.html", ctx)


@login_required
def part_concursos_view(request):
    ctx = participante_base_context(request)
    if ctx is None:
        return redirect("inicio")
    congreso = ctx.get("congreso")
    concursos = []
    if congreso:
        concursos = list(
            Concurso.objects
            .filter(congreso=congreso)
            .annotate(inscritos_count=Count("inscripciones"))
            .order_by("created_at")
        )
        # Calcular capacidad total para concursos grupales (equipos * máx/eq) y dejarla disponible en el objeto
        for c in concursos:
            if c.type == "individual":
                c.capacidad_total = c.cupo_maximo if c.cupo_maximo else None
                # Para concursos individuales no hay equipos: equipos_inscritos None
                c.equipos_inscritos = None
            else:
                equipos = c.numero_equipos or 0
                por_eq = c.max_por_equipo or 0
                total = equipos * por_eq
                c.capacidad_total = total if total > 0 else None
                # No existe todavía un modelo de equipos; mostramos 0 como placeholder.
                # Cuando se implemente Team/Equipo, reemplazar por conteo real distinto.
                c.equipos_inscritos = 0
    ctx.update({"concursos": concursos})
    return render(request, "participantes_vista/part_concursos.html", ctx)


@login_required
def part_conferencias_view(request):
    ctx = participante_base_context(request)
    if ctx is None:
        return redirect("inicio")
    congreso = ctx.get("congreso")
    conferencias = []
    if congreso:
        conferencias = list(
            Conferencia.objects
            .filter(congreso=congreso)
            .annotate(inscritos_count=Count("inscripciones"))
            .order_by("created_at")
        )
    ctx.update({"conferencias": conferencias})
    return render(request, "participantes_vista/part_conferencias.html", ctx)


@login_required
def part_taller_inscribir_view(request, taller_id: int):
    ctx = participante_base_context(request)
    if ctx is None:
        return redirect("inicio")
    congreso = ctx.get("congreso")
    if not congreso:
        return redirect("inicio_participante")
    try:
        taller = Taller.objects.get(pk=taller_id, congreso=congreso)
    except Taller.DoesNotExist:
        messages.error(request, "El taller solicitado no existe en este congreso.")
        return redirect("part_talleres")
    # Manejo de inscripción (POST)
    if request.method == "POST" and request.POST.get("action") == "enroll_taller":
        # Validar aprobación de membresía
        if not ctx.get("approved"):
            messages.error(request, "Tu membresía aún no está aprobada. No puedes inscribirte.")
            return redirect("part_taller_inscribir", taller_id=taller.id)
        # Verificar si ya está inscrito en este taller
        if TallerInscripcion.objects.filter(congreso=congreso, taller=taller, user=request.user).exists():
            messages.info(request, "Ya estás inscrito en este taller.")
            return redirect("part_taller_inscribir", taller_id=taller.id)
        # Verificar si está inscrito en otro taller (solo uno permitido por categoría)
        if TallerInscripcion.objects.filter(congreso=congreso, user=request.user).exclude(taller=taller).exists():
            messages.error(request, "Solo puedes estar inscrito en un taller a la vez.")
            return redirect("part_taller_inscribir", taller_id=taller.id)
        capacidad_tmp = taller.cupo_maximo if taller.cupo_maximo else None
        inscritos_tmp = taller.inscripciones.count()
        if capacidad_tmp is not None and inscritos_tmp >= capacidad_tmp:
            messages.error(request, "El cupo de este taller ya está lleno.")
            return redirect("part_taller_inscribir", taller_id=taller.id)
        # Crear inscripción
        TallerInscripcion.objects.create(
            congreso=congreso,
            taller=taller,
            user=request.user,
        )
        messages.success(request, f"Te has inscrito al taller '{taller.title}'.", extra_tags="taller")
        return redirect("inicio_participante")
    inscritos = taller.inscripciones.count()
    capacidad = taller.cupo_maximo if taller.cupo_maximo else None
    ya_inscrito = TallerInscripcion.objects.filter(congreso=congreso, taller=taller, user=request.user).exists()
    otro_taller_inscrito = False
    if not ya_inscrito:
        otro_taller_inscrito = TallerInscripcion.objects.filter(congreso=congreso, user=request.user).exclude(taller=taller).exists()
    # calcular anterior/siguiente por fecha de creacion (lista mostrada es -created_at)
    prev_taller = (
        Taller.objects.filter(congreso=congreso, created_at__gt=taller.created_at)
        .order_by("created_at")
        .first()
    )
    next_taller = (
        Taller.objects.filter(congreso=congreso, created_at__lt=taller.created_at)
        .order_by("-created_at")
        .first()
    )
    ctx.update({
        "taller": taller,
        "inscritos": inscritos,
        "capacidad": capacidad,
        "ya_inscrito": ya_inscrito,
        "otro_taller_inscrito": otro_taller_inscrito,
        "prev_taller": prev_taller,
        "next_taller": next_taller,
    })
    # Datos completos de participantes (nombre, correo, nivel y campos extra dinámicos)
    extra_fields = []
    participantes_rows = []
    if congreso:
        from .models import ExtraField, UserExtraFieldValue
        inscripciones = list(taller.inscripciones.select_related("user", "performance_level"))
        extra_fields = list(
            ExtraField.objects.filter(
                (Q(congreso=congreso) | Q(congreso__isnull=True)),
                active=True,
                section__in=["registro", "perfil"],
                role_scope__in=["both", "participante"],
            ).order_by("order", "name")
        )
        user_ids = [ins.user_id for ins in inscripciones] or [0]
        field_ids = [f.id for f in extra_fields] or [0]
        values_qs = UserExtraFieldValue.objects.filter(
            (Q(congreso=congreso) | Q(congreso__isnull=True)),
            user_id__in=user_ids,
            field_id__in=field_ids,
        ).values("user_id", "field_id", "value")
        extra_values_by_user = {}
        for v in values_qs:
            d = extra_values_by_user.setdefault(v["user_id"], {})
            d[v["field_id"]] = v["value"]
        for ins in inscripciones:
            ordered = [extra_values_by_user.get(ins.user_id, {}).get(f.id, "") for f in extra_fields]
            participantes_rows.append({"ins": ins, "values_ordered": ordered})
    ctx.update({
        "extra_fields": extra_fields,
        "participantes_rows": participantes_rows,
    })
    return render(request, "participantes_vista/part_taller_inscribir.html", ctx)


@login_required
def part_concurso_inscribir_view(request, concurso_id: int):
    ctx = participante_base_context(request)
    if ctx is None:
        return redirect("inicio")
    congreso = ctx.get("congreso")
    if not congreso:
        return redirect("inicio_participante")
    try:
        concurso = Concurso.objects.get(pk=concurso_id, congreso=congreso)
    except Concurso.DoesNotExist:
        messages.error(request, "El concurso solicitado no existe en este congreso.")
        return redirect("part_concursos")
    # Manejo de inscripción concurso
    if request.method == "POST" and request.POST.get("action") == "enroll_concurso":
        if not ctx.get("approved"):
            messages.error(request, "Tu membresía aún no está aprobada. No puedes inscribirte.")
            return redirect("part_concurso_inscribir", concurso_id=concurso.id)
        if ConcursoInscripcion.objects.filter(congreso=congreso, concurso=concurso, user=request.user).exists():
            messages.info(request, "Ya estás inscrito en este concurso.")
            return redirect("part_concurso_inscribir", concurso_id=concurso.id)
        if ConcursoInscripcion.objects.filter(congreso=congreso, user=request.user).exclude(concurso=concurso).exists():
            messages.error(request, "Solo puedes estar inscrito en un concurso a la vez.")
            return redirect("part_concurso_inscribir", concurso_id=concurso.id)
        # Validación obligatoria de equipo completo para concursos grupales
        if concurso.type == "grupal":
            max_equipo = concurso.max_por_equipo or 0
            if max_equipo < 2:
                messages.error(request, "Este concurso grupal no tiene un tamaño de equipo válido configurado.")
                return redirect("part_concurso_inscribir", concurso_id=concurso.id)
            required_slots = list(range(2, max_equipo + 1))  # integrante #1 es el usuario actual
            miembros_ids = []
            for slot in required_slots:
                field = f"team_member_{slot}"
                raw_val = (request.POST.get(field) or "").strip()
                if not raw_val:
                    messages.error(request, "Debes seleccionar todos los integrantes del equipo antes de inscribirte.")
                    return redirect("part_concurso_inscribir", concurso_id=concurso.id)
                miembros_ids.append(raw_val)
            # Unicidad
            if len(set(miembros_ids)) != len(miembros_ids):
                messages.error(request, "Cada integrante adicional debe ser único. Has repetido algún participante.")
                return redirect("part_concurso_inscribir", concurso_id=concurso.id)
            # No incluirse a sí mismo
            if str(request.user.id) in miembros_ids:
                messages.error(request, "No debes incluirte como integrante adicional; ya eres el integrante #1.")
                return redirect("part_concurso_inscribir", concurso_id=concurso.id)
            # Deben ser participantes aprobados del mismo congreso
            aprobados_ids = set(
                UserCongresoMembership.objects.filter(
                    congreso=congreso,
                    role="participante",
                    status="approved",
                    user_id__in=miembros_ids,
                ).values_list("user_id", flat=True)
            )
            if len(aprobados_ids) != len(miembros_ids):
                messages.error(request, "Alguno de los integrantes seleccionados no está aprobado en este congreso.")
                return redirect("part_concurso_inscribir", concurso_id=concurso.id)
            # Verificar que ninguno esté inscrito ya en este u otro concurso
            already_other = ConcursoInscripcion.objects.filter(congreso=congreso, user_id__in=miembros_ids).exclude(concurso=concurso)
            if already_other.exists():
                messages.error(request, "Alguno de los integrantes ya está inscrito en otro concurso.")
                return redirect("part_concurso_inscribir", concurso_id=concurso.id)
            already_this = ConcursoInscripcion.objects.filter(congreso=congreso, concurso=concurso, user_id__in=miembros_ids)
            if already_this.exists():
                messages.error(request, "Alguno de los integrantes ya está inscrito en este concurso.")
                return redirect("part_concurso_inscribir", concurso_id=concurso.id)
            # Nombre de equipo obligatorio
            team_name = (request.POST.get("team_name") or "").strip()
            if not team_name:
                messages.error(request, "Debes ingresar el nombre del equipo.")
                return redirect("part_concurso_inscribir", concurso_id=concurso.id)
            # Capacidad considerando todos los integrantes a agregar
            capacidad_tmp_check = (concurso.numero_equipos or 0) * (concurso.max_por_equipo or 0) or None
            inscritos_tmp_check = concurso.inscripciones.count()
            if capacidad_tmp_check is not None and inscritos_tmp_check + max_equipo > capacidad_tmp_check:
                messages.error(request, "No hay cupo suficiente para registrar un equipo completo.")
                return redirect("part_concurso_inscribir", concurso_id=concurso.id)
            # Crear equipo y miembros
            equipo = ConcursoEquipo.objects.create(
                congreso=congreso,
                concurso=concurso,
                nombre=team_name,
                lider=request.user,
            )
            # Miembros adicionales
            for uid in miembros_ids:
                ConcursoEquipoMiembro.objects.create(equipo=equipo, user_id=uid)
            # Registrar líder también como miembro
            ConcursoEquipoMiembro.objects.create(equipo=equipo, user=request.user)
            # Crear inscripciones para todos (incluye al líder al final fuera del if global)
            for uid in miembros_ids:
                ConcursoInscripcion.objects.create(congreso=congreso, concurso=concurso, user_id=uid)
        # Capacidad
        if concurso.type == "individual":
            capacidad_tmp = concurso.cupo_maximo if concurso.cupo_maximo else None
        else:
            capacidad_tmp = (concurso.numero_equipos or 0) * (concurso.max_por_equipo or 0) or None
        inscritos_tmp = concurso.inscripciones.count()
        if capacidad_tmp is not None and inscritos_tmp >= capacidad_tmp:
            messages.error(request, "El cupo de este concurso ya está lleno.")
            return redirect("part_concurso_inscribir", concurso_id=concurso.id)
        # Registrar inscripción del solicitante (líder) si todavía no existe
        if not ConcursoInscripcion.objects.filter(congreso=congreso, concurso=concurso, user=request.user).exists():
            ConcursoInscripcion.objects.create(
                congreso=congreso,
                concurso=concurso,
                user=request.user,
            )
        if concurso.type == "grupal":
            messages.success(request, f"Has registrado el equipo '{team_name}' en el concurso '{concurso.title}'.", extra_tags="concurso")
        else:
            messages.success(request, f"Te has inscrito al concurso '{concurso.title}'.", extra_tags="concurso")
        return redirect("inicio_participante")
    inscritos = concurso.inscripciones.count()
    if concurso.type == "individual":
        capacidad = concurso.cupo_maximo if concurso.cupo_maximo else None
        equipos_inscritos = None
    else:
        capacidad = (concurso.numero_equipos or 0) * (concurso.max_por_equipo or 0) or None
        equipos_inscritos = concurso.equipos.count()
    ya_inscrito = ConcursoInscripcion.objects.filter(congreso=congreso, concurso=concurso, user=request.user).exists()
    otro_concurso_inscrito = False
    if not ya_inscrito:
        otro_concurso_inscrito = ConcursoInscripcion.objects.filter(congreso=congreso, user=request.user).exclude(concurso=concurso).exists()
    prev_concurso = (
        Concurso.objects.filter(congreso=congreso, created_at__gt=concurso.created_at)
        .order_by("created_at")
        .first()
    )
    next_concurso = (
        Concurso.objects.filter(congreso=congreso, created_at__lt=concurso.created_at)
        .order_by("-created_at")
        .first()
    )
    # Datos para formulario de equipos (UI)
    equipo_participantes = []
    max_integrantes_equipo = concurso.max_por_equipo or 0
    max_integrantes_adicionales = max(0, max_integrantes_equipo - 1)  # adicionales al usuario actual
    min_integrantes_equipo = max_integrantes_equipo if max_integrantes_equipo >= 2 else 1  # ahora mínimo = máximo para grupos
    team_slots = []  # Números de integrante (2..max) para generar selects individuales
    if concurso.type == "grupal":
        # Participantes aprobados del mismo congreso, excluyéndose a sí mismo
        miembros = (
            UserCongresoMembership.objects
            .filter(congreso=congreso, role="participante", status="approved")
            .exclude(user=request.user)
            .select_related("user")
            .order_by("user__first_name", "user__last_name", "user__username")
        )
        for m in miembros:
            nombre = (m.user.get_full_name() or m.user.username or m.user.email or m.user.username)
            correo = m.user.email or "sin-correo"
            equipo_participantes.append({"id": m.user_id, "name": nombre, "email": correo, "display": f"{nombre} ({correo})"})
        if max_integrantes_equipo >= 2:
            team_slots = list(range(2, max_integrantes_equipo + 1))
    ctx.update({
        "concurso": concurso,
        "inscritos": inscritos,
        "capacidad": capacidad,
        "ya_inscrito": ya_inscrito,
        "otro_concurso_inscrito": otro_concurso_inscrito,
        "equipos_inscritos": equipos_inscritos,
        "prev_concurso": prev_concurso,
        "next_concurso": next_concurso,
        # Equipo UI
        "equipo_participantes": equipo_participantes,
        "max_integrantes_equipo": max_integrantes_equipo,
        "max_integrantes_adicionales": max_integrantes_adicionales,
        "min_integrantes_equipo": min_integrantes_equipo,
        "team_slots": team_slots,
        "required_integrantes_adicionales": max(0, max_integrantes_equipo - 1),
    })
    # Datos completos de participantes (nombre, correo, nivel y campos extra dinámicos)
    extra_fields = []
    participantes_rows = []
    team_by_user = {}
    if congreso:
        from .models import ExtraField, UserExtraFieldValue
        inscripciones = list(concurso.inscripciones.select_related("user", "performance_level"))
        extra_fields = list(
            ExtraField.objects.filter(
                (Q(congreso=congreso) | Q(congreso__isnull=True)),
                active=True,
                section__in=["registro", "perfil"],
                role_scope__in=["both", "participante"],
            ).order_by("order", "name")
        )
        user_ids = [ins.user_id for ins in inscripciones] or [0]
        field_ids = [f.id for f in extra_fields] or [0]
        values_qs = UserExtraFieldValue.objects.filter(
            (Q(congreso=congreso) | Q(congreso__isnull=True)),
            user_id__in=user_ids,
            field_id__in=field_ids,
        ).values("user_id", "field_id", "value")
        extra_values_by_user = {}
        for v in values_qs:
            d = extra_values_by_user.setdefault(v["user_id"], {})
            d[v["field_id"]] = v["value"]
        # Si es grupal, mapear equipo por usuario
        if concurso.type == "grupal" and user_ids:
            membs = (
                ConcursoEquipoMiembro.objects
                .filter(equipo__concurso=concurso, user_id__in=user_ids)
                .select_related("equipo")
            )
            for m in membs:
                team_by_user[m.user_id] = m.equipo.nombre if m.equipo and m.equipo.nombre else ""
        for ins in inscripciones:
            ordered = [extra_values_by_user.get(ins.user_id, {}).get(f.id, "") for f in extra_fields]
            row = {"ins": ins, "values_ordered": ordered}
            if concurso.type == "grupal":
                row["team_name"] = team_by_user.get(ins.user_id, "")
            participantes_rows.append(row)
    ctx.update({
        "extra_fields": extra_fields,
        "participantes_rows": participantes_rows,
        "team_by_user": team_by_user,
    })
    return render(request, "participantes_vista/part_concurso_inscribir.html", ctx)


@login_required
def part_conferencia_inscribir_view(request, conferencia_id: int):
    ctx = participante_base_context(request)
    if ctx is None:
        return redirect("inicio")
    congreso = ctx.get("congreso")
    if not congreso:
        return redirect("inicio_participante")
    try:
        conferencia = Conferencia.objects.get(pk=conferencia_id, congreso=congreso)
    except Conferencia.DoesNotExist:
        messages.error(request, "La conferencia solicitada no existe en este congreso.")
        return redirect("part_conferencias")
    if request.method == "POST" and request.POST.get("action") == "enroll_conferencia":
        if not ctx.get("approved"):
            messages.error(request, "Tu membresía aún no está aprobada. No puedes inscribirte.")
            return redirect("part_conferencia_inscribir", conferencia_id=conferencia.id)
        if ConferenciaInscripcion.objects.filter(congreso=congreso, conferencia=conferencia, user=request.user).exists():
            messages.info(request, "Ya estás inscrito en esta conferencia.")
            return redirect("part_conferencia_inscribir", conferencia_id=conferencia.id)
        if ConferenciaInscripcion.objects.filter(congreso=congreso, user=request.user).exclude(conferencia=conferencia).exists():
            messages.error(request, "Solo puedes estar inscrito en una conferencia a la vez.")
            return redirect("part_conferencia_inscribir", conferencia_id=conferencia.id)
        capacidad_tmp = conferencia.cupo_maximo if conferencia.cupo_maximo else None
        inscritos_tmp = conferencia.inscripciones.count()
        if capacidad_tmp is not None and inscritos_tmp >= capacidad_tmp:
            messages.error(request, "El cupo de esta conferencia ya está lleno.")
            return redirect("part_conferencia_inscribir", conferencia_id=conferencia.id)
        ConferenciaInscripcion.objects.create(
            congreso=congreso,
            conferencia=conferencia,
            user=request.user,
        )
        messages.success(request, f"Te has inscrito a la conferencia '{conferencia.title}'.", extra_tags="conferencia")
        return redirect("inicio_participante")
    inscritos = conferencia.inscripciones.count()
    capacidad = conferencia.cupo_maximo if conferencia.cupo_maximo else None
    ya_inscrito = ConferenciaInscripcion.objects.filter(congreso=congreso, conferencia=conferencia, user=request.user).exists()
    otra_conferencia_inscrita = False
    if not ya_inscrito:
        otra_conferencia_inscrita = ConferenciaInscripcion.objects.filter(congreso=congreso, user=request.user).exclude(conferencia=conferencia).exists()
    prev_conferencia = (
        Conferencia.objects.filter(congreso=congreso, created_at__gt=conferencia.created_at)
        .order_by("created_at")
        .first()
    )
    next_conferencia = (
        Conferencia.objects.filter(congreso=congreso, created_at__lt=conferencia.created_at)
        .order_by("-created_at")
        .first()
    )
    ctx.update({
        "conferencia": conferencia,
        "inscritos": inscritos,
        "capacidad": capacidad,
        "ya_inscrito": ya_inscrito,
        "otra_conferencia_inscrita": otra_conferencia_inscrita,
        "prev_conferencia": prev_conferencia,
        "next_conferencia": next_conferencia,
    })
    # Datos completos de participantes (nombre, correo, nivel y campos extra dinámicos)
    extra_fields = []
    participantes_rows = []
    if congreso:
        from .models import ExtraField, UserExtraFieldValue
        inscripciones = list(conferencia.inscripciones.select_related("user", "performance_level"))
        extra_fields = list(
            ExtraField.objects.filter(
                (Q(congreso=congreso) | Q(congreso__isnull=True)),
                active=True,
                section__in=["registro", "perfil"],
                role_scope__in=["both", "participante"],
            ).order_by("order", "name")
        )
        user_ids = [ins.user_id for ins in inscripciones] or [0]
        field_ids = [f.id for f in extra_fields] or [0]
        values_qs = UserExtraFieldValue.objects.filter(
            (Q(congreso=congreso) | Q(congreso__isnull=True)),
            user_id__in=user_ids,
            field_id__in=field_ids,
        ).values("user_id", "field_id", "value")
        extra_values_by_user = {}
        for v in values_qs:
            d = extra_values_by_user.setdefault(v["user_id"], {})
            d[v["field_id"]] = v["value"]
        for ins in inscripciones:
            ordered = [extra_values_by_user.get(ins.user_id, {}).get(f.id, "") for f in extra_fields]
            participantes_rows.append({"ins": ins, "values_ordered": ordered})
    ctx.update({
        "extra_fields": extra_fields,
        "participantes_rows": participantes_rows,
    })
    return render(request, "participantes_vista/part_conferencia_inscribir.html", ctx)


@login_required
def part_avisos_view(request):
    ctx = participante_base_context(request)
    if ctx is None:
        return redirect("inicio")
    # Cargar avisos del congreso (más recientes primero)
    from .models import Aviso
    congreso = ctx.get("congreso")
    avisos = []
    if congreso:
        avisos = list(
            Aviso.objects.filter(congreso=congreso).order_by("-created_at")
        )
    ctx.update({"avisos": avisos})
    return render(request, "participantes_vista/part_avisos.html", ctx)


@login_required
def part_perfil_view(request):
    ctx = participante_base_context(request)
    if ctx is None:
        return redirect("inicio")
    from .models import PerformanceLevel, ExtraField, UserExtraFieldValue, UserCongresoMembership
    user = request.user
    membership: UserCongresoMembership | None = ctx.get("membership")
    congreso = ctx.get("congreso")

    # Campos extra para sección PERFIL (rol participante + globales/congreso)
    perfil_extra_fields = []
    if congreso:
        perfil_extra_fields = list(
            ExtraField.objects.filter(active=True, section="perfil")
            .filter(Q(congreso=congreso) | Q(congreso__isnull=True))
            .filter(role_scope__in=["both", "participante"])  # visibles para participante
            .order_by("order", "name")
        )

    # Valores actuales de esos campos para el usuario
    extra_values_map: dict[int, str] = {}
    if perfil_extra_fields:
        vals = UserExtraFieldValue.objects.filter(user=user, field__in=perfil_extra_fields).filter(
            Q(congreso=congreso) | Q(congreso__isnull=True)
        )
        for v in vals:
            extra_values_map[v.field_id] = v.value

    levels = []
    if congreso:
        levels = list(PerformanceLevel.objects.filter(congreso=congreso).order_by("name"))

    # Dominios permitidos para este congreso (para cambio de usuario/correo)
    allowed_domains_profile = []
    if congreso:
        allowed_domains_profile = [d.lower().strip() for d in AllowedEmailDomain.objects.filter(congreso=congreso).values_list("domain", flat=True) if d]

    # Procesar acciones del formulario
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "update_public" and membership and membership.status == "approved":
            first_name = request.POST.get("first_name", "").strip()[:150]
            apellido_paterno = request.POST.get("apellido_paterno", "").strip()[:150]
            apellido_materno = request.POST.get("apellido_materno", "").strip()[:150]
            perf_id = request.POST.get("performance_level_id")
            # Actualizar nombre(s) y guardar apellidos en extra fields si existen, sino crear globales si se desea.
            if first_name:
                user.first_name = first_name
            # Guardar apellidos en extra fields: buscar por code
            def _save_extra(code: str, label: str, value: str):
                if not value:
                    return
                ef = next((f for f in perfil_extra_fields if f.code == code), None)
                if ef:
                    uev, _ = UserExtraFieldValue.objects.get_or_create(user=user, congreso=congreso, field=ef)
                    uev.value = value
                    uev.save()
            _save_extra("apellido_paterno", "Apellido Paterno", apellido_paterno)
            _save_extra("apellido_materno", "Apellido Materno", apellido_materno)
            # Performance level
            if perf_id and perf_id.isdigit():
                try:
                    pl = PerformanceLevel.objects.get(id=int(perf_id), congreso=congreso)
                    membership.performance_level = pl
                except PerformanceLevel.DoesNotExist:
                    pass
            membership.save()
            user.save()
            # Otros campos extra dinámicos
            for f in perfil_extra_fields:
                if f.code in ["apellido_paterno", "apellido_materno"]:
                    continue  # ya manejados
                field_name = f"extra_{f.id}"
                raw_val = request.POST.get(field_name, "").strip()
                if f.field_type == "boolean":
                    raw_val = "1" if request.POST.get(field_name) else "0"
                if raw_val or not f.required:
                    # Validar unicidad si aplica
                    if getattr(f, "unique_value", False) and raw_val:
                        q = UserExtraFieldValue.objects.filter(field=f, value=raw_val)
                        if f.congreso_id:
                            q = q.filter(congreso=congreso)
                        else:
                            q = q.filter(congreso__isnull=True)
                        q = q.exclude(user=user)
                        if q.exists():
                            messages.error(request, f"El valor '{raw_val}' ya está en uso para '{f.name}'. Debe ser único.")
                            return redirect("part_perfil")
                    uev, _ = UserExtraFieldValue.objects.get_or_create(user=user, congreso=congreso if f.congreso_id else None, field=f)
                    uev.value = raw_val
                    uev.save()
            messages.success(request, "Datos públicos actualizados.")
            return redirect("part_perfil")
        elif action == "change_user" and membership and membership.status == "approved":
            # Ahora solo un campo: username que también funciona como correo.
            new_username = (request.POST.get("username") or "").strip()
            if not new_username:
                messages.error(request, "El usuario (correo) no puede estar vacío.")
                return redirect("part_perfil")
            # Debe ser correo válido
            try:
                validate_email(new_username)
            except ValidationError:
                messages.error(request, "Debe ser un correo electrónico válido.")
                return redirect("part_perfil")
            # Validar dominio permitido si hay lista configurada
            if allowed_domains_profile:
                email_lc = new_username.lower()
                if not any(email_lc.endswith(d) for d in allowed_domains_profile):
                    pretty = ", ".join(sorted(allowed_domains_profile))
                    messages.error(request, f"El correo no pertenece a un dominio permitido de este congreso: {pretty}")
                    return redirect("part_perfil")
            # Unicidad (username y email se emparejan al mismo valor)
            if User.objects.exclude(pk=user.pk).filter(Q(username__iexact=new_username) | Q(email__iexact=new_username)).exists():
                messages.error(request, "Ese correo ya está en uso por otra cuenta.")
                return redirect("part_perfil")
            user.username = new_username
            user.email = new_username
            user.save()
            membership.save()
            messages.success(request, "Usuario actualizado correctamente.")
            return redirect("part_perfil")
        elif action == "change_password" and membership and membership.status == "approved":
            current_password = request.POST.get("current_password", "")
            new_password = request.POST.get("new_password", "")
            confirm_password = request.POST.get("confirm_password", "")
            if not user.check_password(current_password):
                messages.error(request, "La contraseña actual es incorrecta.")
            elif new_password != confirm_password:
                messages.error(request, "La confirmación no coincide.")
            elif len(new_password) < 8:
                messages.error(request, "La nueva contraseña debe tener al menos 8 caracteres.")
            else:
                user.set_password(new_password)
                user.save()
                update_session_auth_hash(request, user)
                messages.success(request, "Contraseña actualizada correctamente.")
                return redirect("part_perfil")

    ctx.update({
        "perfil_extra_fields": perfil_extra_fields,
        "extra_values_map": extra_values_map,
        "levels": levels,
        "allowed_domains_profile": allowed_domains_profile,
        # IDs de campos apellido si existen para precargar
        "apellido_paterno_field_id": next((f.id for f in perfil_extra_fields if f.code == "apellido_paterno"), None),
        "apellido_materno_field_id": next((f.id for f in perfil_extra_fields if f.code == "apellido_materno"), None),
    })
    return render(request, "participantes_vista/part_perfil.html", ctx)


# -------------------
# INSTRUCTOR: PERFIL Y ALUMNOS
# -------------------
@login_required
@instructors_area_only
def instructor_perfil_view(request):
    if not (request.user.is_superuser or request.user.groups.filter(name__in=["Instructores"]).exists()):
        messages.error(request, "Esta página es solo para instructores.")
        return redirect("inicio")

    # Congreso principal (última membresía aprobada) para mostrar información contextual
    membership = (
        UserCongresoMembership.objects
        .filter(user=request.user, role="instructor", status="approved")
        .order_by("-decided_at", "-id")
        .first()
    )
    congreso = membership.congreso if membership else None

    # Dominios permitidos agregando todos los congresos aprobados del instructor
    domain_set = set()
    approved_memberships = UserCongresoMembership.objects.filter(user=request.user, status="approved").select_related("congreso")
    for mem in approved_memberships:
        doms = AllowedEmailDomain.objects.filter(congreso=mem.congreso).values_list("domain", flat=True)
        for d in doms:
            if d:
                domain_set.add(d.lower().strip())

    # Prellenar apellidos paterno/materno a partir de last_name actual (heurística simple)
    ln = (request.user.last_name or "").strip()
    ln_pat = ""
    ln_mat = ""
    if ln:
        parts = ln.split()
        if len(parts) >= 2:
            ln_pat = parts[0]
            ln_mat = " ".join(parts[1:])
        else:
            ln_pat = ln

    context = {
        "congreso": congreso,
        "username_value": request.user.username,
        "first_name_value": request.user.first_name,
        "last_name_value": request.user.last_name,
        "last_name_paterno_value": ln_pat,
        "last_name_materno_value": ln_mat,
        "allowed_domains_profile": sorted(domain_set),
    }

    if request.method == "POST":
        action = request.POST.get("action")
        # Actualizar nombre y apellidos
        if action == "update_fullname":
            first = request.POST.get("first_name", "").strip()
            last_pat = request.POST.get("last_name_paterno", "").strip()
            last_mat = request.POST.get("last_name_materno", "").strip()
            combined_last = (last_pat + (" " + last_mat if last_mat else "")).strip()
            if len(first) > 150 or len(combined_last) > 150 or len(last_pat) > 150 or len(last_mat) > 150:
                messages.error(request, "Nombre o apellido demasiado largo (máx 150 caracteres).", extra_tags="fullname")
            else:
                request.user.first_name = first
                request.user.last_name = combined_last
                request.user.save(update_fields=["first_name", "last_name"])
                messages.success(request, "Nombre y apellidos actualizados correctamente.", extra_tags="fullname")
                context["first_name_value"] = first
                context["last_name_value"] = combined_last
                context["last_name_paterno_value"] = last_pat
                context["last_name_materno_value"] = last_mat
            return HttpResponseRedirect(reverse("instructor_perfil") + "#fullname")

        elif action == "update_name":
            new_username = request.POST.get("username", "").strip()
            if not new_username:
                messages.error(request, "El usuario (correo) no puede estar vacío.", extra_tags="user")
            elif User.objects.exclude(pk=request.user.pk).filter(username=new_username).exists():
                messages.error(request, "Ese usuario ya está en uso.", extra_tags="user")
            elif len(new_username) > 150:
                messages.error(request, "El usuario no puede exceder 150 caracteres.", extra_tags="user")
            else:
                try:
                    validate_email(new_username)
                except ValidationError:
                    messages.error(request, "Debe ser un correo electrónico válido.", extra_tags="user")
                    return HttpResponseRedirect(reverse("instructor_perfil") + "#user")
                allowed_domains_profile = context.get("allowed_domains_profile") or []
                if "@" in new_username and allowed_domains_profile:
                    email_lc = new_username.lower()
                    if not any(email_lc.endswith(d) for d in allowed_domains_profile):
                        pretty = ", ".join(allowed_domains_profile)
                        messages.error(
                            request,
                            f"El correo no pertenece a un dominio permitido por tus congresos activos: {pretty}",
                            extra_tags="user",
                        )
                        return HttpResponseRedirect(reverse("instructor_perfil") + "#user")
                request.user.username = new_username
                if "@" in new_username:
                    request.user.email = new_username
                request.user.save()
                messages.success(request, "Usuario actualizado correctamente.", extra_tags="user")
                context["username_value"] = new_username
            return HttpResponseRedirect(reverse("instructor_perfil") + "#user")

        elif action == "update_password":
            current = request.POST.get("current_password", "")
            new = request.POST.get("new_password", "")
            confirm = request.POST.get("confirm_password", "")
            if not request.user.check_password(current):
                messages.error(request, "La contraseña actual no es correcta.", extra_tags="password")
            elif new != confirm:
                messages.error(request, "La nueva contraseña y su confirmación no coinciden.", extra_tags="password")
            elif len(new) < 8 or len(new) > 16:
                messages.error(request, "La nueva contraseña debe tener entre 8 y 16 caracteres.", extra_tags="password")
            else:
                specials = set(".!%&$}")
                if not any(c.isupper() for c in new) or not any(c.islower() for c in new) or not any(c in specials for c in new):
                    messages.error(request, "La nueva contraseña debe contener mayúscula, minúscula y símbolo (.!%&$}).", extra_tags="password")
                else:
                    try:
                        request.user.set_password(new)
                        request.user.save()
                        update_session_auth_hash(request, request.user)
                        messages.success(request, "Contraseña actualizada correctamente.", extra_tags="password")
                    except Exception:
                        messages.error(request, "Error al actualizar la contraseña. Intenta nuevamente.", extra_tags="password")
            return HttpResponseRedirect(reverse("instructor_perfil") + "#password")

    return render(request, "instructores_vista/perfil_instructor.html", context)


@login_required
@instructors_area_only
def instructor_participantes_view(request):
    if not (request.user.is_superuser or request.user.groups.filter(name__in=["Instructores"]).exists()):
        messages.error(request, "Esta página es solo para instructores.")
        return redirect("inicio")
    membership = (
        UserCongresoMembership.objects
        .filter(user=request.user, role="instructor", status="approved")
        .order_by("-decided_at", "-id")
        .first()
    )
    congreso = membership.congreso if membership else None
    # Construir tablero de recursos (talleres, concursos, conferencias) del instructor en este congreso
    talleres = list(Taller.objects.filter(congreso=congreso, instructor=request.user).order_by("created_at"))
    concursos = list(Concurso.objects.filter(congreso=congreso, instructor=request.user).order_by("created_at"))
    conferencias = list(Conferencia.objects.filter(congreso=congreso, instructor=request.user).order_by("created_at"))
    recurso_rows = []
    for t in talleres:
        recurso_rows.append({"tipo": "taller", "obj": t, "titulo": t.title, "inscritos": t.inscripciones.count(), "url": f"{reverse('ins_taller_participantes')}?id={t.id}"})
    for c in concursos:
        recurso_rows.append({"tipo": "concurso", "obj": c, "titulo": c.title, "inscritos": c.inscripciones.count(), "url": f"{reverse('ins_concurso_participantes')}?id={c.id}"})
    for cf in conferencias:
        recurso_rows.append({"tipo": "conferencia", "obj": cf, "titulo": cf.title, "inscritos": cf.inscripciones.count(), "url": f"{reverse('ins_conferencia_participantes')}?id={cf.id}"})
    recurso_rows.sort(key=lambda r: (r["tipo"], getattr(r["obj"], "created_at", None)))
    return render(request, "instructores_vista/instructor_participantes.html", {"congreso": congreso, "recurso_rows": recurso_rows})

# -------------------
# INSTRUCTOR: RECURSOS (Talleres / Concursos / Conferencias)
# - Si el instructor ya tiene uno vinculado en el congreso activo, mostrar vista previa tipo detalle.
# - Si no tiene, mostrar formulario básico de creación (similar a las páginas de admin pero simplificado).
# -------------------
@login_required
@instructors_area_only
def _get_active_instructor_congreso(request):
    if not (request.user.is_superuser or request.user.groups.filter(name__in=["Instructores"]).exists()):
        messages.error(request, "Esta página es solo para instructores.")
        return None
    active_id = request.session.get("instructor_active_congreso_id")
    membership = None
    if active_id:
        membership = UserCongresoMembership.objects.filter(
            user=request.user, role="instructor", status="approved", congreso_id=active_id
        ).first()
    if not membership:
        membership = (
            UserCongresoMembership.objects
            .filter(user=request.user, role="instructor", status="approved")
            .order_by("-decided_at", "-id").first()
        )
        if membership:
            request.session["instructor_active_congreso_id"] = membership.congreso_id
    return membership.congreso if membership else None


@login_required
@instructors_area_only
def instructor_talleres_view(request):
    """Listado y vista de talleres para el instructor.

    Ahora soporta múltiples talleres. Permite seleccionar uno vía ?id=<taller_id>,
    crear nuevos con action=create_taller y eliminar el activo con action=delete_taller.
    Modo creación explícito con ?nuevo=1.
    """
    congreso = _get_active_instructor_congreso(request)
    if not congreso:
        return redirect("instructor_pendiente")
    talleres_qs = Taller.objects.filter(congreso=congreso, instructor=request.user).order_by("-created_at")
    nuevo_flag = request.GET.get("nuevo") == "1"
    active_id = request.GET.get("id")
    active = None
    if active_id:
        try:
            active = talleres_qs.get(pk=active_id)
        except Taller.DoesNotExist:
            active = None
    if not active and not nuevo_flag:
        active = talleres_qs.first()
    # Crear
    if request.method == "POST" and request.POST.get("action") == "create_taller":
        title = (request.POST.get("title") or "").strip()
        lugar = (request.POST.get("lugar") or "").strip()
        cupo = request.POST.get("cupo") or None
        description = (request.POST.get("descripcion") or "").strip()
        if not title:
            messages.error(request, "El nombre del taller es obligatorio.")
        else:
            obj = Taller.objects.create(
                congreso=congreso,
                title=title,
                lugar=lugar,
                cupo_maximo=cupo or None,
                description=description,
                instructor=request.user,
            )
            image = request.FILES.get("image")
            if image:
                if image.size > 3 * 1024 * 1024:
                    messages.warning(request, "La imagen supera 3MB, se ignoró.")
                else:
                    obj.image = image
                    obj.save()
            messages.success(request, "Taller creado correctamente.")
            return redirect("ins_talleres")
    # Eliminar activo
    if request.method == "POST" and request.POST.get("action") == "delete_taller" and active:
        if request.user.is_superuser or request.user.id == active.instructor_id or request.user.id == congreso.admin_user_id:
            active.delete()
            messages.success(request, "Taller eliminado correctamente.")
        else:
            messages.error(request, "No tienes permisos para eliminar este taller.")
        return redirect("ins_talleres")
    taller_inscritos = active.inscripciones.count() if active else 0
    taller_capacidad = active.cupo_maximo if active and active.cupo_maximo else None
    # Listado embebido con datos completos de participantes
    participantes_rows = []
    extra_fields = []
    if active:
        inscripciones = list(active.inscripciones.select_related("user", "performance_level"))
        from .models import ExtraField, UserExtraFieldValue
        extra_fields = list(
            ExtraField.objects.filter(
                (Q(congreso=congreso) | Q(congreso__isnull=True)),
                active=True,
                section__in=["registro", "perfil"],
                role_scope__in=["both", "participante"],
            ).order_by("order", "name")
        )
        user_ids = [ins.user_id for ins in inscripciones] or [0]
        field_ids = [f.id for f in extra_fields] or [0]
        values_qs = UserExtraFieldValue.objects.filter(
            (Q(congreso=congreso) | Q(congreso__isnull=True)),
            user_id__in=user_ids,
            field_id__in=field_ids,
        ).values("user_id", "field_id", "value")
        extra_values_by_user = {}
        for v in values_qs:
            d = extra_values_by_user.setdefault(v["user_id"], {})
            d[v["field_id"]] = v["value"]
        for ins in inscripciones:
            ordered = [extra_values_by_user.get(ins.user_id, {}).get(f.id, "") for f in extra_fields]
            participantes_rows.append({"ins": ins, "values_ordered": ordered})
    ctx = {
        "congreso": congreso,
        "talleres": list(talleres_qs),
        "taller": active,
        "taller_inscritos": taller_inscritos,
        "taller_capacidad": taller_capacidad,
        "extra_fields": extra_fields,
        "participantes_rows": participantes_rows,
        "creation_mode": nuevo_flag,
    }
    return render(request, "instructores_vista/ins_taller.html", ctx)


@login_required
@instructors_area_only
def instructor_concursos_view(request):
    congreso = _get_active_instructor_congreso(request)
    if not congreso:
        return redirect("instructor_pendiente")
    concursos_qs = Concurso.objects.filter(congreso=congreso, instructor=request.user).order_by("-created_at")
    nuevo_flag = request.GET.get("nuevo") == "1"
    active_id = request.GET.get("id")
    active = None
    if active_id:
        try:
            active = concursos_qs.get(pk=active_id)
        except Concurso.DoesNotExist:
            active = None
    if not active and not nuevo_flag:
        active = concursos_qs.first()
    # Crear
    if request.method == "POST" and request.POST.get("action") == "create_concurso":
        title = (request.POST.get("title") or "").strip()
        tipo = (request.POST.get("type") or request.POST.get("tipo") or "individual").strip()
        lugar = (request.POST.get("lugar") or "").strip()
        description = (request.POST.get("descripcion") or request.POST.get("description") or "").strip()
        cupo = request.POST.get("cupo") or request.POST.get("cupo_maximo") or None
        equipos = request.POST.get("numero_equipos") or None
        maxeq = request.POST.get("max_por_equipo") or None
        if not title:
            messages.error(request, "El nombre del concurso es obligatorio.")
        else:
            obj = Concurso.objects.create(
                congreso=congreso,
                title=title,
                type=tipo,
                lugar=lugar,
                description=description,
                cupo_maximo=cupo or None,
                numero_equipos=equipos or None,
                max_por_equipo=maxeq or None,
                instructor=request.user,
            )
            image = request.FILES.get("image")
            if image:
                if image.size > 3 * 1024 * 1024:
                    messages.warning(request, "La imagen supera 3MB, se ignoró.")
                else:
                    obj.image = image
                    obj.save()
            messages.success(request, "Concurso creado correctamente.", extra_tags="concurso")
            return redirect("ins_concursos")
    # Eliminar activo
    if request.method == "POST" and request.POST.get("action") == "delete_concurso" and active:
        if request.user.is_superuser or request.user.id == active.instructor_id or request.user.id == congreso.admin_user_id:
            active.delete()
            messages.success(request, "Concurso eliminado correctamente.", extra_tags="concurso")
        else:
            messages.error(request, "No tienes permisos para eliminar este concurso.")
        return redirect("ins_concursos")
    concurso_inscritos = active.inscripciones.count() if active else 0
    if active:
        if active.type == "individual":
            concurso_capacidad = active.cupo_maximo if active.cupo_maximo else None
        else:
            if active.numero_equipos and active.max_por_equipo:
                concurso_capacidad = active.numero_equipos * active.max_por_equipo
            else:
                concurso_capacidad = None
    else:
        concurso_capacidad = None
    # Listado embebido con datos completos de participantes
    participantes_rows = []
    extra_fields = []
    if active:
        inscripciones = list(active.inscripciones.select_related("user", "performance_level"))
        from .models import ExtraField, UserExtraFieldValue
        extra_fields = list(
            ExtraField.objects.filter(
                (Q(congreso=congreso) | Q(congreso__isnull=True)),
                active=True,
                section__in=["registro", "perfil"],
                role_scope__in=["both", "participante"],
            ).order_by("order", "name")
        )
        user_ids = [ins.user_id for ins in inscripciones] or [0]
        field_ids = [f.id for f in extra_fields] or [0]
        values_qs = UserExtraFieldValue.objects.filter(
            (Q(congreso=congreso) | Q(congreso__isnull=True)),
            user_id__in=user_ids,
            field_id__in=field_ids,
        ).values("user_id", "field_id", "value")
        extra_values_by_user = {}
        for v in values_qs:
            d = extra_values_by_user.setdefault(v["user_id"], {})
            d[v["field_id"]] = v["value"]
        for ins in inscripciones:
            ordered = [extra_values_by_user.get(ins.user_id, {}).get(f.id, "") for f in extra_fields]
            participantes_rows.append({"ins": ins, "values_ordered": ordered})
    ctx = {
        "congreso": congreso,
        "concursos": list(concursos_qs),
        "concurso": active,
        "concurso_inscritos": concurso_inscritos,
        "concurso_capacidad": concurso_capacidad,
        "extra_fields": extra_fields,
        "participantes_rows": participantes_rows,
        "creation_mode": nuevo_flag,
    }
    return render(request, "instructores_vista/ins_concurso.html", ctx)


@login_required
@instructors_area_only
def instructor_conferencias_view(request):
    congreso = _get_active_instructor_congreso(request)
    if not congreso:
        return redirect("instructor_pendiente")
    conferencias_qs = Conferencia.objects.filter(congreso=congreso, instructor=request.user).order_by("-created_at")
    nuevo_flag = request.GET.get("nuevo") == "1"
    active_id = request.GET.get("id")
    active = None
    if active_id:
        try:
            active = conferencias_qs.get(pk=active_id)
        except Conferencia.DoesNotExist:
            active = None
    if not active and not nuevo_flag:
        active = conferencias_qs.first()
    # Crear
    if request.method == "POST" and request.POST.get("action") == "create_conferencia":
        title = (request.POST.get("title") or "").strip()
        lugar = (request.POST.get("lugar") or "").strip()
        description = (request.POST.get("descripcion") or request.POST.get("description") or "").strip()
        cupo = request.POST.get("cupo") or request.POST.get("cupo_maximo") or None
        if not title:
            messages.error(request, "El nombre de la conferencia es obligatorio.")
        else:
            obj = Conferencia.objects.create(
                congreso=congreso,
                title=title,
                lugar=lugar,
                description=description,
                cupo_maximo=cupo or None,
                instructor=request.user,
            )
            image = request.FILES.get("image")
            if image:
                if image.size > 3 * 1024 * 1024:
                    messages.warning(request, "La imagen supera 3MB, se ignoró.")
                else:
                    obj.image = image
                    obj.save()
            messages.success(request, "Conferencia creada correctamente.")
            return redirect("ins_conferencias")
    # Eliminar activo
    if request.method == "POST" and request.POST.get("action") == "delete_conferencia" and active:
        if request.user.is_superuser or request.user.id == active.instructor_id or request.user.id == congreso.admin_user_id:
            active.delete()
            messages.success(request, "Conferencia eliminada correctamente.")
        else:
            messages.error(request, "No tienes permisos para eliminar esta conferencia.")
        return redirect("ins_conferencias")
    conferencia_inscritos = active.inscripciones.count() if active else 0
    conferencia_capacidad = active.cupo_maximo if active and active.cupo_maximo else None
    # Listado embebido con datos completos de participantes
    participantes_rows = []
    extra_fields = []
    if active:
        inscripciones = list(active.inscripciones.select_related("user", "performance_level"))
        from .models import ExtraField, UserExtraFieldValue
        extra_fields = list(
            ExtraField.objects.filter(
                (Q(congreso=congreso) | Q(congreso__isnull=True)),
                active=True,
                section__in=["registro", "perfil"],
                role_scope__in=["both", "participante"],
            ).order_by("order", "name")
        )
        user_ids = [ins.user_id for ins in inscripciones] or [0]
        field_ids = [f.id for f in extra_fields] or [0]
        values_qs = UserExtraFieldValue.objects.filter(
            (Q(congreso=congreso) | Q(congreso__isnull=True)),
            user_id__in=user_ids,
            field_id__in=field_ids,
        ).values("user_id", "field_id", "value")
        extra_values_by_user = {}
        for v in values_qs:
            d = extra_values_by_user.setdefault(v["user_id"], {})
            d[v["field_id"]] = v["value"]
        for ins in inscripciones:
            ordered = [extra_values_by_user.get(ins.user_id, {}).get(f.id, "") for f in extra_fields]
            participantes_rows.append({"ins": ins, "values_ordered": ordered})
    ctx = {
        "congreso": congreso,
        "conferencias": list(conferencias_qs),
        "conferencia": active,
        "conferencia_inscritos": conferencia_inscritos,
        "conferencia_capacidad": conferencia_capacidad,
        "extra_fields": extra_fields,
        "participantes_rows": participantes_rows,
        "creation_mode": nuevo_flag,
    }
    return render(request, "instructores_vista/ins_conferencia.html", ctx)


# -------------------
# INSTRUCTOR EDIT PAGES (owner style) for Taller/Concurso/Conferencia
# -------------------
@login_required
@instructors_area_only
def instructor_taller_editar_view(request):
    congreso = _get_active_instructor_congreso(request)
    if not congreso:
        return redirect("instructor_pendiente")
    taller = Taller.objects.filter(congreso=congreso, instructor=request.user).order_by("-created_at").first()
    if not taller:
        messages.info(request, "No tienes un taller creado aún.")
        return redirect("ins_talleres")
    # Permisos: dueño o admin del congreso
    if not (request.user.id == taller.instructor_id or _user_is_congreso_admin(request.user, congreso)):
        messages.error(request, "No tienes permisos para editar este taller.")
        return redirect("ins_talleres")
    if request.method == "POST":
        title = (request.POST.get("title") or "").strip()
        lugar = (request.POST.get("lugar") or "").strip()
        cupo = request.POST.get("cupo")
        descripcion = (request.POST.get("descripcion") or "").strip()
        image = request.FILES.get("image")
        has_error = False
        if not title:
            messages.error(request, "El nombre del taller es obligatorio.")
            has_error = True
        if image and isinstance(image, UploadedFile) and image.size > 3 * 1024 * 1024:
            messages.error(request, "La imagen excede el tamaño máximo de 3MB.")
            has_error = True
        cupo_val = None
        if cupo:
            try:
                cupo_val = max(0, int(cupo))
            except ValueError:
                messages.error(request, "Cupo debe ser un número entero.")
                has_error = True
        if not has_error:
            taller.title = title
            taller.lugar = lugar
            taller.cupo_maximo = cupo_val
            taller.description = descripcion
            if image:
                taller.image = image
            taller.save()
            messages.success(request, "Taller actualizado.")
            return redirect("ins_talleres")
    ctx = {"congreso": congreso, "taller": taller}
    return render(request, "instructores_vista/ins_taller_editar.html", ctx)


@login_required
@instructors_area_only
def instructor_concurso_editar_view(request):
    congreso = _get_active_instructor_congreso(request)
    if not congreso:
        return redirect("instructor_pendiente")
    concurso = Concurso.objects.filter(congreso=congreso, instructor=request.user).order_by("-created_at").first()
    if not concurso:
        messages.info(request, "No tienes un concurso creado aún.")
        return redirect("ins_concursos")
    if not (request.user.id == concurso.instructor_id or _user_is_congreso_admin(request.user, congreso)):
        messages.error(request, "No tienes permisos para editar este concurso.")
        return redirect("ins_concursos")
    if request.method == "POST":
        tipo = (request.POST.get("tipo") or concurso.type or "individual").strip()
        title = (request.POST.get("title") or "").strip()
        lugar = (request.POST.get("lugar") or "").strip()
        descripcion = (request.POST.get("descripcion") or "").strip()
        image = request.FILES.get("image")
        cupo = request.POST.get("cupo")
        numero_equipos = request.POST.get("numero_equipos")
        max_por_equipo = request.POST.get("max_por_equipo")
        has_error = False
        if tipo not in ("individual", "grupal"):
            tipo = "individual"
        if not title:
            messages.error(request, "El nombre del concurso es obligatorio.")
            has_error = True
        if image and isinstance(image, UploadedFile) and image.size > 3 * 1024 * 1024:
            messages.error(request, "La imagen excede el tamaño máximo de 3MB.")
            has_error = True
        cupo_val = None
        num_eq_val = None
        max_eq_val = None
        if tipo == "individual":
            if cupo:
                try:
                    cupo_val = max(0, int(cupo))
                except ValueError:
                    messages.error(request, "Cupo debe ser un número entero.")
                    has_error = True
        else:
            if numero_equipos:
                try:
                    num_eq_val = max(0, int(numero_equipos))
                except ValueError:
                    messages.error(request, "Número de equipos debe ser entero.")
                    has_error = True
            if max_por_equipo:
                try:
                    max_eq_val = max(0, int(max_por_equipo))
                except ValueError:
                    messages.error(request, "Número máximo por equipo debe ser entero.")
                    has_error = True
        if not has_error:
            concurso.type = tipo
            concurso.title = title
            concurso.lugar = lugar
            concurso.description = descripcion
            concurso.cupo_maximo = cupo_val if tipo == "individual" else None
            concurso.numero_equipos = num_eq_val if tipo == "grupal" else None
            concurso.max_por_equipo = max_eq_val if tipo == "grupal" else None
            if image:
                concurso.image = image
            concurso.save()
            messages.success(request, "Concurso actualizado.", extra_tags="concurso")
            return redirect("ins_concursos")
    ctx = {"congreso": congreso, "concurso": concurso}
    return render(request, "instructores_vista/ins_concurso_editar.html", ctx)


@login_required
@instructors_area_only
def instructor_conferencia_editar_view(request):
    congreso = _get_active_instructor_congreso(request)
    if not congreso:
        return redirect("instructor_pendiente")
    conferencia = Conferencia.objects.filter(congreso=congreso, instructor=request.user).order_by("-created_at").first()
    if not conferencia:
        messages.info(request, "No tienes una conferencia creada aún.")
        return redirect("ins_conferencias")
    if not (request.user.id == conferencia.instructor_id or _user_is_congreso_admin(request.user, congreso)):
        messages.error(request, "No tienes permisos para editar esta conferencia.")
        return redirect("ins_conferencias")
    if request.method == "POST":
        title = (request.POST.get("title") or "").strip()
        lugar = (request.POST.get("lugar") or "").strip()
        descripcion = (request.POST.get("descripcion") or "").strip()
        cupo = request.POST.get("cupo") or request.POST.get("cupo_maximo") or None
        image = request.FILES.get("image")
        has_error = False
        if not title:
            messages.error(request, "El nombre de la conferencia es obligatorio.")
            has_error = True
        if image and isinstance(image, UploadedFile) and image.size > 3 * 1024 * 1024:
            messages.error(request, "La imagen excede el tamaño máximo de 3MB.")
            has_error = True
        if not has_error:
            conferencia.title = title
            conferencia.lugar = lugar
            conferencia.description = descripcion
            conferencia.cupo_maximo = cupo or None
            if image:
                conferencia.image = image
            conferencia.save()
            messages.success(request, "Conferencia actualizada.")
            return redirect("ins_conferencias")
    ctx = {"congreso": congreso, "conferencia": conferencia}
    return render(request, "instructores_vista/ins_conferencia_editar.html", ctx)

# -------------------
# INSTRUCTOR PARTICIPANTES PAGES
# -------------------
@login_required
@instructors_area_only
def instructor_taller_participantes_view(request):
    congreso = _get_active_instructor_congreso(request)
    if not congreso:
        return redirect("instructor_pendiente")
    # Permitir elegir un taller específico vía ?id=, con fallback al último creado
    taller_id = request.GET.get("id")
    qs_taller = Taller.objects.filter(congreso=congreso, instructor=request.user)
    if taller_id:
        taller = qs_taller.filter(pk=taller_id).first()
    else:
        taller = qs_taller.order_by("-created_at").first()
    if not taller:
        messages.info(request, "No tienes un taller creado aún.")
        return redirect("ins_talleres")
    inscripciones = list(taller.inscripciones.select_related("user", "performance_level"))
    puede_remover = request.user.is_superuser or request.user.id == congreso.admin_user_id or request.user.id == taller.instructor_id
    # Campos extra dinámicos del participante (registro/perfil, global + congreso)
    extra_fields = list(
        ExtraField.objects.filter(
            (Q(congreso=congreso) | Q(congreso__isnull=True)),
            active=True,
            section__in=["registro", "perfil"],
            role_scope__in=["both", "participante"],
        ).order_by("order", "name")
    )
    user_ids = [ins.user_id for ins in inscripciones] or [0]
    field_ids = [f.id for f in extra_fields] or [0]
    values_qs = UserExtraFieldValue.objects.filter(
        (Q(congreso=congreso) | Q(congreso__isnull=True)),
        user_id__in=user_ids,
        field_id__in=field_ids,
    ).values("user_id", "field_id", "value")
    extra_values_by_user = {}
    for v in values_qs:
        d = extra_values_by_user.setdefault(v["user_id"], {})
        d[v["field_id"]] = v["value"]
    if request.method == "POST" and puede_remover and request.POST.get("action") == "remove":
        insc_id = request.POST.get("insc_id")
        try:
            ins = taller.inscripciones.get(pk=insc_id)
            ins.delete()
            messages.success(request, "Participante removido del taller.")
            if taller_id:
                return redirect(f"{reverse('ins_taller_participantes')}?id={taller.id}")
            return redirect("ins_taller_participantes")
        except Exception:
            messages.error(request, "No se pudo remover la inscripción.")
    # Construir filas con valores ordenados para no depender de filtros en plantilla (siempre para GET y tras errores POST)
    participantes_rows = []
    for ins in inscripciones:
        ordered = []
        for f in extra_fields:
            ordered.append(extra_values_by_user.get(ins.user_id, {}).get(f.id, ""))
        participantes_rows.append({"ins": ins, "values_ordered": ordered})
    ctx = {
        "congreso": congreso,
        "taller": taller,
        "participantes_rows": participantes_rows,
        "puede_remover": puede_remover,
        "extra_fields": extra_fields,
    }
    return render(request, "instructores_vista/ins_taller_participantes.html", ctx)

@login_required
@instructors_area_only
def instructor_concurso_participantes_view(request):
    congreso = _get_active_instructor_congreso(request)
    if not congreso:
        return redirect("instructor_pendiente")
    concurso_id = request.GET.get("id")
    qs_conc = Concurso.objects.filter(congreso=congreso, instructor=request.user)
    if concurso_id:
        concurso = qs_conc.filter(pk=concurso_id).first()
    else:
        concurso = qs_conc.order_by("-created_at").first()
    if not concurso:
        messages.info(request, "No tienes un concurso creado aún.")
        return redirect("ins_concursos")
    inscripciones = list(concurso.inscripciones.select_related("user", "performance_level"))
    puede_remover = bool(concurso) and (request.user.is_superuser or request.user.id == congreso.admin_user_id or request.user.id == concurso.instructor_id)
    extra_fields = list(
        ExtraField.objects.filter(
            (Q(congreso=congreso) | Q(congreso__isnull=True)),
            active=True,
            section__in=["registro", "perfil"],
            role_scope__in=["both", "participante"],
        ).order_by("order", "name")
    )
    user_ids = [ins.user_id for ins in inscripciones] or [0]
    field_ids = [f.id for f in extra_fields] or [0]
    values_qs = UserExtraFieldValue.objects.filter(
        (Q(congreso=congreso) | Q(congreso__isnull=True)),
        user_id__in=user_ids,
        field_id__in=field_ids,
    ).values("user_id", "field_id", "value")
    extra_values_by_user = {}
    for v in values_qs:
        d = extra_values_by_user.setdefault(v["user_id"], {})
        d[v["field_id"]] = v["value"]
    if request.method == "POST" and puede_remover and request.POST.get("action") == "remove":
        insc_id = request.POST.get("insc_id")
        try:
            ins = concurso.inscripciones.get(pk=insc_id)
            ins.delete()
            messages.success(request, "Participante removido del concurso.")
            return redirect("ins_concurso_participantes")
        except Exception:
            messages.error(request, "No se pudo remover la inscripción.")
    participantes_rows = []
    for ins in inscripciones:
        ordered = []
        for f in extra_fields:
            ordered.append(extra_values_by_user.get(ins.user_id, {}).get(f.id, ""))
        participantes_rows.append({"ins": ins, "values_ordered": ordered})
    ctx = {
        "congreso": congreso,
        "concurso": concurso,
        "inscripciones": inscripciones,
        "participantes_rows": participantes_rows,
        "puede_remover": puede_remover,
        "extra_fields": extra_fields,
    }
    return render(request, "instructores_vista/ins_concurso_participantes.html", ctx)

@login_required
@instructors_area_only
def instructor_conferencia_participantes_view(request):
    congreso = _get_active_instructor_congreso(request)
    if not congreso:
        return redirect("instructor_pendiente")
    conferencia_id = request.GET.get("id")
    qs_conf = Conferencia.objects.filter(congreso=congreso, instructor=request.user)
    if conferencia_id:
        conferencia = qs_conf.filter(pk=conferencia_id).first()
    else:
        conferencia = qs_conf.order_by("-created_at").first()
    if not conferencia:
        messages.info(request, "No tienes una conferencia creada aún.")
        return redirect("ins_conferencias")
    inscripciones = list(conferencia.inscripciones.select_related("user", "performance_level"))
    puede_remover = bool(conferencia) and (request.user.is_superuser or request.user.id == congreso.admin_user_id or request.user.id == conferencia.instructor_id)
    extra_fields = list(
        ExtraField.objects.filter(
            (Q(congreso=congreso) | Q(congreso__isnull=True)),
            active=True,
            section__in=["registro", "perfil"],
            role_scope__in=["both", "participante"],
        ).order_by("order", "name")
    )
    user_ids = [ins.user_id for ins in inscripciones] or [0]
    field_ids = [f.id for f in extra_fields] or [0]
    values_qs = UserExtraFieldValue.objects.filter(
        (Q(congreso=congreso) | Q(congreso__isnull=True)),
        user_id__in=user_ids,
        field_id__in=field_ids,
    ).values("user_id", "field_id", "value")
    extra_values_by_user = {}
    for v in values_qs:
        d = extra_values_by_user.setdefault(v["user_id"], {})
        d[v["field_id"]] = v["value"]
    if request.method == "POST" and puede_remover and request.POST.get("action") == "remove":
        insc_id = request.POST.get("insc_id")
        try:
            ins = conferencia.inscripciones.get(pk=insc_id)
            ins.delete()
            messages.success(request, "Participante removido de la conferencia.")
            return redirect("ins_conferencia_participantes")
        except Exception:
            messages.error(request, "No se pudo remover la inscripción.")
    participantes_rows = []
    for ins in inscripciones:
        ordered = []
        for f in extra_fields:
            ordered.append(extra_values_by_user.get(ins.user_id, {}).get(f.id, ""))
        participantes_rows.append({"ins": ins, "values_ordered": ordered})
    ctx = {
        "congreso": congreso,
        "conferencia": conferencia,
        "inscripciones": inscripciones,
        "participantes_rows": participantes_rows,
        "puede_remover": puede_remover,
        "extra_fields": extra_fields,
    }
    return render(request, "instructores_vista/ins_conferencia_participantes.html", ctx)
@login_required
def usuario_participante_nuevo_view(request, congreso_id: int):
    deny = _deny_non_admin_roles(request)
    if deny: return deny
    try:
        congreso = Congreso.objects.get(pk=congreso_id)
    except Congreso.DoesNotExist:
        messages.error(request, "El congreso solicitado no existe.")
        return redirect("congresos_eventos")

    # Enforzar scope: si es Administradores_Congresos con scope y está intentando otro congreso
    scoped_congreso = _get_scoped_congreso_for_group_admin(request.user)
    if scoped_congreso and scoped_congreso.id != congreso.id:
        messages.error(
            request,
            f"Tu cuenta está limitada al congreso '{scoped_congreso.name}'. No puedes acceder a '{congreso.name}'."
        )
        return redirect("inicio_congreso", scoped_congreso.id)
    guard = _ensure_congreso_access_or_redirect(request, congreso)
    if guard:
        return guard
    if not _user_is_congreso_admin(request.user, congreso):
        messages.error(request, "No tienes permisos para agregar participantes en este congreso.")
        return redirect("participantes", congreso.id)

    # Niveles disponibles y campos extra (sección registro) para participantes
    levels = list(PerformanceLevel.objects.filter(congreso=congreso).order_by("name"))
    extra_fields = list(
        ExtraField.objects.filter(
            (Q(congreso=congreso) | Q(congreso__isnull=True)),
            active=True,
            section="registro",
            role_scope__in=["both", "participante"],
        ).order_by("order", "name")
    )

    if request.method == "POST":
        first_name = (request.POST.get("first_name") or "").strip()
        apellido_paterno = (request.POST.get("apellido_paterno") or "").strip()
        apellido_materno = (request.POST.get("apellido_materno") or "").strip()
        email = (request.POST.get("email") or "").strip()
        password = request.POST.get("password")
        congreso_password = request.POST.get("congreso_password")
        performance_level_id = request.POST.get("performance_level_id")

        # Validar email
        try:
            validate_email(email or "")
        except ValidationError:
            messages.error(request, "Ingresa un correo electrónico válido.")
            return render(request, "usuarios_accion/participante.html", {
                "congreso": congreso, "levels": levels, "extra_fields": extra_fields,
                "rol": "Administrador",
            })

        # Validar contraseña del congreso si aplica
        if congreso.has_access_password():
            if not congreso_password or not congreso.check_access_password(congreso_password):
                messages.error(request, "La contraseña de acceso del congreso es incorrecta.")
                return render(request, "usuarios_accion/participante.html", {
                    "congreso": congreso, "levels": levels, "extra_fields": extra_fields,
                    "rol": "Administrador",
                })

        # Dominios permitidos
        dominios = list(AllowedEmailDomain.objects.filter(congreso=congreso).values_list("domain", flat=True))
        if dominios:
            email_lc = email.lower().strip()
            if not any(email_lc.endswith(d) for d in dominios):
                pretty = ", ".join(sorted(dominios))
                messages.error(request, f"Este correo no pertenece a un dominio permitido. Dominios: {pretty}")
                return render(request, "usuarios_accion/participante.html", {
                    "congreso": congreso, "levels": levels, "extra_fields": extra_fields,
                    "rol": "Administrador",
                })

        # Política de contraseña
        specials = set(".!%&$}")
        if not password or len(password) < 8 or len(password) > 16 or not any(c.isupper() for c in password) or not any(c.islower() for c in password) or not any(c in specials for c in password):
            messages.error(request, "La contraseña debe tener 8-16 caracteres e incluir mayúscula, minúscula y símbolo especial (.!%&$}).")
            return render(request, "usuarios_accion/participante.html", {
                "congreso": congreso, "levels": levels, "extra_fields": extra_fields,
                "rol": "Administrador",
            })

        # Validar y recolectar extras
        extra_values = {}
        for f in extra_fields:
            key = f"extra_{f.id}"
            raw = (request.POST.get(key) or "").strip()
            if f.required and not raw and f.field_type != "boolean":
                messages.error(request, f"El campo '{f.name}' es obligatorio.")
                return render(request, "usuarios_accion/participante.html", {"congreso": congreso, "levels": levels, "extra_fields": extra_fields, "rol": "Administrador"})
            if raw:
                if f.field_type == "email":
                    try:
                        validate_email(raw)
                    except ValidationError:
                        messages.error(request, f"Ingresa un correo válido en '{f.name}'.")
                        return render(request, "usuarios_accion/participante.html", {"congreso": congreso, "levels": levels, "extra_fields": extra_fields, "rol": "Administrador"})
                elif f.field_type == "number":
                    try:
                        float(raw)
                    except ValueError:
                        messages.error(request, f"Ingresa un número válido en '{f.name}'.")
                        return render(request, "usuarios_accion/participante.html", {"congreso": congreso, "levels": levels, "extra_fields": extra_fields, "rol": "Administrador"})
                elif f.field_type == "select":
                    choices = f.get_choices()
                    if choices and raw not in choices:
                        messages.error(request, f"Selecciona una opción válida para '{f.name}'.")
                        return render(request, "usuarios_accion/participante.html", {"congreso": congreso, "levels": levels, "extra_fields": extra_fields, "rol": "Administrador"})
            extra_values[f.id] = raw

        # Nivel de desempeño (si existen niveles, es requerido)
        selected_level = None
        if levels:
            try:
                selected_level = PerformanceLevel.objects.get(id=performance_level_id, congreso=congreso)
            except (PerformanceLevel.DoesNotExist, ValueError, TypeError):
                messages.error(request, "Selecciona un nivel de desempeño válido.")
                return render(request, "usuarios_accion/participante.html", {"congreso": congreso, "levels": levels, "extra_fields": extra_fields, "rol": "Administrador"})

        # Crear/usar usuario
        existing = (
            User.objects.filter(Q(username__iexact=email) | Q(email__iexact=email)).order_by("-last_login", "-id").first()
        )
        if existing:
            user = authenticate(request, username=existing.username, password=password)
            if not user:
                messages.error(request, "Ya existe una cuenta con ese correo. Verifica la contraseña." )
                return render(request, "usuarios_accion/participante.html", {"congreso": congreso, "levels": levels, "extra_fields": extra_fields, "rol": "Administrador"})
            if first_name and not user.first_name:
                user.first_name = first_name
            full_last = f"{apellido_paterno} {apellido_materno}".strip()
            if full_last and not user.last_name:
                user.last_name = full_last
            if email:
                if user.username.lower() != email.lower() and not User.objects.exclude(pk=user.pk).filter(username__iexact=email).exists():
                    user.username = email
                user.email = email
            user.save()
        else:
            user = User.objects.create_user(
                username=email,
                email=email,
                password=password,
                first_name=first_name,
                last_name=f"{apellido_paterno} {apellido_materno}"
            )
            # Grupo participante
            try:
                grupo = Group.objects.get(name="Participantes")
                user.groups.add(grupo)
            except Group.DoesNotExist:
                pass

        # Crear membresía aprobada
        m, _ = UserCongresoMembership.objects.get_or_create(
            user=user,
            congreso=congreso,
            defaults={
                "role": "participante",
                "status": "approved",
                "performance_level": selected_level,
                "decided_at": timezone.now(),
                "decided_by": request.user,
            },
        )
        if m.status != "approved" or m.role != "participante" or m.performance_level_id != (selected_level.id if selected_level else None):
            m.role = "participante"
            m.status = "approved"
            m.performance_level = selected_level
            m.decided_at = timezone.now()
            m.decided_by = request.user
            m.save()

        # Guardar extras
        for fid, val in extra_values.items():
            try:
                fobj = next((x for x in extra_fields if x.id == fid), None)
                if not fobj:
                    continue
                UserExtraFieldValue.objects.update_or_create(
                    user=user,
                    congreso=congreso if fobj.congreso_id else None,
                    field=fobj,
                    defaults={"value": val},
                )
            except Exception:
                pass

        messages.success(request, "Participante creado y activado.")
        return redirect("participantes", congreso.id)

    grupos = request.user.groups.values_list("name", flat=True)
    rol = "Administrador" if (request.user.is_superuser or "Administrador" in grupos) else ("Instructor" if "Instructores" in grupos else "Participante")

    ctx = {"rol": rol, "congreso": congreso, "levels": levels, "extra_fields": extra_fields}
    return render(request, "usuarios_accion/participante.html", ctx)


@login_required
def usuario_instructor_nuevo_view(request, congreso_id: int):
    deny = _deny_non_admin_roles(request)
    if deny: return deny
    try:
        congreso = Congreso.objects.get(pk=congreso_id)
    except Congreso.DoesNotExist:
        messages.error(request, "El congreso solicitado no existe.")
        return redirect("congresos_eventos")
    guard = _ensure_congreso_access_or_redirect(request, congreso)
    if guard:
        return guard
    if not _user_is_congreso_admin(request.user, congreso):
        messages.error(request, "No tienes permisos para agregar instructores en este congreso.")
        return redirect("instructores", congreso.id)

    extra_fields = list(
        ExtraField.objects.filter(
            (Q(congreso=congreso) | Q(congreso__isnull=True)),
            active=True,
            section="registro",
            role_scope__in=["both", "instructor"],
        ).order_by("order", "name")
    )

    if request.method == "POST":
        first_name = (request.POST.get("first_name") or "").strip()
        apellido_paterno = (request.POST.get("apellido_paterno") or "").strip()
        apellido_materno = (request.POST.get("apellido_materno") or "").strip()
        email = (request.POST.get("email") or "").strip()
        password = request.POST.get("password")
        congreso_password = request.POST.get("congreso_password")

        try:
            validate_email(email or "")
        except ValidationError:
            messages.error(request, "Ingresa un correo electrónico válido.")
            return render(request, "usuarios_accion/instructor.html", {"congreso": congreso, "extra_fields": extra_fields, "rol": "Administrador"})

        if congreso.has_access_password():
            if not congreso_password or not congreso.check_access_password(congreso_password):
                messages.error(request, "La contraseña de acceso del congreso es incorrecta.")
                return render(request, "usuarios_accion/instructor.html", {"congreso": congreso, "extra_fields": extra_fields, "rol": "Administrador"})

        dominios = list(AllowedEmailDomain.objects.filter(congreso=congreso).values_list("domain", flat=True))
        if dominios:
            email_lc = email.lower().strip()
            if not any(email_lc.endswith(d) for d in dominios):
                pretty = ", ".join(sorted(dominios))
                messages.error(request, f"Este correo no pertenece a un dominio permitido. Dominios: {pretty}")
                return render(request, "usuarios_accion/instructor.html", {"congreso": congreso, "extra_fields": extra_fields, "rol": "Administrador"})

        specials = set(".!%&$}")
        if not password or len(password) < 8 or len(password) > 16 or not any(c.isupper() for c in password) or not any(c.islower() for c in password) or not any(c in specials for c in password):
            messages.error(request, "La contraseña debe tener 8-16 caracteres e incluir mayúscula, minúscula y símbolo especial (.!%&$}).")
            return render(request, "usuarios_accion/instructor.html", {"congreso": congreso, "extra_fields": extra_fields, "rol": "Administrador"})

        extra_values = {}
        for f in extra_fields:
            key = f"extra_{f.id}"
            raw = (request.POST.get(key) or "").strip()
            if f.required and not raw and f.field_type != "boolean":
                messages.error(request, f"El campo '{f.name}' es obligatorio.")
                return render(request, "usuarios_accion/instructor.html", {"congreso": congreso, "extra_fields": extra_fields, "rol": "Administrador"})
            if raw:
                if f.field_type == "email":
                    try:
                        validate_email(raw)
                    except ValidationError:
                        messages.error(request, f"Ingresa un correo válido en '{f.name}'.")
                        return render(request, "usuarios_accion/instructor.html", {"congreso": congreso, "extra_fields": extra_fields, "rol": "Administrador"})
                elif f.field_type == "number":
                    try:
                        float(raw)
                    except ValueError:
                        messages.error(request, f"Ingresa un número válido en '{f.name}'.")
                        return render(request, "usuarios_accion/instructor.html", {"congreso": congreso, "extra_fields": extra_fields, "rol": "Administrador"})
                elif f.field_type == "select":
                    choices = f.get_choices()
                    if choices and raw not in choices:
                        messages.error(request, f"Selecciona una opción válida para '{f.name}'.")
                        return render(request, "usuarios_accion/instructor.html", {"congreso": congreso, "extra_fields": extra_fields, "rol": "Administrador"})
            extra_values[f.id] = raw

        existing = (
            User.objects.filter(Q(username__iexact=email) | Q(email__iexact=email)).order_by("-last_login", "-id").first()
        )
        if existing:
            user = authenticate(request, username=existing.username, password=password)
            if not user:
                messages.error(request, "Ya existe una cuenta con ese correo. Verifica la contraseña.")
                return render(request, "usuarios_accion/instructor.html", {"congreso": congreso, "extra_fields": extra_fields, "rol": "Administrador"})
            if first_name and not user.first_name:
                user.first_name = first_name
            full_last = f"{apellido_paterno} {apellido_materno}".strip()
            if full_last and not user.last_name:
                user.last_name = full_last
            if email:
                if user.username.lower() != email.lower() and not User.objects.exclude(pk=user.pk).filter(username__iexact=email).exists():
                    user.username = email
                user.email = email
            user.save()
        else:
            user = User.objects.create_user(
                username=email,
                email=email,
                password=password,
                first_name=first_name,
                last_name=f"{apellido_paterno} {apellido_materno}"
            )
            try:
                grupo = Group.objects.get(name="Instructores")
                user.groups.add(grupo)
            except Group.DoesNotExist:
                pass

        m, _ = UserCongresoMembership.objects.get_or_create(
            user=user,
            congreso=congreso,
            defaults={
                "role": "instructor",
                "status": "approved",
                "decided_at": timezone.now(),
                "decided_by": request.user,
            },
        )
        if m.status != "approved" or m.role != "instructor":
            m.role = "instructor"
            m.status = "approved"
            m.decided_at = timezone.now()
            m.decided_by = request.user
            m.save()

        for fid, val in extra_values.items():
            try:
                fobj = next((x for x in extra_fields if x.id == fid), None)
                if not fobj:
                    continue
                UserExtraFieldValue.objects.update_or_create(
                    user=user,
                    congreso=congreso if fobj.congreso_id else None,
                    field=fobj,
                    defaults={"value": val},
                )
            except Exception:
                pass

        messages.success(request, "Instructor creado y activado.")
        return redirect("instructores", congreso.id)

    grupos = request.user.groups.values_list("name", flat=True)
    rol = "Administrador" if (request.user.is_superuser or "Administrador" in grupos) else ("Instructor" if "Instructores" in grupos else "Participante")
    ctx = {"rol": rol, "congreso": congreso, "extra_fields": extra_fields}
    return render(request, "usuarios_accion/instructor.html", ctx)


@login_required
def usuario_editar_participante_view(request, congreso_id: int, membership_id: int):
    deny = _deny_non_admin_roles(request)
    if deny: return deny
    try:
        congreso = Congreso.objects.get(pk=congreso_id)
    except Congreso.DoesNotExist:
        messages.error(request, "El congreso solicitado no existe.")
        return redirect("congresos_eventos")

    membership = UserCongresoMembership.objects.filter(id=membership_id, congreso=congreso, role="participante").select_related("user", "performance_level").first()
    if not membership:
        messages.error(request, "El participante solicitado no existe.")
        return redirect("participantes", congreso.id)

    levels = list(PerformanceLevel.objects.filter(congreso=congreso).order_by("name"))
    extra_fields = list(
        ExtraField.objects.filter(
            (Q(congreso=congreso) | Q(congreso__isnull=True)),
            active=True,
            section="registro",
            role_scope__in=["both", "participante"],
        ).order_by("order", "name")
    )
    # Prefill extra values
    extra_values = {}
    vals = UserExtraFieldValue.objects.filter(user=membership.user, field__in=extra_fields).filter(Q(congreso=congreso) | Q(congreso__isnull=True))
    for v in vals:
        extra_values[v.field_id] = v.value

    if request.method == "POST":
        first_name = (request.POST.get("first_name") or "").strip()
        email = (request.POST.get("email") or "").strip()
        password = request.POST.get("password") or ""
        performance_level_id = request.POST.get("performance_level_id")

        try:
            validate_email(email or "")
        except ValidationError:
            messages.error(request, "Ingresa un correo electrónico válido.")
            # fall-through to render
        else:
            # dominio
            dominios = list(AllowedEmailDomain.objects.filter(congreso=congreso).values_list("domain", flat=True))
            if dominios:
                email_lc = email.lower().strip()
                if not any(email_lc.endswith(d) for d in dominios):
                    pretty = ", ".join(sorted(dominios))
                    messages.error(request, f"Este correo no pertenece a un dominio permitido. Dominios: {pretty}")
                else:
                    # actualizar usuario
                    user = membership.user
                    user.first_name = first_name or user.first_name
                    # username = email
                    if email and (user.username.lower() != email.lower()):
                        if not User.objects.exclude(pk=user.pk).filter(username__iexact=email).exists():
                            user.username = email
                    user.email = email
                    # actualizar password si se proporcionó (opcional)
                    if password:
                        specials = set(".!%&$}")
                        if len(password) < 8 or len(password) > 16 or not any(c.isupper() for c in password) or not any(c.islower() for c in password) or not any(c in specials for c in password):
                            messages.error(request, "La contraseña debe tener 8-16 caracteres e incluir mayúscula, minúscula y símbolo especial (.!%&$}).")
                        else:
                            user.set_password(password)
                    user.save()

                    # nivel
                    if levels:
                        try:
                            membership.performance_level = PerformanceLevel.objects.get(id=performance_level_id, congreso=congreso)
                        except (PerformanceLevel.DoesNotExist, ValueError, TypeError):
                            messages.error(request, "Selecciona un nivel de desempeño válido.")
                        else:
                            membership.save(update_fields=["performance_level"])

                    # extras
                    for f in extra_fields:
                        key = f"extra_{f.id}"
                        raw = (request.POST.get(key) or "").strip()
                        if f.required and not raw and f.field_type != "boolean":
                            messages.error(request, f"El campo '{f.name}' es obligatorio.")
                            continue
                        UserExtraFieldValue.objects.update_or_create(
                            user=membership.user,
                            congreso=congreso if f.congreso_id else None,
                            field=f,
                            defaults={"value": raw},
                        )

                    messages.success(request, "Participante actualizado.")
                    return redirect("participantes", congreso.id)

    grupos = request.user.groups.values_list("name", flat=True)
    rol = "Administrador" if (request.user.is_superuser or "Administrador" in grupos) else ("Instructor" if "Instructores" in grupos else "Participante")
    ctx = {"rol": rol, "congreso": congreso, "membership": membership, "levels": levels, "extra_fields": extra_fields, "extra_values": extra_values}
    return render(request, "usuarios_accion/editar_participante.html", ctx)


@login_required
def usuario_editar_instructores_view(request, congreso_id: int, membership_id: int):
    deny = _deny_non_admin_roles(request)
    if deny: return deny
    try:
        congreso = Congreso.objects.get(pk=congreso_id)
    except Congreso.DoesNotExist:
        messages.error(request, "El congreso solicitado no existe.")
        return redirect("congresos_eventos")

    membership = UserCongresoMembership.objects.filter(id=membership_id, congreso=congreso, role="instructor").select_related("user").first()
    if not membership:
        messages.error(request, "El instructor solicitado no existe.")
        return redirect("instructores", congreso.id)

    extra_fields = list(
        ExtraField.objects.filter(
            (Q(congreso=congreso) | Q(congreso__isnull=True)),
            active=True,
            section="registro",
            role_scope__in=["both", "instructor"],
        ).order_by("order", "name")
    )
    extra_values = {}
    vals = UserExtraFieldValue.objects.filter(user=membership.user, field__in=extra_fields).filter(Q(congreso=congreso) | Q(congreso__isnull=True))
    for v in vals:
        extra_values[v.field_id] = v.value

    if request.method == "POST":
        first_name = (request.POST.get("first_name") or "").strip()
        email = (request.POST.get("email") or "").strip()
        password = request.POST.get("password") or ""

        try:
            validate_email(email or "")
        except ValidationError:
            messages.error(request, "Ingresa un correo electrónico válido.")
        else:
            dominios = list(AllowedEmailDomain.objects.filter(congreso=congreso).values_list("domain", flat=True))
            if dominios:
                email_lc = email.lower().strip()
                if not any(email_lc.endswith(d) for d in dominios):
                    pretty = ", ".join(sorted(dominios))
                    messages.error(request, f"Este correo no pertenece a un dominio permitido. Dominios: {pretty}")
                else:
                    user = membership.user
                    user.first_name = first_name or user.first_name
                    if email and (user.username.lower() != email.lower()):
                        if not User.objects.exclude(pk=user.pk).filter(username__iexact=email).exists():
                            user.username = email
                    user.email = email
                    if password:
                        specials = set(".!%&$}")
                        if len(password) < 8 or len(password) > 16 or not any(c.isupper() for c in password) or not any(c.islower() for c in password) or not any(c in specials for c in password):
                            messages.error(request, "La contraseña debe tener 8-16 caracteres e incluir mayúscula, minúscula y símbolo especial (.!%&$}).")
                        else:
                            user.set_password(password)
                    user.save()

                    for f in extra_fields:
                        key = f"extra_{f.id}"
                        raw = (request.POST.get(key) or "").strip()
                        if f.required and not raw and f.field_type != "boolean":
                            messages.error(request, f"El campo '{f.name}' es obligatorio.")
                            continue
                        UserExtraFieldValue.objects.update_or_create(
                            user=membership.user,
                            congreso=congreso if f.congreso_id else None,
                            field=f,
                            defaults={"value": raw},
                        )
                    messages.success(request, "Instructor actualizado.")
                    return redirect("instructores", congreso.id)

    grupos = request.user.groups.values_list("name", flat=True)
    rol = "Administrador" if (request.user.is_superuser or "Administrador" in grupos) else ("Instructor" if "Instructores" in grupos else "Participante")
    ctx = {"rol": rol, "congreso": congreso, "membership": membership, "extra_fields": extra_fields, "extra_values": extra_values}
    return render(request, "usuarios_accion/editar_participantes.html", ctx)




@login_required
def congresos_eventos_view(request):
    deny = _deny_non_admin_roles(request)
    if deny: return deny
    # Bloquear listado general para administradores de congresos con scope
    scoped = _get_scoped_congreso_for_group_admin(request.user)
    if scoped:
        messages.info(request, "Tu acceso está limitado a un congreso. Redirigido a su panel.")
        return redirect("inicio_congreso", scoped.id)
    grupos = request.user.groups.values_list("name", flat=True)
    if "Instructores" in grupos:
        rol = "Instructor"
    elif "Participantes" in grupos:
        rol = "Participante"
    else:
        rol = "Administrador"

    # Asegurar al menos un congreso base para comenzar
    if not Congreso.objects.exists():
        Congreso.objects.create(name="Congreso General")

    # Acciones (eliminar congreso)
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "delete_congreso":
            cid = request.POST.get("congreso_id")
            if cid:
                Congreso.objects.filter(id=cid).delete()
                messages.success(request, "Congreso eliminado.")
            return HttpResponseRedirect(reverse("congresos_eventos"))

    congresos = list(Congreso.objects.all().order_by("name"))
    return render(request, "congresos_eventos.html", {"rol": rol, "congresos": congresos, "is_scoped_admin": False})


@login_required
def congreso_evento_view(request):
    deny = _deny_non_admin_roles(request)
    if deny: return deny
    """Crea un nuevo congreso/evento y, opcionalmente, un usuario administrador para ese congreso.

    Campos esperados:
    - logo (archivo, opcional, máx ~3MB)
    - nombre_congreso (texto)
    - descripcion (texto largo, opcional)
    - nombre_encargado (texto)
    - username (texto o email)
    - password (texto, 8-16, mayúscula+minúscula+símbolo especial .!%&$})
    - access_password (opcional, contraseña de acceso al registro del congreso)
    """
    grupos = request.user.groups.values_list("name", flat=True)
    if "Instructores" in grupos:
        rol = "Instructor"
    elif "Participantes" in grupos:
        rol = "Participante"
    else:
        rol = "Administrador"

    # Prohibir creación a administradores de congresos con scope (a menos que sean Administrador o superuser)
    if request.user.groups.filter(name="Administradores_Congresos").exists() and not (request.user.is_superuser or "Administrador" in grupos):
        scoped = _get_scoped_congreso_for_group_admin(request.user)
        messages.error(request, "No tienes permisos para crear nuevos congresos.")
        return redirect("inicio_congreso", scoped.id if scoped else "inicio")

    if request.method == "POST":
        nombre_congreso = (request.POST.get("nombre_congreso") or "").strip()
        descripcion = (request.POST.get("descripcion") or "").strip()
        nombre_encargado = (request.POST.get("nombre_encargado") or "").strip()
        username = (request.POST.get("username") or "").strip()
        password = request.POST.get("password") or ""
        logo_file = request.FILES.get("logo")
        access_password = (request.POST.get("access_password") or "").strip()

        # Validaciones
        has_error = False
        if not nombre_congreso:
            messages.error(request, "El nombre del congreso o evento no puede estar vacío.")
            has_error = True
        elif Congreso.objects.filter(name__iexact=nombre_congreso).exists():
            messages.error(request, "Ya existe un congreso o evento con ese nombre.")
            has_error = True

        if logo_file:
            # Límite aproximado 3MB
            max_bytes = 3 * 1024 * 1024
            if isinstance(logo_file, UploadedFile) and logo_file.size > max_bytes:
                messages.error(request, "El logo excede el tamaño máximo de 3MB.")
                has_error = True

        # Validaciones de usuario (si se llenan los 3 campos, intentamos crear usuario)
        creating_user = bool(nombre_encargado or username or password)
        if creating_user:
            if not (nombre_encargado and username and password):
                messages.error(request, "Para crear el usuario administrador del congreso, llena Nombre, Usuario y Contraseña.")
                has_error = True
            else:
                if User.objects.filter(username=username).exists():
                    messages.error(request, "El usuario indicado ya existe.")
                    has_error = True
                # Política de contraseña usada en el proyecto
                specials = set(".!%&$}")
                if len(password) < 8 or len(password) > 16 or not any(c.isupper() for c in password) or not any(c.islower() for c in password) or not any(c in specials for c in password):
                    messages.error(request, "La contraseña debe tener entre 8 y 16 caracteres, incluir mayúscula, minúscula y un símbolo especial (.!%&$}).")
                    has_error = True

        if not has_error:
            # Crear congreso
            congreso = Congreso(name=nombre_congreso, description=descripcion or None)
            if logo_file:
                congreso.logo = logo_file
            # Establecer contraseña de acceso si viene
            if access_password:
                congreso.set_access_password(access_password)
            congreso.save()

            scope_user = None
            # Crear usuario si corresponde
            if creating_user and nombre_encargado and username and password:
                user = User.objects.create_user(
                    username=username,
                    email=username if "@" in username else "",
                    password=password,
                    first_name=nombre_encargado,
                )
                group_name = "Administradores_Congresos"
                group, _ = Group.objects.get_or_create(name=group_name)
                user.groups.add(group)
                congreso.admin_user = user
                congreso.save(update_fields=["admin_user"])
                scope_user = user

            # Crear scope automático si hubo usuario admin
            if scope_user:
                try:
                    CongresoAdminScope.objects.get_or_create(user=scope_user, defaults={"congreso": congreso})
                except Exception:
                    pass

            messages.success(request, "Congreso o evento creado correctamente.")
            return redirect("congresos_eventos")

    return render(request, "congreso_evento.html", {"rol": rol, "is_scoped_admin": False})

@login_required
def editar_congreso_evento_view(request, congreso_id: int):
    deny = _deny_non_admin_roles(request)
    if deny: return deny
    """Edita un congreso/evento existente. Permite actualizar nombre y logo.

    Campos opcionales para usuario administrador:
    - nombre_encargado, username, password. Si existen valores y el congreso ya tiene
      admin_user, se actualizan esos campos (password solo si viene). Si no tiene
      admin_user y se llenan los 3, se crea y se asocia.
    """
    grupos = request.user.groups.values_list("name", flat=True)
    if "Instructores" in grupos:
        rol = "Instructor"
    elif "Participantes" in grupos:
        rol = "Participante"
    else:
        rol = "Administrador"

    # Restringir a administradores de congresos acotados: solo pueden editar su propio congreso
    if request.user.groups.filter(name="Administradores_Congresos").exists() and not (request.user.is_superuser or "Administrador" in grupos):
        scoped = _get_scoped_congreso_for_group_admin(request.user)
        if not scoped or scoped.id != congreso_id:
            messages.error(request, "No tienes permisos para editar este congreso.")
            return redirect("inicio_congreso", scoped.id if scoped else "inicio")

    try:
        congreso = Congreso.objects.get(pk=congreso_id)
    except Congreso.DoesNotExist:
        messages.error(request, "El congreso solicitado no existe.")
        return redirect("congresos_eventos")

    if request.method == "POST":
        nombre_congreso = (request.POST.get("nombre_congreso") or "").strip()
        descripcion = (request.POST.get("descripcion") or "").strip()
        nombre_encargado = (request.POST.get("nombre_encargado") or "").strip()
        username = (request.POST.get("username") or "").strip()
        password = request.POST.get("password") or ""
        logo_file = request.FILES.get("logo")
        access_password = (request.POST.get("access_password") or "").strip()
        clear_access_password = request.POST.get("clear_access_password") == "on"

        has_error = False
        if not nombre_congreso:
            messages.error(request, "El nombre del congreso o evento no puede estar vacío.")
            has_error = True
        elif Congreso.objects.exclude(pk=congreso.pk).filter(name__iexact=nombre_congreso).exists():
            messages.error(request, "Ya existe un congreso o evento con ese nombre.")
            has_error = True

        if logo_file:
            max_bytes = 3 * 1024 * 1024
            if isinstance(logo_file, UploadedFile) and logo_file.size > max_bytes:
                messages.error(request, "El logo excede el tamaño máximo de 3MB.")
                has_error = True

        # Actualización/creación/reasignación del usuario administrador del congreso
        specials = set(".!%&$}")
        admin = congreso.admin_user
        if nombre_encargado or username or password:
            target_user = None
            if username:
                target_user = User.objects.filter(username=username).first()

            if admin:
                if target_user and target_user.pk != admin.pk:
                    # Reasignar a un usuario existente distinto
                    # Validar y actualizar datos del nuevo admin si se proporcionan
                    if password:
                        if len(password) < 8 or len(password) > 16 or not any(c.isupper() for c in password) or not any(c.islower() for c in password) or not any(c in specials for c in password):
                            messages.error(request, "La contraseña debe tener entre 8 y 16 caracteres, incluir mayúscula, minúscula y un símbolo especial (.!%&$}).")
                            has_error = True
                        else:
                            target_user.set_password(password)
                    if nombre_encargado:
                        target_user.first_name = nombre_encargado
                    if not has_error:
                        target_user.save()
                        # Asegurar pertenencia al grupo de administradores de congresos
                        group, _ = Group.objects.get_or_create(name="Administradores_Congresos")
                        target_user.groups.add(group)
                        congreso.admin_user = target_user
                        # Crear scope si no existe
                        try:
                            CongresoAdminScope.objects.get_or_create(user=target_user, defaults={"congreso": congreso})
                        except Exception:
                            pass
                else:
                    # Actualizar el admin actual
                    if username and (not target_user or target_user.pk == admin.pk):
                        # Cambiar username del mismo admin
                        if User.objects.exclude(pk=admin.pk).filter(username=username).exists():
                            messages.error(request, "El usuario indicado ya existe.")
                            has_error = True
                        else:
                            admin.username = username
                            if "@" in username:
                                admin.email = username
                    if nombre_encargado:
                        admin.first_name = nombre_encargado
                    if password:
                        if len(password) < 8 or len(password) > 16 or not any(c.isupper() for c in password) or not any(c.islower() for c in password) or not any(c in specials for c in password):
                            messages.error(request, "La contraseña debe tener entre 8 y 16 caracteres, incluir mayúscula, minúscula y un símbolo especial (.!%&$}).")
                            has_error = True
                        else:
                            admin.set_password(password)
                    if not has_error:
                        admin.save()
                        group, _ = Group.objects.get_or_create(name="Administradores_Congresos")
                        admin.groups.add(group)
            else:
                # No hay admin actual: usar usuario existente si se indicó, o crear uno nuevo
                if target_user:
                    if password:
                        if len(password) < 8 or len(password) > 16 or not any(c.isupper() for c in password) or not any(c.islower() for c in password) or not any(c in specials for c in password):
                            messages.error(request, "La contraseña debe tener entre 8 y 16 caracteres, incluir mayúscula, minúscula y un símbolo especial (.!%&$}).")
                            has_error = True
                        else:
                            target_user.set_password(password)
                    if nombre_encargado:
                        target_user.first_name = nombre_encargado
                    if not has_error:
                        target_user.save()
                        group, _ = Group.objects.get_or_create(name="Administradores_Congresos")
                        target_user.groups.add(group)
                        congreso.admin_user = target_user
                else:
                    # Crear nuevo requiere al menos nombre, usuario y contraseña
                    if nombre_encargado or username or password:
                        if User.objects.filter(username=username).exists():
                            messages.error(request, "El usuario indicado ya existe.")
                            has_error = True
                        elif len(password) < 8 or len(password) > 16 or not any(c.isupper() for c in password) or not any(c.islower() for c in password) or not any(c in specials for c in password):
                            messages.error(request, "La contraseña debe tener entre 8 y 16 caracteres, incluir mayúscula, minúscula y un símbolo especial (.!%&$}).")
                            has_error = True
                        else:
                            user = User.objects.create_user(
                                username=username,
                                email=username if "@" in username else "",
                                password=password,
                                first_name=nombre_encargado,
                            )
                            group, _ = Group.objects.get_or_create(name="Administradores_Congresos")
                            user.groups.add(group)
                            congreso.admin_user = user
                            try:
                                CongresoAdminScope.objects.get_or_create(user=user, defaults={"congreso": congreso})
                            except Exception:
                                pass
                    else:
                        messages.error(request, "Para crear el usuario administrador del congreso, llena Nombre, Usuario y Contraseña o indica un usuario existente.")
                        has_error = True

        if not has_error:
            congreso.name = nombre_congreso
            congreso.description = descripcion or None
            if logo_file:
                congreso.logo = logo_file
            # Actualizar contraseña de acceso
            if clear_access_password:
                congreso.set_access_password(None)
            elif access_password:
                congreso.set_access_password(access_password)
            congreso.save()
            messages.success(request, "Congreso o evento actualizado correctamente.")
            return redirect("congresos_eventos")

    # Valores por defecto para el formulario
    initial = {
        "nombre_congreso": congreso.name,
        "descripcion": congreso.description or "",
        "nombre_encargado": congreso.admin_user.first_name if congreso.admin_user else "",
        "username": congreso.admin_user.username if congreso.admin_user else "",
    }

    return render(request, "editar_congreso_evento.html", {"rol": rol, "congreso": congreso, "is_scoped_admin": False, **initial})

@login_required
def correos_view(request):
    deny = _deny_non_admin_roles(request)
    if deny: return deny
    # Redirigir a admins de congresos con scope al panel del congreso
    scoped = _get_scoped_congreso_for_group_admin(request.user)
    if scoped:
        return redirect("inicio_congreso", scoped.id)
    grupos = request.user.groups.values_list("name", flat=True)
    if "Instructores" in grupos:
        rol = "Instructor"
    elif "Participantes" in grupos:
        rol = "Participante"
    else:
        rol = "Administrador"

    # Crear congreso rápido si no hay ninguno aún (para empezar a trabajar)
    if not Congreso.objects.exists():
        Congreso.objects.create(name="Congreso General")

    # Selección de congreso actual (por querystring ?c=ID o por post oculto)
    selected_id = request.POST.get("congreso_id") or request.GET.get("c")
    try:
        congreso = Congreso.objects.get(pk=selected_id) if selected_id else Congreso.objects.order_by("name").first()
    except Congreso.DoesNotExist:
        congreso = Congreso.objects.order_by("name").first()

    # Acciones
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "create_congreso":
            name = (request.POST.get("name") or "").strip()
            if not name:
                messages.error(request, "El nombre del congreso no puede estar vacío.")
            elif Congreso.objects.filter(name__iexact=name).exists():
                messages.error(request, "Ya existe un congreso con ese nombre.")
            else:
                congreso = Congreso.objects.create(name=name)
                messages.success(request, "Congreso creado.")
            return HttpResponseRedirect(reverse("correos") + f"?c={congreso.id}")

        elif action == "add_domain":
            raw = (request.POST.get("domain") or "").strip().lower()
            # Normalizar: asegurar que empiece con '@'
            if raw and not raw.startswith("@"):
                raw = "@" + raw
            if not raw or "@" not in raw or "." not in raw:
                messages.error(request, "Ingresa un dominio válido. Ej.: @gmail.com")
            else:
                try:
                    AllowedEmailDomain.objects.create(congreso=congreso, domain=raw)
                    messages.success(request, "Dominio agregado.")
                except Exception:
                    messages.error(request, "Ese dominio ya existe para este congreso.")
            return HttpResponseRedirect(reverse("correos") + f"?c={congreso.id}")

        elif action == "delete_domain":
            did = request.POST.get("domain_id")
            AllowedEmailDomain.objects.filter(id=did, congreso=congreso).delete()
            messages.success(request, "Dominio eliminado.")
            return HttpResponseRedirect(reverse("correos") + f"?c={congreso.id}")

        elif action == "update_domain":
            did = request.POST.get("domain_id")
            raw = (request.POST.get("domain") or "").strip().lower()
            if raw and not raw.startswith("@"):
                raw = "@" + raw
            if not raw or "@" not in raw or "." not in raw:
                messages.error(request, "Ingresa un dominio válido. Ej.: @gmail.com")
            else:
                try:
                    dom = AllowedEmailDomain.objects.get(id=did, congreso=congreso)
                    if AllowedEmailDomain.objects.exclude(id=did).filter(congreso=congreso, domain=raw).exists():
                        messages.error(request, "Ese dominio ya existe para este congreso.")
                    else:
                        dom.domain = raw
                        dom.save(update_fields=["domain"])
                        messages.success(request, "Dominio actualizado.")
                except AllowedEmailDomain.DoesNotExist:
                    messages.error(request, "El dominio a actualizar no existe.")
            return HttpResponseRedirect(reverse("correos") + f"?c={congreso.id}")

    congresos = list(Congreso.objects.all().order_by("name"))
    dominios = list(AllowedEmailDomain.objects.filter(congreso=congreso).order_by("domain"))
    # Modo edición opcional via querystring ?edit=<id>
    edit_id = request.GET.get("edit")
    edit_domain = None
    if edit_id:
        try:
            edit_domain = AllowedEmailDomain.objects.get(id=edit_id, congreso=congreso)
        except AllowedEmailDomain.DoesNotExist:
            edit_domain = None

    ctx = {
        "rol": rol,
        "congresos": congresos,
        "congreso": congreso,
        "dominios": dominios,
        "edit_domain": edit_domain,
    }
    return render(request, "correos.html", ctx)

@login_required
def nivel_desempeno_view(request):
    deny = _deny_non_admin_roles(request)
    if deny: return deny
    # Redirigir a admins de congresos con scope al panel del congreso
    scoped = _get_scoped_congreso_for_group_admin(request.user)
    if scoped:
        return redirect("inicio_congreso", scoped.id)
    grupos = request.user.groups.values_list("name", flat=True)
    if request.user.is_superuser or "Administrador" in grupos:
        rol = "Administrador"
    elif "Instructores" in grupos:
        rol = "Instructor"
    elif "Participantes" in grupos:
        rol = "Participante"
    else:
        rol = "Administrador"

    if not Congreso.objects.exists():
        Congreso.objects.create(name="Congreso General")

    selected_id = request.POST.get("congreso_id") or request.GET.get("c")
    try:
        congreso = Congreso.objects.get(pk=selected_id) if selected_id else Congreso.objects.order_by("name").first()
    except Congreso.DoesNotExist:
        congreso = Congreso.objects.order_by("name").first()

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "add_level":
            name = (request.POST.get("name") or "").strip()

            if not name:
                messages.error(request, "El nombre del nivel no puede estar vacío.")
            else:
                try:
                    PerformanceLevel.objects.create(congreso=congreso, name=name)
                    messages.success(request, "Nivel agregado.")
                except Exception:
                    messages.error(request, "Ese nivel ya existe para este congreso.")
            return HttpResponseRedirect(reverse("nivel_desempeno") + f"?c={congreso.id}")

        elif action == "delete_level":
            lid = request.POST.get("level_id")
            PerformanceLevel.objects.filter(id=lid, congreso=congreso).delete()
            messages.success(request, "Nivel eliminado.")
            return HttpResponseRedirect(reverse("nivel_desempeno") + f"?c={congreso.id}")

        elif action == "update_level":
            lid = request.POST.get("level_id")
            name = (request.POST.get("name") or "").strip()
            if not name:
                messages.error(request, "El nombre del nivel no puede estar vacío.")
            else:
                try:
                    lvl = PerformanceLevel.objects.get(id=lid, congreso=congreso)
                    if PerformanceLevel.objects.exclude(id=lid).filter(congreso=congreso, name__iexact=name).exists():
                        messages.error(request, "Ese nivel ya existe para este congreso.")
                    else:
                        lvl.name = name
                        lvl.save(update_fields=["name"])
                        messages.success(request, "Nivel actualizado.")
                except PerformanceLevel.DoesNotExist:
                    messages.error(request, "El nivel a actualizar no existe.")
            return HttpResponseRedirect(reverse("nivel_desempeno") + f"?c={congreso.id}")

    congresos = list(Congreso.objects.all().order_by("name"))
    niveles = list(PerformanceLevel.objects.filter(congreso=congreso).order_by("name"))
    edit_id = request.GET.get("edit")
    edit_level = None
    if edit_id:
        try:
            edit_level = PerformanceLevel.objects.get(id=edit_id, congreso=congreso)
        except PerformanceLevel.DoesNotExist:
            edit_level = None
    return render(request, "nivel_desempeno.html", {"rol": rol, "congresos": congresos, "congreso": congreso, "niveles": niveles, "edit_level": edit_level})

@login_required
def perfil_view(request):
    deny = _deny_non_admin_roles(request)
    if deny: return deny
    # Determinar rol como en otras vistas
    grupos = request.user.groups.values_list("name", flat=True)
    if request.user.is_superuser or "Administrador" in grupos:
        rol = "Administrador"
    elif "Instructores" in grupos:
        rol = "Instructor"
    elif "Participantes" in grupos:
        rol = "Participante"
    else:
        rol = "Administrador"

    # Construir lista de dominios permitidos a nivel perfil global en base a las
    # membresías aprobadas del usuario. Si cambia su username (que en este sistema
    # suele ser el correo), debe respetar al menos uno de estos dominios.
    domain_set = set()
    memberships = UserCongresoMembership.objects.filter(user=request.user, status="approved").select_related("congreso")
    for mem in memberships:
        dominios = AllowedEmailDomain.objects.filter(congreso=mem.congreso).values_list("domain", flat=True)
        for d in dominios:
            if d:
                domain_set.add(d.lower().strip())

    context = {
        "rol": rol,
        "username_value": request.user.username,
        "allowed_domains_profile": sorted(domain_set),
    }

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "update_name":
            new_username = request.POST.get("username", "").strip()
            if not new_username:
                messages.error(request, "El nombre de usuario no puede estar vacío.")
            elif User.objects.exclude(pk=request.user.pk).filter(username=new_username).exists():
                messages.error(request, "Ese nombre de usuario ya está en uso.")
            elif len(new_username) > 150:
                messages.error(request, "El nombre de usuario no puede exceder 150 caracteres.")
            else:
                # Validar formato de correo obligatorio
                try:
                    validate_email(new_username)
                except ValidationError:
                    messages.error(request, "El usuario debe ser un correo electrónico válido.")
                    return HttpResponseRedirect(reverse("perfil") + "#user")
                # Si el username parece un correo y existen dominios permitidos por alguna
                # membresía aprobada, exigir que el dominio esté permitido.
                allowed_domains_profile = context.get("allowed_domains_profile") or []
                if "@" in new_username and allowed_domains_profile:
                    email_lc = new_username.lower()
                    if not any(email_lc.endswith(d) for d in allowed_domains_profile):
                        pretty = ", ".join(allowed_domains_profile)
                        messages.error(
                            request,
                            f"El correo no pertenece a un dominio permitido por tus congresos activos. Dominios permitidos: {pretty}",
                        )
                        return HttpResponseRedirect(reverse("perfil") + "#user")
                request.user.username = new_username
                # También actualizar el email si el username es un correo
                if "@" in new_username:
                    request.user.email = new_username
                request.user.save()
                messages.success(request, "Nombre de usuario actualizado correctamente.")
                context["username_value"] = new_username
            # PRG: redirigir para mostrar mensajes y evitar reenvío
            return HttpResponseRedirect(reverse("perfil") + "#user")
        

        elif action == "update_password":
            current = request.POST.get("current_password", "")
            new = request.POST.get("new_password", "")
            confirm = request.POST.get("confirm_password", "")

            # Validaciones básicas
            if not request.user.check_password(current):
                messages.error(request, "La contraseña actual no es correcta.")
            elif new != confirm:
                messages.error(request, "La nueva contraseña y su confirmación no coinciden.")
            elif len(new) < 8 or len(new) > 16:
                messages.error(request, "La nueva contraseña debe tener entre 8 y 16 caracteres.")
            else:
                specials = set(".!%&$}")
                if not any(c.isupper() for c in new) or not any(c.islower() for c in new) or not any(c in specials for c in new):
                    messages.error(request, "La nueva contraseña debe contener al menos una letra mayúscula, una letra minúscula y un símbolo especial (.!%&$}).")
                else:
                    try:
                        request.user.set_password(new)
                        request.user.save()
                        update_session_auth_hash(request, request.user)  # mantener sesión
                        messages.success(request, "Contraseña actualizada correctamente.")
                    except Exception:
                        messages.error(request, "Ocurrió un error al actualizar la contraseña. Inténtalo de nuevo.")
            # PRG: redirigir al bloque de contraseña
            return HttpResponseRedirect(reverse("perfil") + "#password")

    return render(request, "perfil.html", context)


@login_required
def inicio_congreso_view(request, congreso_id: int):
    """Página de inicio/administración para un congreso específico."""
    grupos = request.user.groups.values_list("name", flat=True)
    if request.user.is_superuser or "Administrador" in grupos:
        rol = "Administrador"
    elif "Instructores" in grupos:
        rol = "Instructor"
    elif "Participantes" in grupos:
        rol = "Participante"
    else:
        rol = "Administrador"

    try:
        congreso = Congreso.objects.get(pk=congreso_id)
    except Congreso.DoesNotExist:
        messages.error(request, "El congreso solicitado no existe.")
        return redirect("congresos_eventos")

    # Validación de acceso por membresía (los administradores de congreso scoped también pasan)
    if not _user_is_congreso_admin(request.user, congreso):
        guard = _ensure_congreso_access_or_redirect(request, congreso)
        if guard:
            return guard

    # Aprobaciones pendientes
    pending = UserCongresoMembership.objects.filter(congreso=congreso, status="pending").select_related("user").order_by("created_at")

    # Procesar acciones de aprobación/rechazo
    if request.method == "POST":
        action = request.POST.get("action")
        mem_id = request.POST.get("membership_id")
        # Permisos: admin del congreso o del grupo Administradores_Congresos
        is_congreso_admin = (congreso.admin_user_id == request.user.id)
        in_group_admins = request.user.groups.filter(name="Administradores_Congresos").exists()
        if action in {"approve_membership", "reject_membership"} and mem_id and (is_congreso_admin or in_group_admins or request.user.is_superuser):
            try:
                membership = UserCongresoMembership.objects.get(id=mem_id, congreso=congreso)
                membership.status = "approved" if action == "approve_membership" else "rejected"
                membership.decided_at = timezone.now()
                membership.decided_by = request.user
                membership.save(update_fields=["status", "decided_at", "decided_by"])
                if action == "approve_membership":
                    messages.success(
                        request,
                        f"Se aprobó el acceso de {membership.user.username} como {membership.role}.",
                        extra_tags="approval",
                    )
                    # Enviar correo de notificación de aprobación (silencioso si falla)
                    user_email = membership.user.email
                    if user_email:
                        try:
                            login_link = request.build_absolute_uri(reverse("login"))
                            subject = f"Tu cuenta fue aprobada en {congreso.name}"
                            saludo_nombre = membership.user.first_name or membership.user.username
                            body = (
                                f"Hola {saludo_nombre},\n\n"
                                f"Tu cuenta ha sido aprobada en el congreso '{congreso.name}'. Ya puedes iniciar sesión y participar.\n\n"
                                f"Ingresa aquí: {login_link}\n\n"
                                "Saludos."
                            )
                            html_body = (
                                f"<p>Hola {saludo_nombre},</p>"
                                f"<p>Tu cuenta ha sido aprobada en el congreso '<strong>{congreso.name}</strong>'. Ya puedes iniciar sesión y participar.</p>"
                                f"<p><a href='{login_link}' style='display:inline-block;padding:10px 16px;background:#0d6efd;color:#fff;text-decoration:none;border-radius:6px;'>Iniciar sesión</a></p>"
                                "<p>Saludos.</p>"
                            )
                            from_email = getattr(settings, "DEFAULT_FROM_EMAIL", None) or getattr(
                                settings, "EMAIL_HOST_USER", "no-reply@localhost"
                            )
                            send_mail(subject, body, from_email, [user_email], fail_silently=True, html_message=html_body)
                        except Exception:
                            pass
                else:
                    # Usar 'error' para que el aviso se muestre en rojo
                    messages.error(
                        request,
                        f"Se rechazó el acceso de {membership.user.username}.",
                        extra_tags="approval",
                    )
            except UserCongresoMembership.DoesNotExist:
                messages.error(request, "La solicitud indicada no existe.")
            return HttpResponseRedirect(reverse("inicio_congreso", args=[congreso.id]))

    is_admin = _user_is_congreso_admin(request.user, congreso)
    # Construir tabla de usuarios aprobados con valores de campos extra
    approved_memberships = list(
        UserCongresoMembership.objects.filter(congreso=congreso, status="approved")
        .select_related("user", "performance_level")
        .order_by("user__first_name", "user__last_name", "user__username")
    )

    # Campos extra activos (registro y perfil) del congreso y globales
    extra_fields = list(
        ExtraField.objects.filter(
            (Q(congreso=congreso) | Q(congreso__isnull=True)),
            active=True,
            section__in=["registro", "perfil"],
        ).order_by("order", "name")
    )

    # Valores de campos extra para los usuarios en esta lista
    approved_user_ids = [m.user_id for m in approved_memberships] or [0]
    field_ids = [f.id for f in extra_fields] or [0]
    values_qs = UserExtraFieldValue.objects.filter(
        (Q(congreso=congreso) | Q(congreso__isnull=True)),
        user_id__in=approved_user_ids,
        field_id__in=field_ids,
    ).values("user_id", "field_id", "value")
    value_map = {(row["user_id"], row["field_id"]): row["value"] for row in values_qs}

    # Preparar filas para la plantilla
    approved_rows = []
    for m in approved_memberships:
        # Mapa por id y lista alineada con extra_fields para fácil render en plantilla
        row_values_map = {}
        row_values_ordered = []
        for f in extra_fields:
            if f.role_scope == "participante" and m.role != "participante":
                row_values_ordered.append("")
                continue
            if f.role_scope == "instructor" and m.role != "instructor":
                row_values_ordered.append("")
                continue
            val = value_map.get((m.user_id, f.id), "")
            row_values_map[f.id] = val
            row_values_ordered.append(val)
        approved_rows.append({
            "membership": m,
            "user": m.user,
            "role": m.role,
            "performance_level": m.performance_level,
            "values": row_values_map,
            "values_ordered": row_values_ordered,
        })

    ctx = {
        "rol": rol,
        "congreso": congreso,
        "pending_memberships": list(pending),
        "is_admin": is_admin,
        "extra_fields": extra_fields,
        "approved_rows": approved_rows,
        "is_scoped_admin": bool(_get_scoped_congreso_for_group_admin(request.user)),
    }
    return render(request, "inicio_congreso.html", ctx)


@login_required
def perfil_congreso_view(request, congreso_id: int):
    """Perfil dentro del contexto de un congreso específico.

    Permite actualizar nombre de usuario y contraseña, manteniendo la sesión activa
    al cambiar la contraseña. Replica la política y validaciones de "perfil_view",
    pero renderiza una plantilla enfocada al área del congreso y conserva el id del
    congreso en las redirecciones.
    """
    # Rol del usuario
    grupos = request.user.groups.values_list("name", flat=True)
    if request.user.is_superuser or "Administrador" in grupos:
        rol = "Administrador"
    elif "Instructores" in grupos:
        rol = "Instructor"
    elif "Participantes" in grupos:
        rol = "Participante"
    else:
        rol = "Administrador"

    # Congreso en contexto
    try:
        congreso = Congreso.objects.get(pk=congreso_id)
    except Congreso.DoesNotExist:
        messages.error(request, "El congreso solicitado no existe.")
        return redirect("congresos_eventos")

    # Validación de acceso por membresía
    guard = _ensure_congreso_access_or_redirect(request, congreso)
    if guard:
        return guard

    # Determinar si debemos ocultar paneles (admins o administradores de congresos)
    hide_extras_and_level = (
        request.user.is_superuser
        or request.user.groups.filter(name__in=["Administrador", "Administradores_Congresos"]).exists()
    )

    # Nombre completo del encargado actual (si existe)
    def _full_name(u) -> str:
        if not u:
            return ""
        return (u.first_name or "").strip() + (" " + (u.last_name or "").strip() if (u.last_name or "").strip() else "")

    # Cargar membresía del usuario en el congreso
    membership = (
        UserCongresoMembership.objects.filter(user=request.user, congreso=congreso)
        .order_by("-created_at")
        .first()
    )
    # Niveles disponibles en el congreso
    levels = list(PerformanceLevel.objects.filter(congreso=congreso).order_by("name"))

    # Cargar campos extra (globales + del congreso) activos y visibles por rol
    role_for_extras = membership.role if membership else ("instructor" if "Instructores" in grupos else ("participante" if "Participantes" in grupos else None))
    extra_fields = []
    if role_for_extras:
        ef_q = ExtraField.objects.filter(Q(congreso=congreso) | Q(congreso__isnull=True), active=True)
        if role_for_extras == "instructor":
            ef_q = ef_q.filter(Q(role_scope="both") | Q(role_scope="instructor"))
        else:
            ef_q = ef_q.filter(Q(role_scope="both") | Q(role_scope="participante"))
        extra_fields = list(ef_q.order_by("order", "name"))

    # Prefill valores
    extra_values = {}
    extra_items = []
    if extra_fields:
        vals = UserExtraFieldValue.objects.filter(user=request.user, field__in=extra_fields).filter(
            Q(congreso=congreso) | Q(congreso__isnull=True)
        )
        for v in vals:
            extra_values[v.field_id] = v.value
        for f in extra_fields:
            extra_items.append({"field": f, "value": extra_values.get(f.id, "")})

    context = {
        "rol": rol,
        "username_value": request.user.username,
        "congreso": congreso,
        # Valores iniciales para el bloque "Datos públicos"
        "congreso_name": congreso.name,
        "encargado_name": _full_name(congreso.admin_user) or _full_name(request.user),
        # Dominios permitidos del congreso para validar username si es correo
        "allowed_domains_congreso": list(
            AllowedEmailDomain.objects.filter(congreso=congreso).values_list("domain", flat=True)
        ),
        "membership": membership,
        "levels": levels,
        "extra_fields": extra_fields,
        "extra_values": extra_values,
        "extra_items": extra_items,
        "hide_extras_and_level": hide_extras_and_level,
    }

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "update_name":
            new_username = request.POST.get("username", "").strip()
            if not new_username:
                messages.error(request, "El nombre de usuario no puede estar vacío.")
            elif User.objects.exclude(pk=request.user.pk).filter(username=new_username).exists():
                messages.error(request, "Ese nombre de usuario ya está en uso.")
            elif len(new_username) > 150:
                messages.error(request, "El nombre de usuario no puede exceder 150 caracteres.")
            else:
                # Validar formato de correo obligatorio
                try:
                    validate_email(new_username)
                except ValidationError:
                    messages.error(request, "El usuario debe ser un correo electrónico válido.")
                    return HttpResponseRedirect(reverse("perfil_congreso", args=[congreso.id]) + "#user")
                dominios = [d.lower().strip() for d in context.get("allowed_domains_congreso", []) if d]
                if dominios and "@" in new_username:
                    email_lc = new_username.lower()
                    if not any(email_lc.endswith(d) for d in dominios):
                        pretty = ", ".join(sorted(dominios))
                        messages.error(
                            request,
                            f"El correo no pertenece a un dominio permitido para este congreso. Dominios permitidos: {pretty}",
                        )
                        return HttpResponseRedirect(reverse("perfil_congreso", args=[congreso.id]) + "#user")
                request.user.username = new_username
                if "@" in new_username:
                    request.user.email = new_username
                request.user.save()
                messages.success(request, "Nombre de usuario actualizado correctamente.")
                context["username_value"] = new_username
            # PRG con ancla del bloque
            return HttpResponseRedirect(reverse("perfil_congreso", args=[congreso.id]) + "#user")

        elif action == "update_password":
            current = request.POST.get("current_password", "")
            new = request.POST.get("new_password", "")
            confirm = request.POST.get("confirm_password", "")

            if not request.user.check_password(current):
                messages.error(request, "La contraseña actual no es correcta.")
            elif new != confirm:
                messages.error(request, "La nueva contraseña y su confirmación no coinciden.")
            elif len(new) < 8 or len(new) > 16:
                messages.error(request, "La nueva contraseña debe tener entre 8 y 16 caracteres.")
            else:
                specials = set(".!%&$}")
                if not any(c.isupper() for c in new) or not any(c.islower() for c in new) or not any(c in specials for c in new):
                    messages.error(request, "La nueva contraseña debe contener al menos una letra mayúscula, una letra minúscula y un símbolo especial (.!%&$}).")
                else:
                    try:
                        request.user.set_password(new)
                        request.user.save()
                        update_session_auth_hash(request, request.user)
                        messages.success(request, "Contraseña actualizada correctamente.")
                    except Exception:
                        messages.error(request, "Ocurrió un error al actualizar la contraseña. Inténtalo de nuevo.")
            return HttpResponseRedirect(reverse("perfil_congreso", args=[congreso.id]) + "#password")

        elif action == "update_public":
            nombre_congreso = (request.POST.get("nombre_congreso") or "").strip()
            nombre_encargado = (request.POST.get("nombre_encargado") or "").strip()

            has_error = False
            # Validar nombre de congreso
            if not nombre_congreso:
                messages.error(request, "El nombre del congreso o evento no puede estar vacío.")
                has_error = True
            elif Congreso.objects.exclude(pk=congreso.pk).filter(name__iexact=nombre_congreso).exists():
                messages.error(request, "Ya existe un congreso o evento con ese nombre.")
                has_error = True

            # Si hay errores, conservar valores en el contexto y recargar
            if has_error:
                context["congreso_name"] = nombre_congreso
                context["encargado_name"] = nombre_encargado
                return HttpResponseRedirect(reverse("perfil_congreso", args=[congreso.id]) + "#public")

            # Actualizar nombre del congreso
            congreso.name = nombre_congreso
            congreso.save(update_fields=["name"]) 

            # Actualizar nombre del usuario asignado al congreso (first_name y last_name)
            if nombre_encargado:
                if congreso.admin_user:
                    admin = congreso.admin_user
                    parts = [p for p in nombre_encargado.split(" ") if p]
                    if len(parts) >= 2:
                        admin.first_name = parts[0]
                        admin.last_name = " ".join(parts[1:])
                    else:
                        admin.first_name = nombre_encargado
                        admin.last_name = ""
                    admin.save(update_fields=["first_name", "last_name"])
                else:
                    # Si el congreso no tiene admin asignado, asociar al usuario actual
                    admin = request.user
                    parts = [p for p in nombre_encargado.split(" ") if p]
                    if len(parts) >= 2:
                        admin.first_name = parts[0]
                        admin.last_name = " ".join(parts[1:])
                    else:
                        admin.first_name = nombre_encargado
                        admin.last_name = ""
                    admin.save(update_fields=["first_name", "last_name"]) 
                    congreso.admin_user = admin
                    congreso.save(update_fields=["admin_user"]) 

            messages.success(request, "Datos públicos actualizados.")
            return HttpResponseRedirect(reverse("perfil_congreso", args=[congreso.id]) + "#public")

        elif action == "update_level":
            if hide_extras_and_level:
                messages.info(request, "Este ajuste no aplica para administradores.")
                return HttpResponseRedirect(reverse("perfil_congreso", args=[congreso.id]) + "#level")
            # Solo participantes pueden seleccionar nivel
            if not membership or membership.role != "participante":
                messages.error(request, "Este ajuste aplica únicamente a Participantes.")
                return HttpResponseRedirect(reverse("perfil_congreso", args=[congreso.id]) + "#level")
            level_id = request.POST.get("performance_level_id")
            if level_id:
                try:
                    lvl = PerformanceLevel.objects.get(id=level_id, congreso=congreso)
                    membership.performance_level = lvl
                    membership.save(update_fields=["performance_level"])
                    messages.success(request, "Nivel de desempeño actualizado.")
                except PerformanceLevel.DoesNotExist:
                    messages.error(request, "Nivel inválido para este congreso.")
            else:
                # Permitir limpiar selección
                membership.performance_level = None
                membership.save(update_fields=["performance_level"])
                messages.success(request, "Nivel de desempeño actualizado.")
            return HttpResponseRedirect(reverse("perfil_congreso", args=[congreso.id]) + "#level")

        elif action == "update_extras":
            if hide_extras_and_level:
                messages.info(request, "Este ajuste no aplica para administradores.")
                return HttpResponseRedirect(reverse("perfil_congreso", args=[congreso.id]) + "#extras")
            if not extra_fields:
                return HttpResponseRedirect(reverse("perfil_congreso", args=[congreso.id]) + "#extras")
            # Validar y guardar similares al registro
            for f in extra_fields:
                # Respetar rol_scope (ya filtrado)
                key = f"extra_{f.id}"
                raw_val = (request.POST.get(key) or "").strip()
                if f.required and not raw_val and f.field_type != "boolean":
                    messages.error(request, f"El campo '{f.name}' es obligatorio.")
                    return HttpResponseRedirect(reverse("perfil_congreso", args=[congreso.id]) + "#extras")
                if raw_val:
                    if f.field_type == "email":
                        try:
                            validate_email(raw_val)
                        except ValidationError:
                            messages.error(request, f"Ingresa un correo válido en el campo '{f.name}'.")
                            return HttpResponseRedirect(reverse("perfil_congreso", args=[congreso.id]) + "#extras")
                    elif f.field_type == "number":
                        try:
                            float(raw_val)
                        except ValueError:
                            messages.error(request, f"Ingresa un número válido en el campo '{f.name}'.")
                            return HttpResponseRedirect(reverse("perfil_congreso", args=[congreso.id]) + "#extras")
                    elif f.field_type == "select":
                        choices = f.get_choices()
                        if choices and raw_val not in choices:
                            messages.error(request, f"Selecciona una opción válida para '{f.name}'.")
                            return HttpResponseRedirect(reverse("perfil_congreso", args=[congreso.id]) + "#extras")
                # Guardar por ámbito (global o congreso)
                UserExtraFieldValue.objects.update_or_create(
                    user=request.user,
                    congreso=congreso if f.congreso_id else None,
                    field=f,
                    defaults={"value": raw_val},
                )
            messages.success(request, "Campos adicionales actualizados.")
            return HttpResponseRedirect(reverse("perfil_congreso", args=[congreso.id]) + "#extras")

    return render(request, "perfil_congreso.html", context)


# -------------------
# PLACEHOLDER PÁGINAS DEL ÁREA DE CONGRESO
# -------------------
@login_required
def talleres_list_view(request, congreso_id: int):
    try:
        congreso = Congreso.objects.get(pk=congreso_id)
    except Congreso.DoesNotExist:
        messages.error(request, "El congreso solicitado no existe.")
        return redirect("congresos_eventos")
    guard = _ensure_congreso_access_or_redirect(request, congreso)
    if guard:
        return guard
    # Rol del usuario (para sidebar)
    grupos = request.user.groups.values_list("name", flat=True)
    if "Instructores" in grupos:
        rol = "Instructor"
    elif "Participantes" in grupos:
        rol = "Participante"
    else:
        rol = "Administrador"
    # Permisos de administración para eliminar
    is_admin = (
        (congreso.admin_user_id == request.user.id)
        or request.user.groups.filter(name__in=["Administradores_Congresos", "Administrador"]).exists()
        or request.user.is_superuser
    )

    # Acciones POST
    if request.method == "POST" and is_admin:
        action = request.POST.get("action")
        if action == "delete_taller":
            tid = request.POST.get("taller_id")
            if tid:
                try:
                    Taller.objects.get(id=tid, congreso=congreso).delete()
                    messages.success(request, "Taller eliminado.")
                except Taller.DoesNotExist:
                    messages.error(request, "El taller indicado no existe.")
            else:
                messages.error(request, "No se indicó el taller a eliminar.")
            return redirect("talleres_list", congreso.id)
    # Lista sencilla (se puede enriquecer después)
    items = list(Taller.objects.filter(congreso=congreso).order_by("created_at"))
    return render(request, "talleres_list.html", {"rol": rol, "congreso": congreso, "talleres": items, "is_admin": is_admin, "is_scoped_admin": bool(_get_scoped_congreso_for_group_admin(request.user))})


@login_required
def export_talleres_excel_view(request, congreso_id: int):
    """Genera un archivo Excel con todos los talleres del congreso indicado.

    Columnas:
    1) Numeración
    2) ID Taller
    3) Nombre del taller
    4) Lugar del taller
    5) Cupos máximo
    6) Descripción del taller
    7) ID de instructor
    8) Nombre del instructor
    9) Correo del instructor
    """
    from django.http import HttpResponse
    try:
        congreso = Congreso.objects.get(pk=congreso_id)
    except Congreso.DoesNotExist:
        messages.error(request, "El congreso solicitado no existe.")
        return redirect("congresos_eventos")
    guard = _ensure_congreso_access_or_redirect(request, congreso)
    if guard:
        return guard

    talleres = (
        Taller.objects.filter(congreso=congreso)
        .select_related("instructor")
        .order_by("title")
    )

    # Función para normalizar (quitar acentos / diacríticos)
    import unicodedata
    def _normalize(text: str | None) -> str:
        if not text:
            return ""
        # NFD separa caracteres base y marcas; filtramos las marcas (categoria Mn)
        return "".join(c for c in unicodedata.normalize("NFD", text) if unicodedata.category(c) != "Mn")

    # Preparar workbook (openpyxl). Si no está disponible (ImportError), retornar CSV como fallback.
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Talleres"
        headers = [
            _normalize("#"),
            _normalize("ID Taller"),
            _normalize("Nombre del taller"),
            _normalize("Lugar"),
            _normalize("Cupo maximo"),
            _normalize("Descripcion"),
            _normalize("ID Instructor"),
            _normalize("Nombre Instructor"),
            _normalize("Correo Instructor"),
        ]
        ws.append(headers)
        for idx, t in enumerate(talleres, start=1):
            instructor = t.instructor
            ws.append([
                idx,
                t.id,
                _normalize(t.title),
                _normalize(t.lugar or ""),
                t.cupo_maximo if t.cupo_maximo is not None else "",
                _normalize(t.description or ""),
                instructor.id if instructor else "",
                _normalize((instructor.first_name or instructor.username) if instructor else ""),
                _normalize((instructor.email or instructor.username) if instructor else ""),
            ])
        # Ajuste sencillo de ancho de columnas
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value) if cell.value is not None else ""
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=Talleres.xlsx"
        wb.save(response)
        return response
    except ImportError:
        # Fallback CSV si openpyxl no está instalado; instale con: pip install openpyxl
        import csv, io
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow([
            _normalize("#"),
            _normalize("ID Taller"),
            _normalize("Nombre del taller"),
            _normalize("Lugar"),
            _normalize("Cupo maximo"),
            _normalize("Descripcion"),
            _normalize("ID Instructor"),
            _normalize("Nombre Instructor"),
            _normalize("Correo Instructor"),
        ])
        for idx, t in enumerate(talleres, start=1):
            instructor = t.instructor
            writer.writerow([
                idx,
                t.id,
                _normalize(t.title),
                _normalize(t.lugar or ""),
                t.cupo_maximo if t.cupo_maximo is not None else "",
                _normalize((t.description or "").replace("\n"," ").strip()),
                instructor.id if instructor else "",
                _normalize((instructor.first_name or instructor.username) if instructor else ""),
                _normalize((instructor.email or instructor.username) if instructor else ""),
            ])
        response = HttpResponse(buffer.getvalue(), content_type="text/csv")
        response["Content-Disposition"] = "attachment; filename=Talleres.csv"
        return response


@login_required
def export_conferencias_excel_view(request, congreso_id: int):
    """Genera un archivo Excel con todas las conferencias del congreso indicado.

    Columnas:
    1) Numeración
    2) ID Conferencia
    3) Nombre de la conferencia
    4) Lugar de la conferencia
    5) Cupos máximo
    6) Descripción de la conferencia
    7) ID de instructor
    8) Nombre del instructor
    9) Correo del instructor
    """
    from django.http import HttpResponse
    try:
        congreso = Congreso.objects.get(pk=congreso_id)
    except Congreso.DoesNotExist:
        messages.error(request, "El congreso solicitado no existe.")
        return redirect("congresos_eventos")
    guard = _ensure_congreso_access_or_redirect(request, congreso)
    if guard:
        return guard

    conferencias = (
        Conferencia.objects.filter(congreso=congreso)
        .select_related("instructor")
        .order_by("title")
    )

    import unicodedata
    def _normalize(text: str | None) -> str:
        if not text:
            return ""
        return "".join(c for c in unicodedata.normalize("NFD", text) if unicodedata.category(c) != "Mn")

    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Conferencias"
        headers = [
            _normalize("#"),
            _normalize("ID Conferencia"),
            _normalize("Nombre de la conferencia"),
            _normalize("Lugar"),
            _normalize("Cupo maximo"),
            _normalize("Descripcion"),
            _normalize("ID Instructor"),
            _normalize("Nombre Instructor"),
            _normalize("Correo Instructor"),
        ]
        ws.append(headers)
        for idx, c in enumerate(conferencias, start=1):
            instructor = c.instructor
            ws.append([
                idx,
                c.id,
                _normalize(c.title),
                _normalize(c.lugar or ""),
                c.cupo_maximo if c.cupo_maximo is not None else "",
                _normalize(c.description or ""),
                instructor.id if instructor else "",
                _normalize((instructor.first_name or instructor.username) if instructor else ""),
                _normalize((instructor.email or instructor.username) if instructor else ""),
            ])
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value) if cell.value is not None else ""
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=Conferencias.xlsx"
        wb.save(response)
        return response
    except ImportError:
        import csv, io
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow([
            _normalize("#"),
            _normalize("ID Conferencia"),
            _normalize("Nombre de la conferencia"),
            _normalize("Lugar"),
            _normalize("Cupo maximo"),
            _normalize("Descripcion"),
            _normalize("ID Instructor"),
            _normalize("Nombre Instructor"),
            _normalize("Correo Instructor"),
        ])
        for idx, c in enumerate(conferencias, start=1):
            instructor = c.instructor
            writer.writerow([
                idx,
                c.id,
                _normalize(c.title),
                _normalize(c.lugar or ""),
                c.cupo_maximo if c.cupo_maximo is not None else "",
                _normalize((c.description or "").replace("\n"," ").strip()),
                instructor.id if instructor else "",
                _normalize((instructor.first_name or instructor.username) if instructor else ""),
                _normalize((instructor.email or instructor.username) if instructor else ""),
            ])
        response = HttpResponse(buffer.getvalue(), content_type="text/csv")
        response["Content-Disposition"] = "attachment; filename=Conferencias.csv"
        return response


@login_required
def export_concursos_excel_view(request, congreso_id: int):
    """Genera un archivo Excel con todos los concursos del congreso indicado.

    Columnas:
    1) Numeración
    2) ID Concurso
    3) Nombre del concurso
    4) Modalidad (individual/grupal)
    5) Lugar
    6) Descripción
    7) ID Instructor
    8) Nombre Instructor
    9) Correo Instructor
    10) Cupo máximo (solo individual)
    11) Número de equipos (solo grupal)
    12) Máx por equipo (solo grupal)
    """
    from django.http import HttpResponse
    try:
        congreso = Congreso.objects.get(pk=congreso_id)
    except Congreso.DoesNotExist:
        messages.error(request, "El congreso solicitado no existe.")
        return redirect("congresos_eventos")
    guard = _ensure_congreso_access_or_redirect(request, congreso)
    if guard:
        return guard

    concursos = (
        Concurso.objects.filter(congreso=congreso)
        .select_related("instructor")
        .order_by("title")
    )

    import unicodedata
    def _normalize(text: str | None) -> str:
        if not text:
            return ""
        return "".join(c for c in unicodedata.normalize("NFD", text) if unicodedata.category(c) != "Mn")

    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Concursos"
        headers = [
            _normalize("#"),
            _normalize("ID Concurso"),
            _normalize("Nombre del concurso"),
            _normalize("Modalidad"),
            _normalize("Lugar"),
            _normalize("Descripcion"),
            _normalize("ID Instructor"),
            _normalize("Nombre Instructor"),
            _normalize("Correo Instructor"),
            _normalize("Cupo maximo (individual)"),
            _normalize("Numero de equipos (grupal)"),
            _normalize("Max por equipo (grupal)"),
        ]
        ws.append(headers)
        for idx, x in enumerate(concursos, start=1):
            instructor = x.instructor
            is_grupal = (x.type == "grupal")
            cupo = x.cupo_maximo if not is_grupal else ""
            n_equipos = x.numero_equipos if is_grupal else ""
            max_equipo = x.max_por_equipo if is_grupal else ""
            ws.append([
                idx,
                x.id,
                _normalize(x.title),
                _normalize(x.type),
                _normalize(x.lugar or ""),
                _normalize(x.description or ""),
                instructor.id if instructor else "",
                _normalize((instructor.first_name or instructor.username) if instructor else ""),
                _normalize((instructor.email or instructor.username) if instructor else ""),
                cupo,
                n_equipos,
                max_equipo,
            ])
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value) if cell.value is not None else ""
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=Concursos.xlsx"
        wb.save(response)
        return response
    except ImportError:
        import csv, io
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow([
            _normalize("#"),
            _normalize("ID Concurso"),
            _normalize("Nombre del concurso"),
            _normalize("Modalidad"),
            _normalize("Lugar"),
            _normalize("Descripcion"),
            _normalize("ID Instructor"),
            _normalize("Nombre Instructor"),
            _normalize("Correo Instructor"),
            _normalize("Cupo maximo (individual)"),
            _normalize("Numero de equipos (grupal)"),
            _normalize("Max por equipo (grupal)"),
        ])
        for idx, x in enumerate(concursos, start=1):
            instructor = x.instructor
            is_grupal = (x.type == "grupal")
            cupo = x.cupo_maximo if not is_grupal else ""
            n_equipos = x.numero_equipos if is_grupal else ""
            max_equipo = x.max_por_equipo if is_grupal else ""
            writer.writerow([
                idx,
                x.id,
                _normalize(x.title),
                _normalize(x.type),
                _normalize(x.lugar or ""),
                _normalize((x.description or "").replace("\n"," ").strip()),
                instructor.id if instructor else "",
                _normalize((instructor.first_name or instructor.username) if instructor else ""),
                _normalize((instructor.email or instructor.username) if instructor else ""),
                cupo,
                n_equipos,
                max_equipo,
            ])
        response = HttpResponse(buffer.getvalue(), content_type="text/csv")
        response["Content-Disposition"] = "attachment; filename=Concursos.csv"
        return response

@login_required
def export_taller_excel_view(request, congreso_id: int, taller_id: int):
    """Exporta un único taller en formato Excel (o CSV fallback) con las mismas columnas
    usadas en la exportación global de talleres."""
    from django.http import HttpResponse
    try:
        congreso = Congreso.objects.get(pk=congreso_id)
    except Congreso.DoesNotExist:
        messages.error(request, "El congreso solicitado no existe.")
        return redirect("congresos_eventos")
    guard = _ensure_congreso_access_or_redirect(request, congreso)
    if guard:
        return guard
    try:
        taller = Taller.objects.select_related("instructor").get(pk=taller_id, congreso=congreso)
    except Taller.DoesNotExist:
        messages.error(request, "El taller solicitado no existe en este congreso.")
        return redirect("talleres_list", congreso.id)

    import unicodedata
    def _normalize(text: str | None) -> str:
        if not text:
            return ""
        return "".join(c for c in unicodedata.normalize("NFD", text) if unicodedata.category(c) != "Mn")

    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Taller"
        headers = [
            _normalize("#"),
            _normalize("ID Taller"),
            _normalize("Nombre del taller"),
            _normalize("Lugar"),
            _normalize("Cupo maximo"),
            _normalize("Descripcion"),
            _normalize("ID Instructor"),
            _normalize("Nombre Instructor"),
            _normalize("Correo Instructor"),
        ]
        ws.append(headers)
        inst = taller.instructor
        ws.append([
            1,
            taller.id,
            _normalize(taller.title),
            _normalize(taller.lugar or ""),
            taller.cupo_maximo if taller.cupo_maximo is not None else "",
            _normalize(taller.description or ""),
            inst.id if inst else "",
            _normalize((inst.first_name or inst.username) if inst else ""),
            _normalize((inst.email or inst.username) if inst else ""),
        ])
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value) if cell.value is not None else ""
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        from django.utils.text import slugify as _slug
        response["Content-Disposition"] = f"attachment; filename=Taller_{_slug(taller.title)[:50]}.xlsx"
        wb.save(response)
        return response
    except ImportError:
        import csv, io
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow([
            _normalize("#"),
            _normalize("ID Taller"),
            _normalize("Nombre del taller"),
            _normalize("Lugar"),
            _normalize("Cupo maximo"),
            _normalize("Descripcion"),
            _normalize("ID Instructor"),
            _normalize("Nombre Instructor"),
            _normalize("Correo Instructor"),
        ])
        inst = taller.instructor
        writer.writerow([
            1,
            taller.id,
            _normalize(taller.title),
            _normalize(taller.lugar or ""),
            taller.cupo_maximo if taller.cupo_maximo is not None else "",
            _normalize((taller.description or "").replace("\n"," ").strip()),
            inst.id if inst else "",
            _normalize((inst.first_name or inst.username) if inst else ""),
            _normalize((inst.email or inst.username) if inst else ""),
        ])
        response = HttpResponse(buffer.getvalue(), content_type="text/csv")
        from django.utils.text import slugify as _slug
        response["Content-Disposition"] = f"attachment; filename=Taller_{_slug(taller.title)[:50]}.csv"
        return response

@login_required
def export_conferencia_excel_view(request, congreso_id: int, conferencia_id: int):
    """Exporta una única conferencia en formato Excel (o CSV) con columnas análogas al export global."""
    from django.http import HttpResponse
    try:
        congreso = Congreso.objects.get(pk=congreso_id)
    except Congreso.DoesNotExist:
        messages.error(request, "El congreso solicitado no existe.")
        return redirect("congresos_eventos")
    guard = _ensure_congreso_access_or_redirect(request, congreso)
    if guard:
        return guard
    try:
        conferencia = Conferencia.objects.select_related("instructor").get(pk=conferencia_id, congreso=congreso)
    except Conferencia.DoesNotExist:
        messages.error(request, "La conferencia solicitada no existe en este congreso.")
        return redirect("conferencias_list", congreso.id)

    import unicodedata
    def _normalize(text: str | None) -> str:
        if not text:
            return ""
        return "".join(c for c in unicodedata.normalize("NFD", text) if unicodedata.category(c) != "Mn")

    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Conferencia"
        headers = [
            _normalize("#"),
            _normalize("ID Conferencia"),
            _normalize("Nombre de la conferencia"),
            _normalize("Lugar"),
            _normalize("Cupo maximo"),
            _normalize("Descripcion"),
            _normalize("ID Instructor"),
            _normalize("Nombre Instructor"),
            _normalize("Correo Instructor"),
        ]
        ws.append(headers)
        inst = conferencia.instructor
        ws.append([
            1,
            conferencia.id,
            _normalize(conferencia.title),
            _normalize(conferencia.lugar or ""),
            conferencia.cupo_maximo if conferencia.cupo_maximo is not None else "",
            _normalize(conferencia.description or ""),
            inst.id if inst else "",
            _normalize((inst.first_name or inst.username) if inst else ""),
            _normalize((inst.email or inst.username) if inst else ""),
        ])
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value) if cell.value is not None else ""
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        from django.utils.text import slugify as _slug
        response["Content-Disposition"] = f"attachment; filename=Conferencia_{_slug(conferencia.title)[:50]}.xlsx"
        wb.save(response)
        return response
    except ImportError:
        import csv, io
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow([
            _normalize("#"),
            _normalize("ID Conferencia"),
            _normalize("Nombre de la conferencia"),
            _normalize("Lugar"),
            _normalize("Cupo maximo"),
            _normalize("Descripcion"),
            _normalize("ID Instructor"),
            _normalize("Nombre Instructor"),
            _normalize("Correo Instructor"),
        ])
        inst = conferencia.instructor
        writer.writerow([
            1,
            conferencia.id,
            _normalize(conferencia.title),
            _normalize(conferencia.lugar or ""),
            conferencia.cupo_maximo if conferencia.cupo_maximo is not None else "",
            _normalize((conferencia.description or "").replace("\n"," ").strip()),
            inst.id if inst else "",
            _normalize((inst.first_name or inst.username) if inst else ""),
            _normalize((inst.email or inst.username) if inst else ""),
        ])
        response = HttpResponse(buffer.getvalue(), content_type="text/csv")
        from django.utils.text import slugify as _slug
        response["Content-Disposition"] = f"attachment; filename=Conferencia_{_slug(conferencia.title)[:50]}.csv"
        return response

@login_required
def export_concurso_excel_view(request, congreso_id: int, concurso_id: int):
    """Exporta un único concurso en formato Excel/CSV con modalidad y campos grupales/individuales."""
    from django.http import HttpResponse
    try:
        congreso = Congreso.objects.get(pk=congreso_id)
    except Congreso.DoesNotExist:
        messages.error(request, "El congreso solicitado no existe.")
        return redirect("congresos_eventos")
    guard = _ensure_congreso_access_or_redirect(request, congreso)
    if guard:
        return guard
    try:
        concurso = Concurso.objects.select_related("instructor").get(pk=concurso_id, congreso=congreso)
    except Concurso.DoesNotExist:
        messages.error(request, "El concurso solicitado no existe en este congreso.")
        return redirect("concursos_list", congreso.id)

    import unicodedata
    def _normalize(text: str | None) -> str:
        if not text:
            return ""
        return "".join(c for c in unicodedata.normalize("NFD", text) if unicodedata.category(c) != "Mn")

    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Concurso"
        headers = [
            _normalize("#"),
            _normalize("ID Concurso"),
            _normalize("Nombre del concurso"),
            _normalize("Modalidad"),
            _normalize("Lugar"),
            _normalize("Descripcion"),
            _normalize("ID Instructor"),
            _normalize("Nombre Instructor"),
            _normalize("Correo Instructor"),
            _normalize("Cupo maximo (individual)"),
            _normalize("Numero de equipos (grupal)"),
            _normalize("Max por equipo (grupal)"),
        ]
        ws.append(headers)
        inst = concurso.instructor
        is_grupal = concurso.type == "grupal"
        ws.append([
            1,
            concurso.id,
            _normalize(concurso.title),
            _normalize(concurso.type),
            _normalize(concurso.lugar or ""),
            _normalize(concurso.description or ""),
            inst.id if inst else "",
            _normalize((inst.first_name or inst.username) if inst else ""),
            _normalize((inst.email or inst.username) if inst else ""),
            concurso.cupo_maximo if not is_grupal and concurso.cupo_maximo is not None else "",
            concurso.numero_equipos if is_grupal and concurso.numero_equipos is not None else "",
            concurso.max_por_equipo if is_grupal and concurso.max_por_equipo is not None else "",
        ])
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value) if cell.value is not None else ""
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        from django.utils.text import slugify as _slug
        response["Content-Disposition"] = f"attachment; filename=Concurso_{_slug(concurso.title)[:50]}.xlsx"
        wb.save(response)
        return response
    except ImportError:
        import csv, io
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow([
            _normalize("#"),
            _normalize("ID Concurso"),
            _normalize("Nombre del concurso"),
            _normalize("Modalidad"),
            _normalize("Lugar"),
            _normalize("Descripcion"),
            _normalize("ID Instructor"),
            _normalize("Nombre Instructor"),
            _normalize("Correo Instructor"),
            _normalize("Cupo maximo (individual)"),
            _normalize("Numero de equipos (grupal)"),
            _normalize("Max por equipo (grupal)"),
        ])
        inst = concurso.instructor
        is_grupal = concurso.type == "grupal"
        writer.writerow([
            1,
            concurso.id,
            _normalize(concurso.title),
            _normalize(concurso.type),
            _normalize(concurso.lugar or ""),
            _normalize((concurso.description or "").replace("\n"," ").strip()),
            inst.id if inst else "",
            _normalize((inst.first_name or inst.username) if inst else ""),
            _normalize((inst.email or inst.username) if inst else ""),
            concurso.cupo_maximo if not is_grupal and concurso.cupo_maximo is not None else "",
            concurso.numero_equipos if is_grupal and concurso.numero_equipos is not None else "",
            concurso.max_por_equipo if is_grupal and concurso.max_por_equipo is not None else "",
        ])
        response = HttpResponse(buffer.getvalue(), content_type="text/csv")
        from django.utils.text import slugify as _slug
        response["Content-Disposition"] = f"attachment; filename=Concurso_{_slug(concurso.title)[:50]}.csv"
        return response


@login_required
def export_instructores_excel_view(request, congreso_id: int):
    """Exporta en Excel (o CSV) todas las membresías de instructores del congreso.

    Columnas:
    1) #
    2) ID Usuario
    3) Nombre
    4) Correo
    5) Status membresía
    6..N) Campos extra dinámicos visibles para instructores (registro/perfil; global + congreso)
    Luego columnas de asociaciones:
      - Talleres IDs
      - Talleres Nombres
      - Conferencias IDs
      - Conferencias Nombres
      - Concursos IDs
      - Concursos Nombres
    """
    from django.http import HttpResponse
    try:
        congreso = Congreso.objects.get(pk=congreso_id)
    except Congreso.DoesNotExist:
        messages.error(request, "El congreso solicitado no existe.")
        return redirect("congresos_eventos")
    guard = _ensure_congreso_access_or_redirect(request, congreso)
    if guard:
        return guard

    membresias = list(
        UserCongresoMembership.objects.filter(congreso=congreso, role="instructor")
        .select_related("user")
        .order_by("user__first_name", "user__username")
    )
    user_ids = [m.user_id for m in membresias] or [0]

    # Campos extra (registro + perfil) visibles para instructores
    extra_fields = list(
        ExtraField.objects.filter(
            (Q(congreso=congreso) | Q(congreso__isnull=True)),
            active=True,
            section__in=["registro", "perfil"],
            role_scope__in=["both", "instructor"],
        ).order_by("order", "name")
    )
    field_ids = [f.id for f in extra_fields] or [0]
    value_qs = UserExtraFieldValue.objects.filter(
        (Q(congreso=congreso) | Q(congreso__isnull=True)),
        user_id__in=user_ids,
        field_id__in=field_ids,
    ).values("user_id", "field_id", "value")
    value_map = {(r["user_id"], r["field_id"]): r["value"] for r in value_qs}

    # Asociaciones de recursos impartidos
    talleres_qs = Taller.objects.filter(congreso=congreso, instructor_id__in=user_ids).values("instructor_id", "id", "title")
    conferencias_qs = Conferencia.objects.filter(congreso=congreso, instructor_id__in=user_ids).values("instructor_id", "id", "title")
    concursos_qs = Concurso.objects.filter(congreso=congreso, instructor_id__in=user_ids).values("instructor_id", "id", "title")
    talleres_map: dict[int, list[tuple[int,str]]] = {}
    for r in talleres_qs:
        talleres_map.setdefault(r["instructor_id"], []).append((r["id"], r["title"]))
    conferencias_map: dict[int, list[tuple[int,str]]] = {}
    for r in conferencias_qs:
        conferencias_map.setdefault(r["instructor_id"], []).append((r["id"], r["title"]))
    concursos_map: dict[int, list[tuple[int,str]]] = {}
    for r in concursos_qs:
        concursos_map.setdefault(r["instructor_id"], []).append((r["id"], r["title"]))

    import unicodedata
    def _normalize(text: str | None) -> str:
        if not text:
            return ""
        return "".join(c for c in unicodedata.normalize("NFD", text) if unicodedata.category(c) != "Mn")

    headers = [
        _normalize("#"),
        _normalize("ID Usuario"),
        _normalize("Nombre"),
        _normalize("Correo"),
        _normalize("Status"),
    ] + [_normalize(f.name) for f in extra_fields] + [
        _normalize("Talleres IDs"),
        _normalize("Talleres"),
        _normalize("Conferencias IDs"),
        _normalize("Conferencias"),
        _normalize("Concursos IDs"),
        _normalize("Concursos"),
    ]

    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Instructores"
        ws.append(headers)
        for idx, m in enumerate(membresias, start=1):
            u = m.user
            ordered_values = []
            for f in extra_fields:
                if f.role_scope == "participante":
                    ordered_values.append("")
                else:
                    ordered_values.append(value_map.get((m.user_id, f.id), ""))
            tall = talleres_map.get(u.id, [])
            conf = conferencias_map.get(u.id, [])
            conc = concursos_map.get(u.id, [])
            tall_ids = "; ".join(str(t[0]) for t in tall)
            tall_titles = "; ".join(_normalize(t[1]) for t in tall)
            conf_ids = "; ".join(str(c[0]) for c in conf)
            conf_titles = "; ".join(_normalize(c[1]) for c in conf)
            conc_ids = "; ".join(str(cu[0]) for cu in conc)
            conc_titles = "; ".join(_normalize(cu[1]) for cu in conc)
            ws.append([
                idx,
                u.id,
                _normalize(u.first_name or u.username or u.email),
                _normalize(u.email or u.username),
                _normalize(m.status),
                *[ _normalize(v) for v in ordered_values ],
                tall_ids,
                tall_titles,
                conf_ids,
                conf_titles,
                conc_ids,
                conc_titles,
            ])
        # Auto ancho
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value) if cell.value is not None else ""
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=Instructores.xlsx"
        wb.save(response)
        return response
    except ImportError:
        import csv, io
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(headers)
        for idx, m in enumerate(membresias, start=1):
            u = m.user
            ordered_values = []
            for f in extra_fields:
                if f.role_scope == "participante":
                    ordered_values.append("")
                else:
                    ordered_values.append(value_map.get((m.user_id, f.id), ""))
            tall = talleres_map.get(u.id, [])
            conf = conferencias_map.get(u.id, [])
            conc = concursos_map.get(u.id, [])
            tall_ids = "; ".join(str(t[0]) for t in tall)
            tall_titles = "; ".join(_normalize(t[1]) for t in tall)
            conf_ids = "; ".join(str(c[0]) for c in conf)
            conf_titles = "; ".join(_normalize(c[1]) for c in conf)
            conc_ids = "; ".join(str(cu[0]) for cu in conc)
            conc_titles = "; ".join(_normalize(cu[1]) for cu in conc)
            writer.writerow([
                idx,
                u.id,
                _normalize(u.first_name or u.username or u.email),
                _normalize(u.email or u.username),
                _normalize(m.status),
                *[ _normalize(v) for v in ordered_values ],
                tall_ids,
                tall_titles,
                conf_ids,
                conf_titles,
                conc_ids,
                conc_titles,
            ])
        response = HttpResponse(buffer.getvalue(), content_type="text/csv")
        response["Content-Disposition"] = "attachment; filename=Instructores.csv"
        return response


@login_required
def export_participantes_excel_view(request, congreso_id: int):
    """Exporta en Excel (o CSV) todas las membresías de participantes del congreso.

    Columnas:
    1) #
    2) ID Usuario
    3) Nombre
    4) Correo
    5) Status membresía
    6) Nivel de desempeño (si aplica)
    7..N) Campos extra dinámicos visibles para participantes
    Asociaciones:
      - Talleres inscritos IDs / Nombres
      - Conferencias inscritas IDs / Nombres
      - Concursos inscritos IDs / Nombres
      - Equipos concursos grupales (Nombre equipo (Concurso))
    """
    from django.http import HttpResponse
    try:
        congreso = Congreso.objects.get(pk=congreso_id)
    except Congreso.DoesNotExist:
        messages.error(request, "El congreso solicitado no existe.")
        return redirect("congresos_eventos")
    guard = _ensure_congreso_access_or_redirect(request, congreso)
    if guard:
        return guard

    membresias = list(
        UserCongresoMembership.objects.filter(congreso=congreso, role="participante")
        .select_related("user", "performance_level")
        .order_by("user__first_name", "user__username")
    )
    user_ids = [m.user_id for m in membresias] or [0]

    extra_fields = list(
        ExtraField.objects.filter(
            (Q(congreso=congreso) | Q(congreso__isnull=True)),
            active=True,
            section__in=["registro", "perfil"],
            role_scope__in=["both", "participante"],
        ).order_by("order", "name")
    )
    field_ids = [f.id for f in extra_fields] or [0]
    value_qs = UserExtraFieldValue.objects.filter(
        (Q(congreso=congreso) | Q(congreso__isnull=True)),
        user_id__in=user_ids,
        field_id__in=field_ids,
    ).values("user_id", "field_id", "value")
    value_map = {(r["user_id"], r["field_id"]): r["value"] for r in value_qs}

    # Inscripciones
    talleres_insc = TallerInscripcion.objects.filter(congreso=congreso, user_id__in=user_ids).select_related("taller")
    conferencias_insc = ConferenciaInscripcion.objects.filter(congreso=congreso, user_id__in=user_ids).select_related("conferencia")
    concursos_insc = ConcursoInscripcion.objects.filter(congreso=congreso, user_id__in=user_ids).select_related("concurso")
    tall_map: dict[int, list[tuple[int,str]]] = {}
    for ins in talleres_insc:
        tall_map.setdefault(ins.user_id, []).append((ins.taller_id, ins.taller.title))
    conf_map: dict[int, list[tuple[int,str]]] = {}
    for ins in conferencias_insc:
        conf_map.setdefault(ins.user_id, []).append((ins.conferencia_id, ins.conferencia.title))
    conc_map: dict[int, list[tuple[int,str]]] = {}
    for ins in concursos_insc:
        conc_map.setdefault(ins.user_id, []).append((ins.concurso_id, ins.concurso.title))
    # Equipos para concursos grupales
    equipos_miembros = ConcursoEquipoMiembro.objects.filter(
        equipo__concurso__congreso=congreso, user_id__in=user_ids
    ).select_related("equipo", "equipo__concurso")
    equipos_map: dict[int, list[str]] = {}
    for em in equipos_miembros:
        equipos_map.setdefault(em.user_id, []).append(f"{em.equipo.nombre} ({em.equipo.concurso.title})")

    import unicodedata
    def _normalize(text: str | None) -> str:
        if not text:
            return ""
        return "".join(c for c in unicodedata.normalize("NFD", text) if unicodedata.category(c) != "Mn")

    headers = [
        _normalize("#"),
        _normalize("ID Usuario"),
        _normalize("Nombre"),
        _normalize("Correo"),
        _normalize("Status"),
        _normalize("Nivel de desempeno"),
    ] + [_normalize(f.name) for f in extra_fields] + [
        _normalize("Talleres IDs"),
        _normalize("Talleres"),
        _normalize("Conferencias IDs"),
        _normalize("Conferencias"),
        _normalize("Concursos IDs"),
        _normalize("Concursos"),
        _normalize("Equipos concursos grupales"),
    ]

    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Participantes"
        ws.append(headers)
        for idx, m in enumerate(membresias, start=1):
            u = m.user
            ordered_values = []
            for f in extra_fields:
                if f.role_scope == "instructor":
                    ordered_values.append("")
                else:
                    ordered_values.append(value_map.get((m.user_id, f.id), ""))
            tall = tall_map.get(u.id, [])
            conf = conf_map.get(u.id, [])
            conc = conc_map.get(u.id, [])
            equipos = equipos_map.get(u.id, [])
            ws.append([
                idx,
                u.id,
                _normalize(u.first_name or u.username or u.email),
                _normalize(u.email or u.username),
                _normalize(m.status),
                _normalize(m.performance_level.name if m.performance_level else ""),
                *[ _normalize(v) for v in ordered_values ],
                "; ".join(str(t[0]) for t in tall),
                "; ".join(_normalize(t[1]) for t in tall),
                "; ".join(str(c[0]) for c in conf),
                "; ".join(_normalize(c[1]) for c in conf),
                "; ".join(str(cu[0]) for cu in conc),
                "; ".join(_normalize(cu[1]) for cu in conc),
                "; ".join(_normalize(e) for e in equipos),
            ])
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value) if cell.value is not None else ""
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=Participantes.xlsx"
        wb.save(response)
        return response
    except ImportError:
        import csv, io
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(headers)
        for idx, m in enumerate(membresias, start=1):
            u = m.user
            ordered_values = []
            for f in extra_fields:
                if f.role_scope == "instructor":
                    ordered_values.append("")
                else:
                    ordered_values.append(value_map.get((m.user_id, f.id), ""))
            tall = tall_map.get(u.id, [])
            conf = conf_map.get(u.id, [])
            conc = conc_map.get(u.id, [])
            equipos = equipos_map.get(u.id, [])
            writer.writerow([
                idx,
                u.id,
                _normalize(u.first_name or u.username or u.email),
                _normalize(u.email or u.username),
                _normalize(m.status),
                _normalize(m.performance_level.name if m.performance_level else ""),
                *[ _normalize(v) for v in ordered_values ],
                "; ".join(str(t[0]) for t in tall),
                "; ".join(_normalize(t[1]) for t in tall),
                "; ".join(str(c[0]) for c in conf),
                "; ".join(_normalize(c[1]) for c in conf),
                "; ".join(str(cu[0]) for cu in conc),
                "; ".join(_normalize(cu[1]) for cu in conc),
                "; ".join(_normalize(e) for e in equipos),
            ])
        response = HttpResponse(buffer.getvalue(), content_type="text/csv")
        response["Content-Disposition"] = "attachment; filename=Participantes.csv"
        return response

@login_required
def taller_participantes_print_view(request, congreso_id: int, taller_id: int):
    try:
        congreso = Congreso.objects.get(pk=congreso_id)
    except Congreso.DoesNotExist:
        messages.error(request, "El congreso solicitado no existe.")
        return redirect("congresos_eventos")
    guard = _ensure_congreso_access_or_redirect(request, congreso, allow_instructors=True)
    if guard:
        return guard
    try:
        taller = Taller.objects.get(pk=taller_id, congreso=congreso)
    except Taller.DoesNotExist:
        messages.error(request, "El taller solicitado no existe en este congreso.")
        return redirect("talleres_list", congreso.id)

    inscripciones = (
        TallerInscripcion.objects
        .filter(congreso=congreso, taller=taller)
        .select_related("user", "performance_level")
        .order_by("user__first_name", "user__username")
    )
    from .models import ExtraField, UserExtraFieldValue
    extra_fields = ExtraField.objects.filter(active=True).filter(
        Q(congreso=congreso) | Q(congreso__isnull=True)
    ).filter(role_scope__in=["both", "participante"]).order_by("order", "name")
    user_ids = list(inscripciones.values_list("user_id", flat=True))
    extra_vals = (
        UserExtraFieldValue.objects
        .filter(congreso=congreso, user_id__in=user_ids, field__in=extra_fields)
        .select_related("field")
    )
    extra_values_by_user: dict[int, dict[int, str]] = {}
    for v in extra_vals:
        d = extra_values_by_user.setdefault(v.user_id, {})
        d[v.field_id] = v.value
    # Si es Instructor, exigir que sea dueño del recurso
    if request.user.is_authenticated and request.user.groups.filter(name="Instructores").exists() and not _user_is_congreso_admin(request.user, congreso):
        if getattr(taller, "instructor_id", None) != request.user.id:
            messages.error(request, "No tienes permiso para imprimir este taller.")
            return redirect("ins_talleres")

    ctx = {
        "congreso": congreso,
        "taller": taller,
        "inscripciones": inscripciones,
        "extra_fields": extra_fields,
        "extra_values_by_user": extra_values_by_user,
    }
    return render(request, "taller_accion/taller_participantes_print.html", ctx)


@login_required
def conferencia_participantes_print_view(request, congreso_id: int, conferencia_id: int):
    try:
        congreso = Congreso.objects.get(pk=congreso_id)
    except Congreso.DoesNotExist:
        messages.error(request, "El congreso solicitado no existe.")
        return redirect("congresos_eventos")
    guard = _ensure_congreso_access_or_redirect(request, congreso, allow_instructors=True)
    if guard:
        return guard
    try:
        conferencia = Conferencia.objects.get(pk=conferencia_id, congreso=congreso)
    except Conferencia.DoesNotExist:
        messages.error(request, "La conferencia solicitada no existe en este congreso.")
        return redirect("conferencias_list", congreso.id)

    inscripciones = (
        ConferenciaInscripcion.objects
        .filter(congreso=congreso, conferencia=conferencia)
        .select_related("user", "performance_level")
        .order_by("user__first_name", "user__username")
    )
    from .models import ExtraField, UserExtraFieldValue
    extra_fields = ExtraField.objects.filter(active=True).filter(
        Q(congreso=congreso) | Q(congreso__isnull=True)
    ).filter(role_scope__in=["both", "participante"]).order_by("order", "name")
    user_ids = list(inscripciones.values_list("user_id", flat=True))
    extra_vals = (
        UserExtraFieldValue.objects
        .filter(congreso=congreso, user_id__in=user_ids, field__in=extra_fields)
        .select_related("field")
    )
    extra_values_by_user: dict[int, dict[int, str]] = {}
    for v in extra_vals:
        d = extra_values_by_user.setdefault(v.user_id, {})
        d[v.field_id] = v.value
    # Si es Instructor, exigir que sea dueño del recurso
    if request.user.is_authenticated and request.user.groups.filter(name="Instructores").exists() and not _user_is_congreso_admin(request.user, congreso):
        if getattr(conferencia, "instructor_id", None) != request.user.id:
            messages.error(request, "No tienes permiso para imprimir esta conferencia.")
            return redirect("ins_conferencias")

    ctx = {
        "congreso": congreso,
        "conferencia": conferencia,
        "inscripciones": inscripciones,
        "extra_fields": extra_fields,
        "extra_values_by_user": extra_values_by_user,
    }
    return render(request, "conferencia_accion/conferencia_participantes_print.html", ctx)


@login_required
def concurso_participantes_print_view(request, congreso_id: int, concurso_id: int):
    try:
        congreso = Congreso.objects.get(pk=congreso_id)
    except Congreso.DoesNotExist:
        messages.error(request, "El congreso solicitado no existe.")
        return redirect("congresos_eventos")
    guard = _ensure_congreso_access_or_redirect(request, congreso, allow_instructors=True)
    if guard:
        return guard
    try:
        concurso = Concurso.objects.get(pk=concurso_id, congreso=congreso)
    except Concurso.DoesNotExist:
        messages.error(request, "El concurso solicitado no existe en este congreso.")
        return redirect("concursos_list", congreso.id)

    inscripciones = (
        ConcursoInscripcion.objects
        .filter(congreso=congreso, concurso=concurso)
        .select_related("user", "performance_level")
        .order_by("user__first_name", "user__username")
    )
    from .models import ExtraField, UserExtraFieldValue
    extra_fields = ExtraField.objects.filter(active=True).filter(
        Q(congreso=congreso) | Q(congreso__isnull=True)
    ).filter(role_scope__in=["both", "participante"]).order_by("order", "name")
    user_ids = list(inscripciones.values_list("user_id", flat=True))
    extra_vals = (
        UserExtraFieldValue.objects
        .filter(congreso=congreso, user_id__in=user_ids, field__in=extra_fields)
        .select_related("field")
    )
    extra_values_by_user: dict[int, dict[int, str]] = {}
    for v in extra_vals:
        d = extra_values_by_user.setdefault(v.user_id, {})
        d[v.field_id] = v.value
    # Si es Instructor, exigir que sea dueño del recurso
    if request.user.is_authenticated and request.user.groups.filter(name="Instructores").exists() and not _user_is_congreso_admin(request.user, congreso):
        if getattr(concurso, "instructor_id", None) != request.user.id:
            messages.error(request, "No tienes permiso para imprimir este concurso.")
            return redirect("ins_concursos")

    ctx = {
        "congreso": congreso,
        "concurso": concurso,
        "inscripciones": inscripciones,
        "extra_fields": extra_fields,
        "extra_values_by_user": extra_values_by_user,
    }
    return render(request, "concurso_accion/concurso_participantes_print.html", ctx)

@login_required
def concursos_list_view(request, congreso_id: int):
    try:
        congreso = Congreso.objects.get(pk=congreso_id)
    except Congreso.DoesNotExist:
        messages.error(request, "El congreso solicitado no existe.")
        return redirect("congresos_eventos")
    guard = _ensure_congreso_access_or_redirect(request, congreso)
    if guard:
        return guard
    grupos = request.user.groups.values_list("name", flat=True)
    if "Instructores" in grupos:
        rol = "Instructor"
    elif "Participantes" in grupos:
        rol = "Participante"
    else:
        rol = "Administrador"
    is_admin = (
        (congreso.admin_user_id == request.user.id)
        or request.user.groups.filter(name__in=["Administradores_Congresos", "Administrador"]).exists()
        or request.user.is_superuser
    )

    if request.method == "POST" and is_admin:
        action = request.POST.get("action")
        if action == "delete_concurso":
            cid = request.POST.get("concurso_id")
            if cid:
                try:
                    Concurso.objects.get(id=cid, congreso=congreso).delete()
                    messages.success(request, "Concurso eliminado.", extra_tags="concurso")
                except Concurso.DoesNotExist:
                    messages.error(request, "El concurso indicado no existe.")
            else:
                messages.error(request, "No se indicó el concurso a eliminar.")
            return redirect("concursos_list", congreso.id)
    items = list(Concurso.objects.filter(congreso=congreso).order_by("created_at"))
    return render(request, "concursos_list.html", {"rol": rol, "congreso": congreso, "concursos": items, "is_admin": is_admin, "is_scoped_admin": bool(_get_scoped_congreso_for_group_admin(request.user))})


@login_required
def conferencias_list_view(request, congreso_id: int):
    try:
        congreso = Congreso.objects.get(pk=congreso_id)
    except Congreso.DoesNotExist:
        messages.error(request, "El congreso solicitado no existe.")
        return redirect("congresos_eventos")
    guard = _ensure_congreso_access_or_redirect(request, congreso)
    if guard:
        return guard
    grupos = request.user.groups.values_list("name", flat=True)
    if "Instructores" in grupos:
        rol = "Instructor"
    elif "Participantes" in grupos:
        rol = "Participante"
    else:
        rol = "Administrador"
    is_admin = (
        (congreso.admin_user_id == request.user.id)
        or request.user.groups.filter(name__in=["Administradores_Congresos", "Administrador"]).exists()
        or request.user.is_superuser
    )

    if request.method == "POST" and is_admin:
        action = request.POST.get("action")
        if action == "delete_conferencia":
            fid = request.POST.get("conferencia_id")
            if fid:
                try:
                    Conferencia.objects.get(id=fid, congreso=congreso).delete()
                    messages.success(request, "Conferencia eliminada.")
                except Conferencia.DoesNotExist:
                    messages.error(request, "La conferencia indicada no existe.")
            else:
                messages.error(request, "No se indicó la conferencia a eliminar.")
            return redirect("conferencias_list", congreso.id)
    items = list(Conferencia.objects.filter(congreso=congreso).order_by("created_at"))
    return render(request, "conferencias_list.html", {"rol": rol, "congreso": congreso, "conferencias": items, "is_admin": is_admin, "is_scoped_admin": bool(_get_scoped_congreso_for_group_admin(request.user))})


# -------------------
# ACCIONES: NUEVO/DETALLE (PLACEHOLDER)
# -------------------
@login_required
def taller_nuevo_view(request, congreso_id: int):
    try:
        congreso = Congreso.objects.get(pk=congreso_id)
    except Congreso.DoesNotExist:
        messages.error(request, "El congreso solicitado no existe.")
        return redirect("congresos_eventos")
    guard = _ensure_congreso_access_or_redirect(request, congreso)
    if guard:
        return guard
    grupos = request.user.groups.values_list("name", flat=True)
    if "Instructores" in grupos:
        rol = "Instructor"
    elif "Participantes" in grupos:
        rol = "Participante"
    else:
        rol = "Administrador"
    # Instructores aprobados para este congreso
    instructors = list(
        User.objects.filter(
            congreso_memberships__congreso=congreso,
            congreso_memberships__role="instructor",
            congreso_memberships__status="approved",
        ).distinct().order_by("first_name", "username")
    )

    if request.method == "POST":
        title = (request.POST.get("title") or "").strip()
        lugar = (request.POST.get("lugar") or "").strip()
        cupo = request.POST.get("cupo")
        instructor_id = request.POST.get("instructor_id")
        descripcion = (request.POST.get("descripcion") or "").strip()
        image = request.FILES.get("image")

        has_error = False
        if not title:
            messages.error(request, "El nombre del taller es obligatorio.")
            has_error = True
        elif len(title) > 300:
            messages.error(request, "El nombre del taller es demasiado largo (máximo 300 caracteres).")
            has_error = True
        elif Taller.objects.filter(congreso=congreso, title__iexact=title).exists():
            messages.error(request, "Ya existe un taller con ese nombre en este congreso.")
            has_error = True

        # Validar imagen ~3MB
        if image and isinstance(image, UploadedFile) and image.size > 3 * 1024 * 1024:
            messages.error(request, "La imagen excede el tamaño máximo de 3MB.")
            has_error = True

        # Instructor (opcional, pero si viene debe pertenecer al set)
        instructor = None
        if instructor_id:
            try:
                instructor = next(u for u in instructors if str(u.id) == str(instructor_id))
            except StopIteration:
                messages.error(request, "Instructor inválido para este congreso.")
                has_error = True

        # Cupo
        cupo_val = None
        if cupo:
            try:
                cupo_val = max(0, int(cupo))
            except ValueError:
                messages.error(request, "Cupo debe ser un número entero.")
                has_error = True

        if not has_error:
            t = Taller(
                congreso=congreso,
                title=title,
                lugar=lugar,
                cupo_maximo=cupo_val,
                instructor=instructor,
                description=descripcion,
            )
            if image:
                t.image = image
            t.save()
            messages.success(request, "Taller creado correctamente.")
            return redirect("talleres_list", congreso.id)

    return render(request, "taller_accion/taller.html", {"rol": rol, "congreso": congreso, "instructors": instructors, "is_scoped_admin": bool(_get_scoped_congreso_for_group_admin(request.user))})


@login_required
def taller_editar_view(request, congreso_id: int, taller_id: int):
    try:
        congreso = Congreso.objects.get(pk=congreso_id)
    except Congreso.DoesNotExist:
        messages.error(request, "El congreso solicitado no existe.")
        return redirect("congresos_eventos")
    # Permitir acceso a instructores dueños del recurso, además de admins/superusers
    grupos = request.user.groups.values_list("name", flat=True)
    if "Instructores" in grupos:
        rol = "Instructor"
    elif "Participantes" in grupos:
        rol = "Participante"
    else:
        rol = "Administrador"

    try:
        taller = Taller.objects.get(pk=taller_id, congreso=congreso)
    except Taller.DoesNotExist:
        messages.error(request, "El taller solicitado no existe en este congreso.")
        return redirect("talleres_list", congreso.id)

    is_owner_instructor = (request.user.groups.filter(name__in=["Instructores"]).exists() and taller.instructor_id == request.user.id)
    is_admin_user = _user_is_congreso_admin(request.user, congreso)
    if not (is_owner_instructor or is_admin_user):
        # Para otros roles mantener guardias habituales
        guard = _ensure_congreso_access_or_redirect(request, congreso)
        if guard:
            return guard

    instructors = list(
        User.objects.filter(
            congreso_memberships__congreso=congreso,
            congreso_memberships__role="instructor",
            congreso_memberships__status="approved",
        ).distinct().order_by("first_name", "username")
    )

    if request.method == "POST":
        title = (request.POST.get("title") or "").strip()
        lugar = (request.POST.get("lugar") or "").strip()
        cupo = request.POST.get("cupo")
        instructor_id = request.POST.get("instructor_id")
        descripcion = (request.POST.get("descripcion") or "").strip()
        image = request.FILES.get("image")

        has_error = False
        if not title:
            messages.error(request, "El nombre del taller es obligatorio.")
            has_error = True
        elif Taller.objects.exclude(pk=taller.pk).filter(congreso=congreso, title__iexact=title).exists():
            messages.error(request, "Ya existe un taller con ese nombre en este congreso.")
            has_error = True

        if image and isinstance(image, UploadedFile) and image.size > 3 * 1024 * 1024:
            messages.error(request, "La imagen excede el tamaño máximo de 3MB.")
            has_error = True

        # Si es el dueño (instructor), forzar instructor a sí mismo; si es admin, permitir cambio
        instructor = taller.instructor
        if is_admin_user:
            if instructor_id:
                try:
                    instructor = next(u for u in instructors if str(u.id) == str(instructor_id))
                except StopIteration:
                    messages.error(request, "Instructor inválido para este congreso.")
                    has_error = True
        else:
            instructor = request.user

        cupo_val = None
        if cupo:
            try:
                cupo_val = max(0, int(cupo))
            except ValueError:
                messages.error(request, "Cupo debe ser un número entero.")
                has_error = True

        if not has_error:
            taller.title = title
            taller.lugar = lugar
            taller.cupo_maximo = cupo_val
            taller.instructor = instructor
            taller.description = descripcion
            if image:
                taller.image = image
            taller.save()
            messages.success(request, "Taller actualizado correctamente.")
            return redirect("talleres_list", congreso.id)

    return render(
        request,
        "taller_accion/editar_taller.html",
        {"rol": rol, "congreso": congreso, "instructors": instructors, "taller": taller, "is_scoped_admin": bool(_get_scoped_congreso_for_group_admin(request.user))},
    )


@login_required
def concurso_nuevo_view(request, congreso_id: int):
    try:
        congreso = Congreso.objects.get(pk=congreso_id)
    except Congreso.DoesNotExist:
        messages.error(request, "El congreso solicitado no existe.")
        return redirect("congresos_eventos")
    guard = _ensure_congreso_access_or_redirect(request, congreso)
    if guard:
        return guard
    grupos = request.user.groups.values_list("name", flat=True)
    if "Instructores" in grupos:
        rol = "Instructor"
    elif "Participantes" in grupos:
        rol = "Participante"
    else:
        rol = "Administrador"

    # Instructores aprobados para este congreso
    instructors = list(
        User.objects.filter(
            congreso_memberships__congreso=congreso,
            congreso_memberships__role="instructor",
            congreso_memberships__status="approved",
        ).distinct().order_by("first_name", "username")
    )

    if request.method == "POST":
        tipo = (request.POST.get("tipo") or "individual").strip()
        title = (request.POST.get("title") or "").strip()
        lugar = (request.POST.get("lugar") or "").strip()
        instructor_id = request.POST.get("instructor_id")
        descripcion = (request.POST.get("descripcion") or "").strip()
        image = request.FILES.get("image")
        cupo = request.POST.get("cupo")
        numero_equipos = request.POST.get("numero_equipos")
        max_por_equipo = request.POST.get("max_por_equipo")

        has_error = False
        if tipo not in ("individual", "grupal"):
            tipo = "individual"
        if not title:
            messages.error(request, "El nombre del concurso es obligatorio.")
            has_error = True
        elif Concurso.objects.filter(congreso=congreso, title__iexact=title).exists():
            messages.error(request, "Ya existe un concurso con ese nombre en este congreso.")
            has_error = True

        if image and isinstance(image, UploadedFile) and image.size > 3 * 1024 * 1024:
            messages.error(request, "La imagen excede el tamaño máximo de 3MB.")
            has_error = True

        instructor = None
        if instructor_id:
            try:
                instructor = next(u for u in instructors if str(u.id) == str(instructor_id))
            except StopIteration:
                messages.error(request, "Instructor inválido para este congreso.")
                has_error = True

        cupo_val = None
        num_eq_val = None
        max_eq_val = None
        if tipo == "individual":
            if cupo:
                try:
                    cupo_val = max(0, int(cupo))
                except ValueError:
                    messages.error(request, "Cupo debe ser un número entero.")
                    has_error = True
        else:  # grupal
            if numero_equipos:
                try:
                    num_eq_val = max(0, int(numero_equipos))
                except ValueError:
                    messages.error(request, "Número de equipos debe ser entero.")
                    has_error = True
            if max_por_equipo:
                try:
                    max_eq_val = max(0, int(max_por_equipo))
                except ValueError:
                    messages.error(request, "Número máximo por equipo debe ser entero.")
                    has_error = True

        if not has_error:
            c = Concurso(
                congreso=congreso,
                type=tipo,
                title=title,
                lugar=lugar,
                instructor=instructor,
                description=descripcion,
                cupo_maximo=cupo_val if tipo == "individual" else None,
                numero_equipos=num_eq_val if tipo == "grupal" else None,
                max_por_equipo=max_eq_val if tipo == "grupal" else None,
            )
            if image:
                c.image = image
            c.save()
            messages.success(request, "Concurso creado correctamente.", extra_tags="concurso")
            return redirect("concursos_list", congreso.id)

    return render(request, "concurso_accion/concurso.html", {"rol": rol, "congreso": congreso, "instructors": instructors, "is_scoped_admin": bool(_get_scoped_congreso_for_group_admin(request.user))})


@login_required
def concurso_editar_view(request, congreso_id: int, concurso_id: int):
    try:
        congreso = Congreso.objects.get(pk=congreso_id)
    except Congreso.DoesNotExist:
        messages.error(request, "El congreso solicitado no existe.")
        return redirect("congresos_eventos")
    # Permitir acceso a instructores dueños del recurso, además de admins/superusers
    grupos = request.user.groups.values_list("name", flat=True)
    if "Instructores" in grupos:
        rol = "Instructor"
    elif "Participantes" in grupos:
        rol = "Participante"
    else:
        rol = "Administrador"

    try:
        concurso = Concurso.objects.get(pk=concurso_id, congreso=congreso)
    except Concurso.DoesNotExist:
        messages.error(request, "El concurso solicitado no existe en este congreso.")
        return redirect("concursos_list", congreso.id)

    is_owner_instructor = (request.user.groups.filter(name__in=["Instructores"]).exists() and concurso.instructor_id == request.user.id)
    is_admin_user = _user_is_congreso_admin(request.user, congreso)
    if not (is_owner_instructor or is_admin_user):
        guard = _ensure_congreso_access_or_redirect(request, congreso)
        if guard:
            return guard

    instructors = list(
        User.objects.filter(
            congreso_memberships__congreso=congreso,
            congreso_memberships__role="instructor",
            congreso_memberships__status="approved",
        ).distinct().order_by("first_name", "username")
    )

    if request.method == "POST":
        tipo = (request.POST.get("tipo") or concurso.type or "individual").strip()
        title = (request.POST.get("title") or "").strip()
        lugar = (request.POST.get("lugar") or "").strip()
        instructor_id = request.POST.get("instructor_id")
        descripcion = (request.POST.get("descripcion") or "").strip()
        image = request.FILES.get("image")
        cupo = request.POST.get("cupo")
        numero_equipos = request.POST.get("numero_equipos")
        max_por_equipo = request.POST.get("max_por_equipo")

        has_error = False
        if tipo not in ("individual", "grupal"):
            tipo = "individual"
        if not title:
            messages.error(request, "El nombre del concurso es obligatorio.")
            has_error = True
        elif Concurso.objects.exclude(pk=concurso.pk).filter(congreso=congreso, title__iexact=title).exists():
            messages.error(request, "Ya existe un concurso con ese nombre en este congreso.")
            has_error = True

        if image and isinstance(image, UploadedFile) and image.size > 3 * 1024 * 1024:
            messages.error(request, "La imagen excede el tamaño máximo de 3MB.")
            has_error = True

        # Si es el dueño (instructor), forzar instructor a sí mismo; si es admin, permitir cambio
        instructor = concurso.instructor
        if is_admin_user:
            if instructor_id:
                try:
                    instructor = next(u for u in instructors if str(u.id) == str(instructor_id))
                except StopIteration:
                    messages.error(request, "Instructor inválido para este congreso.")
                    has_error = True
        else:
            instructor = request.user

        cupo_val = None
        num_eq_val = None
        max_eq_val = None
        if tipo == "individual":
            if cupo:
                try:
                    cupo_val = max(0, int(cupo))
                except ValueError:
                    messages.error(request, "Cupo debe ser un número entero.")
                    has_error = True
        else:
            if numero_equipos:
                try:
                    num_eq_val = max(0, int(numero_equipos))
                except ValueError:
                    messages.error(request, "Número de equipos debe ser entero.")
                    has_error = True
            if max_por_equipo:
                try:
                    max_eq_val = max(0, int(max_por_equipo))
                except ValueError:
                    messages.error(request, "Número máximo por equipo debe ser entero.")
                    has_error = True

        if not has_error:
            concurso.type = tipo
            concurso.title = title
            concurso.lugar = lugar
            concurso.instructor = instructor
            concurso.description = descripcion
            concurso.cupo_maximo = cupo_val if tipo == "individual" else None
            concurso.numero_equipos = num_eq_val if tipo == "grupal" else None
            concurso.max_por_equipo = max_eq_val if tipo == "grupal" else None
            if image:
                concurso.image = image
            concurso.save()
            messages.success(request, "Concurso actualizado correctamente.", extra_tags="concurso")
            return redirect("concursos_list", congreso.id)

    return render(
        request,
        "concurso_accion/editar_concurso.html",
        {"rol": rol, "congreso": congreso, "instructors": instructors, "concurso": concurso, "is_scoped_admin": bool(_get_scoped_congreso_for_group_admin(request.user))},
    )


@login_required
def conferencia_nueva_view(request, congreso_id: int):
    try:
        congreso = Congreso.objects.get(pk=congreso_id)
    except Congreso.DoesNotExist:
        messages.error(request, "El congreso solicitado no existe.")
        return redirect("congresos_eventos")
    guard = _ensure_congreso_access_or_redirect(request, congreso)
    if guard:
        return guard
    grupos = request.user.groups.values_list("name", flat=True)
    if "Instructores" in grupos:
        rol = "Instructor"
    elif "Participantes" in grupos:
        rol = "Participante"
    else:
        rol = "Administrador"
    instructors = list(
        User.objects.filter(
            congreso_memberships__congreso=congreso,
            congreso_memberships__role="instructor",
            congreso_memberships__status="approved",
        ).distinct().order_by("first_name", "username")
    )

    if request.method == "POST":
        title = (request.POST.get("title") or "").strip()
        lugar = (request.POST.get("lugar") or "").strip()
        instructor_id = request.POST.get("instructor_id")
        descripcion = (request.POST.get("descripcion") or "").strip()
        cupo = request.POST.get("cupo") or request.POST.get("cupo_maximo") or None
        image = request.FILES.get("image")

        has_error = False
        if not title:
            messages.error(request, "El nombre de la conferencia es obligatorio.")
            has_error = True
        elif Conferencia.objects.filter(congreso=congreso, title__iexact=title).exists():
            messages.error(request, "Ya existe una conferencia con ese nombre en este congreso.")
            has_error = True

        if image and isinstance(image, UploadedFile) and image.size > 3 * 1024 * 1024:
            messages.error(request, "La imagen excede el tamaño máximo de 3MB.")
            has_error = True

        instructor = None
        if instructor_id:
            try:
                instructor = next(u for u in instructors if str(u.id) == str(instructor_id))
            except StopIteration:
                messages.error(request, "Instructor inválido para este congreso.")
                has_error = True

        if not has_error:
            c = Conferencia(
                congreso=congreso,
                title=title,
                lugar=lugar,
                instructor=instructor,
                description=descripcion,
                cupo_maximo=cupo or None,
            )
            if image:
                c.image = image
            c.save()
            messages.success(request, "Conferencia creada correctamente.")
            return redirect("conferencias_list", congreso.id)

    return render(request, "conferencia_accion/conferencia.html", {"rol": rol, "congreso": congreso, "instructors": instructors, "is_scoped_admin": bool(_get_scoped_congreso_for_group_admin(request.user))})


@login_required
def conferencia_editar_view(request, congreso_id: int, conferencia_id: int):
    try:
        congreso = Congreso.objects.get(pk=congreso_id)
    except Congreso.DoesNotExist:
        messages.error(request, "El congreso solicitado no existe.")
        return redirect("congresos_eventos")
    # El acceso de administradores se valida más abajo; permitir dueños instructores
    grupos = request.user.groups.values_list("name", flat=True)
    if "Instructores" in grupos:
        rol = "Instructor"
    elif "Participantes" in grupos:
        rol = "Participante"
    else:
        rol = "Administrador"

    try:
        conferencia = Conferencia.objects.get(pk=conferencia_id, congreso=congreso)
    except Conferencia.DoesNotExist:
        messages.error(request, "La conferencia solicitada no existe en este congreso.")
        return redirect("conferencias_list", congreso.id)

    is_owner_instructor = (request.user.groups.filter(name__in=["Instructores"]).exists() and conferencia.instructor_id == request.user.id)
    is_admin_user = _user_is_congreso_admin(request.user, congreso)
    if not (is_owner_instructor or is_admin_user):
        guard = _ensure_congreso_access_or_redirect(request, congreso)
        if guard:
            return guard

    instructors = list(
        User.objects.filter(
            congreso_memberships__congreso=congreso,
            congreso_memberships__role="instructor",
            congreso_memberships__status="approved",
        ).distinct().order_by("first_name", "username")
    )

    if request.method == "POST":
        title = (request.POST.get("title") or "").strip()
        lugar = (request.POST.get("lugar") or "").strip()
        instructor_id = request.POST.get("instructor_id")
        descripcion = (request.POST.get("descripcion") or "").strip()
        cupo = request.POST.get("cupo") or request.POST.get("cupo_maximo") or None
        image = request.FILES.get("image")

        has_error = False
        if not title:
            messages.error(request, "El nombre de la conferencia es obligatorio.")
            has_error = True
        elif Conferencia.objects.exclude(pk=conferencia.pk).filter(congreso=congreso, title__iexact=title).exists():
            messages.error(request, "Ya existe una conferencia con ese nombre en este congreso.")
            has_error = True

        if image and isinstance(image, UploadedFile) and image.size > 3 * 1024 * 1024:
            messages.error(request, "La imagen excede el tamaño máximo de 3MB.")
            has_error = True

        # Si es dueño instructor, forzar a sí mismo; si es admin, permitir cambio
        instructor = conferencia.instructor
        if is_admin_user:
            if instructor_id:
                try:
                    instructor = next(u for u in instructors if str(u.id) == str(instructor_id))
                except StopIteration:
                    messages.error(request, "Instructor inválido para este congreso.")
                    has_error = True
        else:
            instructor = request.user

        if not has_error:
            conferencia.title = title
            conferencia.lugar = lugar
            conferencia.instructor = instructor
            conferencia.description = descripcion
            conferencia.cupo_maximo = cupo or None
            if image:
                conferencia.image = image
            conferencia.save()
            messages.success(request, "Conferencia actualizada correctamente.")
            return redirect("conferencias_list", congreso.id)

    return render(
        request,
        "conferencia_accion/editar_conferencia.html",
        {"rol": rol, "congreso": congreso, "instructors": instructors, "conferencia": conferencia, "is_scoped_admin": bool(_get_scoped_congreso_for_group_admin(request.user))},
    )


@login_required
def taller_participantes_view(request, congreso_id: int, taller_id: int):
    # Acceso y objetos
    try:
        congreso = Congreso.objects.get(pk=congreso_id)
    except Congreso.DoesNotExist:
        messages.error(request, "El congreso solicitado no existe.")
        return redirect("congresos_eventos")
    guard = _ensure_congreso_access_or_redirect(request, congreso)
    if guard:
        return guard
    try:
        taller = Taller.objects.get(pk=taller_id, congreso=congreso)
    except Taller.DoesNotExist:
        messages.error(request, "El taller solicitado no existe en este congreso.")
        return redirect("talleres_list", congreso.id)

    # Acciones: quitar inscripción
    if request.method == "POST" and request.POST.get("action") == "remove" and request.POST.get("insc_id"):
        insc_id = request.POST.get("insc_id")
        TallerInscripcion.objects.filter(id=insc_id, taller=taller).delete()
        messages.success(request, "Participante removido del taller.")
        return redirect("taller_participantes", congreso.id, taller.id)

    inscripciones = (
        TallerInscripcion.objects
        .filter(congreso=congreso, taller=taller)
        .select_related("user", "performance_level")
        .order_by("user__first_name", "user__username")
    )
    # Extra fields visibles para participantes en este congreso (o globales)
    from .models import ExtraField, UserExtraFieldValue
    extra_fields = ExtraField.objects.filter(active=True).filter(
        Q(congreso=congreso) | Q(congreso__isnull=True)
    ).filter(role_scope__in=["both", "participante"]).order_by("order", "name")
    user_ids = list(inscripciones.values_list("user_id", flat=True))
    extra_vals = (
        UserExtraFieldValue.objects
        .filter(congreso=congreso, user_id__in=user_ids, field__in=extra_fields)
        .select_related("field")
    )
    extra_values_by_user: dict[int, dict[int, str]] = {}
    for v in extra_vals:
        d = extra_values_by_user.setdefault(v.user_id, {})
        d[v.field_id] = v.value
    ctx = {
        "congreso": congreso,
        "taller": taller,
        "inscripciones": inscripciones,
        "extra_fields": extra_fields,
        "extra_values_by_user": extra_values_by_user,
    }
    grupos = request.user.groups.values_list("name", flat=True)
    if "Instructores" in grupos:
        rol = "Instructor"
    elif "Participantes" in grupos:
        rol = "Participante"
    else:
        rol = "Administrador"
    ctx.update({"rol": rol})
    ctx["is_scoped_admin"] = bool(_get_scoped_congreso_for_group_admin(request.user))
    return render(request, "taller_accion/taller_participantes.html", ctx)


@login_required
def concurso_participantes_view(request, congreso_id: int, concurso_id: int):
    try:
        congreso = Congreso.objects.get(pk=congreso_id)
    except Congreso.DoesNotExist:
        messages.error(request, "El congreso solicitado no existe.")
        return redirect("congresos_eventos")
    guard = _ensure_congreso_access_or_redirect(request, congreso)
    if guard:
        return guard
    try:
        concurso = Concurso.objects.get(pk=concurso_id, congreso=congreso)
    except Concurso.DoesNotExist:
        messages.error(request, "El concurso solicitado no existe en este congreso.")
        return redirect("concursos_list", congreso.id)

    if request.method == "POST" and request.POST.get("action") == "remove" and request.POST.get("insc_id"):
        insc_id = request.POST.get("insc_id")
        ConcursoInscripcion.objects.filter(id=insc_id, concurso=concurso).delete()
        messages.success(request, "Participante removido del concurso.", extra_tags="concurso")
        return redirect("concurso_participantes", congreso.id, concurso.id)

    inscripciones = (
        ConcursoInscripcion.objects
        .filter(congreso=congreso, concurso=concurso)
        .select_related("user", "performance_level")
        .order_by("user__first_name", "user__username")
    )
    # Extra fields visibles para participantes en este congreso (o globales)
    from .models import ExtraField, UserExtraFieldValue
    extra_fields = ExtraField.objects.filter(active=True).filter(
        Q(congreso=congreso) | Q(congreso__isnull=True)
    ).filter(role_scope__in=["both", "participante"]).order_by("order", "name")
    user_ids = list(inscripciones.values_list("user_id", flat=True))
    extra_vals = (
        UserExtraFieldValue.objects
        .filter(congreso=congreso, user_id__in=user_ids, field__in=extra_fields)
        .select_related("field")
    )
    extra_values_by_user: dict[int, dict[int, str]] = {}
    for v in extra_vals:
        d = extra_values_by_user.setdefault(v.user_id, {})
        d[v.field_id] = v.value
    ctx = {
        "congreso": congreso,
        "concurso": concurso,
        "inscripciones": inscripciones,
        "extra_fields": extra_fields,
        "extra_values_by_user": extra_values_by_user,
    }
    grupos = request.user.groups.values_list("name", flat=True)
    if "Instructores" in grupos:
        rol = "Instructor"
    elif "Participantes" in grupos:
        rol = "Participante"
    else:
        rol = "Administrador"
    ctx.update({"rol": rol})
    ctx["is_scoped_admin"] = bool(_get_scoped_congreso_for_group_admin(request.user))
    return render(request, "concurso_accion/concurso_participantes.html", ctx)


@login_required
def conferencia_participantes_view(request, congreso_id: int, conferencia_id: int):
    try:
        congreso = Congreso.objects.get(pk=congreso_id)
    except Congreso.DoesNotExist:
        messages.error(request, "El congreso solicitado no existe.")
        return redirect("congresos_eventos")
    guard = _ensure_congreso_access_or_redirect(request, congreso)
    if guard:
        return guard
    try:
        conferencia = Conferencia.objects.get(pk=conferencia_id, congreso=congreso)
    except Conferencia.DoesNotExist:
        messages.error(request, "La conferencia solicitada no existe en este congreso.")
        return redirect("conferencias_list", congreso.id)

    if request.method == "POST" and request.POST.get("action") == "remove" and request.POST.get("insc_id"):
        insc_id = request.POST.get("insc_id")
        ConferenciaInscripcion.objects.filter(id=insc_id, conferencia=conferencia).delete()
        messages.success(request, "Participante removido de la conferencia.")
        return redirect("conferencia_participantes", congreso.id, conferencia.id)

    inscripciones = (
        ConferenciaInscripcion.objects
        .filter(congreso=congreso, conferencia=conferencia)
        .select_related("user", "performance_level")
        .order_by("user__first_name", "user__username")
    )
    # Extra fields visibles para participantes en este congreso (o globales)
    from .models import ExtraField, UserExtraFieldValue
    extra_fields = ExtraField.objects.filter(active=True).filter(
        Q(congreso=congreso) | Q(congreso__isnull=True)
    ).filter(role_scope__in=["both", "participante"]).order_by("order", "name")
    user_ids = list(inscripciones.values_list("user_id", flat=True))
    extra_vals = (
        UserExtraFieldValue.objects
        .filter(congreso=congreso, user_id__in=user_ids, field__in=extra_fields)
        .select_related("field")
    )
    extra_values_by_user: dict[int, dict[int, str]] = {}
    for v in extra_vals:
        d = extra_values_by_user.setdefault(v.user_id, {})
        d[v.field_id] = v.value
    ctx = {
        "congreso": congreso,
        "conferencia": conferencia,
        "inscripciones": inscripciones,
        "extra_fields": extra_fields,
        "extra_values_by_user": extra_values_by_user,
    }
    grupos = request.user.groups.values_list("name", flat=True)
    if "Instructores" in grupos:
        rol = "Instructor"
    elif "Participantes" in grupos:
        rol = "Participante"
    else:
        rol = "Administrador"
    ctx.update({"rol": rol})
    ctx["is_scoped_admin"] = bool(_get_scoped_congreso_for_group_admin(request.user))
    return render(request, "conferencia_accion/conferencia_participantes.html", ctx)


@login_required
def instructores_view(request, congreso_id: int):
    try:
        congreso = Congreso.objects.get(pk=congreso_id)
    except Congreso.DoesNotExist:
        messages.error(request, "El congreso solicitado no existe.")
        return redirect("congresos_eventos")
    guard = _ensure_congreso_access_or_redirect(request, congreso)
    if guard:
        return guard

    # Permisos de administración para cambiar estatus
    is_admin = (
        (congreso.admin_user_id == request.user.id)
        or request.user.groups.filter(name__in=["Administradores_Congresos", "Administrador"]).exists()
        or request.user.is_superuser
    )

    # Acciones POST administrativas (cambiar estatus / eliminar membresía)
    if request.method == "POST" and is_admin:
        action = request.POST.get("action")
        if action == "set_status":
            mem_id = request.POST.get("membership_id")
            new_status = (request.POST.get("status") or "").strip()
            if new_status in {"approved", "pending", "rejected"} and mem_id:
                try:
                    m = UserCongresoMembership.objects.get(id=mem_id, congreso=congreso)
                    m.status = new_status
                    m.decided_at = timezone.now()
                    m.decided_by = request.user
                    m.save(update_fields=["status", "decided_at", "decided_by"])
                    messages.success(request, "Estatus actualizado.")
                except UserCongresoMembership.DoesNotExist:
                    messages.error(request, "La membresía indicada no existe.")
            else:
                messages.error(request, "Parámetros inválidos para cambiar estatus.")
            return redirect("instructores", congreso.id)
        elif action == "activate_all":
            # Activar todos los instructores (los no aprobados)
            qs = UserCongresoMembership.objects.filter(
                congreso=congreso,
                role="instructor",
            ).exclude(status="approved")
            updated = qs.update(status="approved", decided_at=timezone.now(), decided_by_id=request.user.id)
            if updated == 0:
                messages.warning(request, "Todos los instructores ya estaban activados.")
            else:
                messages.success(request, f"Se activaron {updated} instructores.")
            return redirect("instructores", congreso.id)
        elif action == "delete_membership":
            mem_id = request.POST.get("membership_id")
            if mem_id:
                try:
                    with transaction.atomic():
                        membership = UserCongresoMembership.objects.select_related("user").get(id=mem_id, congreso=congreso)
                        user_to_clean = membership.user
                        # Eliminar membresía
                        membership.delete()
                        # Eliminar datos asociados a ESTE congreso
                        TallerInscripcion.objects.filter(congreso=congreso, user=user_to_clean).delete()
                        ConferenciaInscripcion.objects.filter(congreso=congreso, user=user_to_clean).delete()
                        ConcursoInscripcion.objects.filter(congreso=congreso, user=user_to_clean).delete()
                        UserExtraFieldValue.objects.filter(congreso=congreso, user=user_to_clean).delete()
                        # Si el usuario ya no tiene ninguna otra membresía en ningún congreso, eliminar al usuario por completo
                        has_other_memberships = UserCongresoMembership.objects.filter(user=user_to_clean).exists()
                        if not has_other_memberships:
                            user_to_clean.delete()
                            messages.success(request, "Membresía eliminada y usuario borrado definitivamente.")
                        else:
                            messages.success(request, "Membresía eliminada.")
                except UserCongresoMembership.DoesNotExist:
                    messages.error(request, "La membresía indicada no existe.")
            else:
                messages.error(request, "No se indicó la membresía a eliminar.")
            return redirect("instructores", congreso.id)

    # Todas las membresías del congreso separadas por rol
    membresias = (
        UserCongresoMembership.objects
        .filter(congreso=congreso)
        .select_related("user", "performance_level")
        .order_by("user__first_name", "user__username")
    )
    instr = [m for m in membresias if m.role == "instructor"]
    parts = [m for m in membresias if m.role == "participante"]

    # Campos extra visibles para instructores (congreso + globales; registro y perfil)
    extra_fields_instr = list(
        ExtraField.objects.filter(
            (Q(congreso=congreso) | Q(congreso__isnull=True)),
            active=True,
            section__in=["registro", "perfil"],
            role_scope__in=["both", "instructor"],
        ).order_by("order", "name")
    )
    user_ids_instr = [m.user_id for m in instr] or [0]
    field_ids_instr = [f.id for f in extra_fields_instr] or [0]
    values_qs_instr = UserExtraFieldValue.objects.filter(
        (Q(congreso=congreso) | Q(congreso__isnull=True)),
        user_id__in=user_ids_instr,
        field_id__in=field_ids_instr,
    ).values("user_id", "field_id", "value")
    value_map_instr = {(r["user_id"], r["field_id"]): r["value"] for r in values_qs_instr}
    instructores_rows = []
    for m in instr:
        ordered = []
        for f in extra_fields_instr:
            # role_scope ya filtrado pero por consistencia
            if f.role_scope == "participante":
                ordered.append("")
                continue
            ordered.append(value_map_instr.get((m.user_id, f.id), ""))
        instructores_rows.append({
            "membership": m,
            "user": m.user,
            "status": m.status,
            "values_ordered": ordered,
        })

    # Totales para instructores (activados vs no activados)
    total_instr_aprob = sum(1 for m in instr if m.status == "approved")
    total_instr_no_aprob = sum(1 for m in instr if m.status != "approved")

    grupos = request.user.groups.values_list("name", flat=True)
    if "Instructores" in grupos:
        rol = "Instructor"
    elif "Participantes" in grupos:
        rol = "Participante"
    else:
        rol = "Administrador"
    ctx = {
        "rol": rol,
        "congreso": congreso,
        "instructores": instr,
        "extra_fields_instructores": extra_fields_instr,
        "instructores_rows": instructores_rows,
        "participantes": parts,
        "is_admin": is_admin,
        "total_instructores_aprob": total_instr_aprob,
        "total_instructores_no_aprob": total_instr_no_aprob,
    }
    ctx["is_scoped_admin"] = bool(_get_scoped_congreso_for_group_admin(request.user))
    return render(request, "instructores.html", ctx)


@login_required
def participantes_view(request, congreso_id: int):
    try:
        congreso = Congreso.objects.get(pk=congreso_id)
    except Congreso.DoesNotExist:
        messages.error(request, "El congreso solicitado no existe.")
        return redirect("congresos_eventos")
    guard = _ensure_congreso_access_or_redirect(request, congreso)
    if guard:
        return guard

    # Permisos de administración para cambiar estatus
    is_admin = (
        (congreso.admin_user_id == request.user.id)
        or request.user.groups.filter(name__in=["Administradores_Congresos", "Administrador"]).exists()
        or request.user.is_superuser
    )

    if request.method == "POST" and is_admin:
        action = request.POST.get("action")
        if action == "set_status":
            mem_id = request.POST.get("membership_id")
            new_status = (request.POST.get("status") or "").strip()
            if new_status in {"approved", "pending", "rejected"} and mem_id:
                try:
                    m = UserCongresoMembership.objects.get(id=mem_id, congreso=congreso)
                    m.status = new_status
                    m.decided_at = timezone.now()
                    m.decided_by = request.user
                    m.save(update_fields=["status", "decided_at", "decided_by"])
                    messages.success(request, "Estatus actualizado.")
                except UserCongresoMembership.DoesNotExist:
                    messages.error(request, "La membresía indicada no existe.")
            else:
                messages.error(request, "Parámetros inválidos para cambiar estatus.")
            return redirect("participantes", congreso.id)
        elif action == "activate_all":
            # Activar todos los participantes (los no aprobados)
            qs = UserCongresoMembership.objects.filter(
                congreso=congreso,
                role="participante",
            ).exclude(status="approved")
            updated = qs.update(status="approved", decided_at=timezone.now(), decided_by_id=request.user.id)
            if updated == 0:
                messages.warning(request, "Todos los participantes ya estaban activados.")
            else:
                messages.success(request, f"Se activaron {updated} participantes.")
            return redirect("participantes", congreso.id)
        elif action == "delete_membership":
            mem_id = request.POST.get("membership_id")
            if mem_id:
                try:
                    with transaction.atomic():
                        membership = UserCongresoMembership.objects.select_related("user").get(id=mem_id, congreso=congreso)
                        user_to_clean = membership.user
                        # Eliminar membresía
                        membership.delete()
                        # Eliminar datos asociados a ESTE congreso
                        TallerInscripcion.objects.filter(congreso=congreso, user=user_to_clean).delete()
                        ConferenciaInscripcion.objects.filter(congreso=congreso, user=user_to_clean).delete()
                        ConcursoInscripcion.objects.filter(congreso=congreso, user=user_to_clean).delete()
                        UserExtraFieldValue.objects.filter(congreso=congreso, user=user_to_clean).delete()
                        # Si el usuario ya no tiene ninguna otra membresía en ningún congreso, eliminar al usuario por completo
                        has_other_memberships = UserCongresoMembership.objects.filter(user=user_to_clean).exists()
                        if not has_other_memberships:
                            user_to_clean.delete()
                            messages.success(request, "Membresía eliminada y usuario borrado definitivamente.")
                        else:
                            messages.success(request, "Membresía eliminada.")
                except UserCongresoMembership.DoesNotExist:
                    messages.error(request, "La membresía indicada no existe.")
            else:
                messages.error(request, "No se indicó la membresía a eliminar.")
            return redirect("participantes", congreso.id)

    membresias = (
        UserCongresoMembership.objects
        .filter(congreso=congreso)
        .select_related("user", "performance_level")
        .order_by("user__first_name", "user__username")
    )
    instr = [m for m in membresias if m.role == "instructor"]
    parts = [m for m in membresias if m.role == "participante"]

    # Campos extra visibles para participantes (congreso + globales; registro y perfil)
    extra_fields_part = list(
        ExtraField.objects.filter(
            (Q(congreso=congreso) | Q(congreso__isnull=True)),
            active=True,
            section__in=["registro", "perfil"],
            role_scope__in=["both", "participante"],
        ).order_by("order", "name")
    )
    user_ids_part = [m.user_id for m in parts] or [0]
    field_ids_part = [f.id for f in extra_fields_part] or [0]
    values_qs_part = UserExtraFieldValue.objects.filter(
        (Q(congreso=congreso) | Q(congreso__isnull=True)),
        user_id__in=user_ids_part,
        field_id__in=field_ids_part,
    ).values("user_id", "field_id", "value")
    value_map_part = {(r["user_id"], r["field_id"]): r["value"] for r in values_qs_part}
    participantes_rows = []
    for m in parts:
        ordered = []
        for f in extra_fields_part:
            # role_scope ya filtrado pero por consistencia
            if f.role_scope == "instructor":
                ordered.append("")
                continue
            ordered.append(value_map_part.get((m.user_id, f.id), ""))
        participantes_rows.append({
            "membership": m,
            "user": m.user,
            "status": m.status,
            "performance_level": m.performance_level,
            "values_ordered": ordered,
        })

    total_parts_aprob = sum(1 for m in parts if m.status == "approved")
    total_parts_no_aprob = sum(1 for m in parts if m.status != "approved")

    grupos = request.user.groups.values_list("name", flat=True)
    if "Instructores" in grupos:
        rol = "Instructor"
    elif "Participantes" in grupos:
        rol = "Participante"
    else:
        rol = "Administrador"
    ctx = {
        "rol": rol,
        "congreso": congreso,
        "instructores": instr,
        "participantes": parts,
        "extra_fields_participantes": extra_fields_part,
        "participantes_rows": participantes_rows,
        "total_participantes_aprob": total_parts_aprob,
        "total_participantes_no_aprob": total_parts_no_aprob,
        "is_admin": is_admin,
    }
    ctx["is_scoped_admin"] = bool(_get_scoped_congreso_for_group_admin(request.user))
    return render(request, "participantes.html", ctx)


@login_required
def logo_view(request, congreso_id: int):
    try:
        congreso = Congreso.objects.get(pk=congreso_id)
    except Congreso.DoesNotExist:
        messages.error(request, "El congreso solicitado no existe.")
        return redirect("congresos_eventos")
    guard = _ensure_congreso_access_or_redirect(request, congreso)
    if guard:
        return guard

    if request.method == "POST" and request.FILES.get("logo"):
        img = request.FILES.get("logo")
        # Validar un tamaño razonable (opcional, 5MB)
        if getattr(img, "size", 0) > 5 * 1024 * 1024:
            messages.error(request, "El logo supera el tamaño máximo de 5MB.")
        else:
            congreso.logo = img
            congreso.save()
            messages.success(request, "Logo del congreso actualizado.")
            return redirect("logo", congreso.id)

    grupos = request.user.groups.values_list("name", flat=True)
    if "Instructores" in grupos:
        rol = "Instructor"
    elif "Participantes" in grupos:
        rol = "Participante"
    else:
        rol = "Administrador"
    return render(request, "logo.html", {"rol": rol, "congreso": congreso, "is_scoped_admin": bool(_get_scoped_congreso_for_group_admin(request.user))})


@login_required
def avisos_view(request, congreso_id: int):
    try:
        congreso = Congreso.objects.get(pk=congreso_id)
    except Congreso.DoesNotExist:
        messages.error(request, "El congreso solicitado no existe.")
        return redirect("congresos_eventos")
    guard = _ensure_congreso_access_or_redirect(request, congreso)
    if guard:
        return guard

    from .models import Aviso

    # Si viene ?edit=<id>, precargar el aviso a editar (solo si pertenece al congreso)
    aviso_to_edit = None
    edit_id = request.GET.get("edit")
    if edit_id:
        try:
            aviso_to_edit = Aviso.objects.get(id=edit_id, congreso=congreso)
        except Aviso.DoesNotExist:
            aviso_to_edit = None

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "create":
            content = (request.POST.get("content") or "").strip()
            if content:
                Aviso.objects.create(congreso=congreso, content=content)
                messages.success(request, "Aviso publicado.")
            else:
                messages.error(request, "Escribe el contenido del aviso.")
            return redirect("avisos", congreso.id)
        elif action == "update":
            aviso_id = request.POST.get("aviso_id")
            content = (request.POST.get("content") or "").strip()
            if not aviso_id:
                messages.error(request, "No se ha indicado el aviso a actualizar.")
                return redirect("avisos", congreso.id)
            try:
                aviso = Aviso.objects.get(id=aviso_id, congreso=congreso)
            except Aviso.DoesNotExist:
                messages.error(request, "El aviso no existe o no pertenece a este congreso.")
                return redirect("avisos", congreso.id)

            if not content:
                messages.error(request, "Escribe el contenido del aviso.")
                # Mantenerse en modo edición del mismo aviso
                return HttpResponseRedirect(reverse("avisos", args=[congreso.id]) + f"?edit={aviso_id}")

            aviso.content = content
            aviso.save()
            messages.success(request, "Aviso actualizado.")
            return redirect("avisos", congreso.id)
        elif action == "delete" and request.POST.get("aviso_id"):
            Aviso.objects.filter(id=request.POST.get("aviso_id"), congreso=congreso).delete()
            messages.success(request, "Aviso eliminado.")
            return redirect("avisos", congreso.id)

    avisos = list(congreso.avisos.all())
    grupos = request.user.groups.values_list("name", flat=True)
    if "Instructores" in grupos:
        rol = "Instructor"
    elif "Participantes" in grupos:
        rol = "Participante"
    else:
        rol = "Administrador"
    return render(request, "avisos.html", {"rol": rol, "congreso": congreso, "avisos": avisos, "aviso_to_edit": aviso_to_edit, "is_scoped_admin": bool(_get_scoped_congreso_for_group_admin(request.user))})


# -------------------
# RECUPERAR CONTRASEÑA (solicitud por correo)
# -------------------
def recuperar_contrasena_view(request):
    """Solicita el correo y, si existe, envía un código y un enlace de restablecimiento.

    - Si el correo es válido y pertenece a un usuario: muestra mensaje de confirmación grande
      y envía un correo con un código de 4 dígitos y un enlace a una página (placeholder).
    - Si el correo no existe o es inválido: despliega error y deja el formulario visible.
    """
    # Mostrar pista en UI si estamos usando el backend de consola
    using_console = getattr(settings, "EMAIL_BACKEND", "").endswith("console.EmailBackend")
    # Mantener el logo del congreso seleccionado (si llega ?c=ID)
    congreso_id = request.GET.get("c") or request.POST.get("c")
    selected_congreso = None
    if congreso_id:
        try:
            selected_congreso = Congreso.objects.get(pk=congreso_id)
        except Congreso.DoesNotExist:
            selected_congreso = None

    context = {"email_sent": False, "using_console_backend": using_console, "selected_congreso": selected_congreso}

    if request.method == "POST":
        email = request.POST.get("email", "").strip()

        # Validación básica de email (formato)
        email_regex = r"^[^@\s]+@[^@\s]+\.[^@\s]+$"
        if not re.match(email_regex, email):
            messages.error(request, "Ingresa un correo válido.")
            return render(request, "password_reset.html", context)

        # Verificar si el correo existe en usuarios
        try:
            user = User.objects.get(email__iexact=email)
        except User.DoesNotExist:
            messages.error(request, "No existe una cuenta con ese correo.")
            return render(request, "password_reset.html", context)

        # Generar código y enlace
        codigo = random.randint(1000, 9999)
        enlace = request.build_absolute_uri(
            reverse("code_password") + f"?code={codigo}"
            + (f"&c={selected_congreso.id}" if selected_congreso else "")
        )

        # Persistir código con expiración de 30 minutos y anular anteriores
        PasswordResetCode.objects.filter(user=user, is_used=False).delete()
        expires_at = timezone.now() + timedelta(minutes=30)
        PasswordResetCode.objects.create(user=user, code=str(codigo), expires_at=expires_at)

        # Preparar correo (texto + HTML con enlace clicable)
        subject = "Restablecer contraseña (válido 30 min)"
        body = (
            "Estimad@ usuario\n\n"
            "Recientemente se envió una solicitud para restablecer una contraseña para su cuenta.\n"
            "Si esto fue un error, simplemente ignore este correo electrónico y no pasará nada.\n"
            "Para restablecer su contraseña tenga en cuenta el siguiente código:\n\n"
            f"{codigo}\n\n"
            "Este código caduca en 30 minutos.\n\n"
            f"Para restablecer su contraseña da clic aquí: {enlace}\n\n"
            "Saludos"
        )

        html_body = f"""
            <p>Estimad@ usuario</p>
            <p>
              Recientemente se envió una solicitud para restablecer una contraseña para su cuenta.<br/>
              Si esto fue un error, simplemente ignore este correo electrónico y no pasará nada.
            </p>
            <p>Para restablecer su contraseña tenga en cuenta el siguiente código:</p>
            <p style=\"font-size:22px;font-weight:700;letter-spacing:3px;margin:8px 0;\">{codigo}</p>
            <p style=\"margin:8px 0;\"><strong>Este código caduca en 30 minutos.</strong></p>
            <p>
              <a href=\"{enlace}\" style=\"display:inline-block;padding:10px 16px;background:#0d6efd;color:#fff;text-decoration:none;border-radius:6px;\">
                Para restablecer su contraseña da clic aquí
              </a>
            </p>
            <p>Saludos</p>
    """

        from_email = getattr(settings, "DEFAULT_FROM_EMAIL", None) or getattr(
            settings, "EMAIL_HOST_USER", "no-reply@localhost"
        )

        try:
            send_mail(subject, body, from_email, [email], fail_silently=False, html_message=html_body)
            context["email_sent"] = True
        except Exception:
            # Si hay problema con la configuración de correo, avisar pero no fallar.
            messages.warning(
                request,
                "No se pudo enviar el correo en este momento. Verifica la configuración de correo del sistema."
            )

        return render(request, "password_reset.html", context)

    return render(request, "password_reset.html", context)


def code_password_view(request):
    """Verifica el código y permite cambiar la contraseña con la política definida."""
    prefill_code = request.GET.get("code", "").strip()
    context = {"prefill_code": prefill_code}
    # Mantener logo de congreso seleccionado si llega ?c=ID
    congreso_id = request.GET.get("c") or request.POST.get("c")
    selected_congreso = None
    if congreso_id:
        try:
            selected_congreso = Congreso.objects.get(pk=congreso_id)
        except Congreso.DoesNotExist:
            selected_congreso = None
    context["selected_congreso"] = selected_congreso

    if request.method == "POST":
        code = request.POST.get("code", "").strip()
        new = request.POST.get("new_password", "")
        confirm = request.POST.get("confirm_password", "")

        # Validaciones de código
        if not code:
            messages.error(request, "Ingresa el código enviado a tu correo.")
            return render(request, "code_password.html", context)

        now = timezone.now()
        prc = (
            PasswordResetCode.objects.filter(code=code, is_used=False, expires_at__gt=now)
            .order_by("-created_at")
            .first()
        )
        if not prc:
            messages.error(request, "El código es inválido o ha expirado.")
            return render(request, "code_password.html", context)

        # Validaciones de contraseña
        if new != confirm:
            messages.error(request, "La nueva contraseña y su confirmación no coinciden.")
            return render(request, "code_password.html", context)
        if len(new) < 8 or len(new) > 16:
            messages.error(request, "La nueva contraseña debe tener entre 8 y 16 caracteres.")
            return render(request, "code_password.html", context)
        specials = set(".!%&$}")
        if not any(c.isupper() for c in new) or not any(c.islower() for c in new) or not any(c in specials for c in new):
            messages.error(
                request,
                "La nueva contraseña debe contener al menos una letra mayúscula, una letra minúscula y un símbolo especial (.!%&$}).",
            )
            return render(request, "code_password.html", context)

        # Cambiar contraseña del usuario dueño del código
        user = prc.user
        try:
            user.set_password(new)
            user.save()
            # Marcar código como usado e invalidar otros pendientes
            prc.is_used = True
            prc.save(update_fields=["is_used"])
            PasswordResetCode.objects.filter(user=user, is_used=False).exclude(pk=prc.pk).delete()
            messages.success(request, "Tu contraseña fue restablecida. Ahora puedes iniciar sesión.")
            return redirect("login")
        except Exception:
            messages.error(request, "No se pudo restablecer la contraseña en este momento.")

    return render(request, "code_password.html", context)